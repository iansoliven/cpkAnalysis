#!/usr/bin/env python3
"""Generate histogram charts grouped by Test Name and Lot.

The script scans every worksheet whose name starts with "Measurements" inside the
combined workbook produced by mergeXlsx.py. For each distinct Test Name and Lot it
builds a histogram of the Measurement column, scales the X axis so the documented
limits remain visible, and embeds the rendered chart image into a Charts worksheet.
Charts are arranged with one Test Name per row and one Lot per column, each sized
approximately 10" x 5" (width x height).
"""

from __future__ import annotations

import argparse
import math
from collections import OrderedDict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")  # Always render off-screen
import matplotlib.pyplot as plt  # type: ignore
import numpy as np  # type: ignore
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

MEASUREMENTS_PREFIX = "Measurements"
INCH_TO_PIXELS = 96  # Excel assumes ~96 DPI for embedded images
CHART_WIDTH_IN = 10
CHART_HEIGHT_IN = 5
CHART_WIDTH_PX = int(CHART_WIDTH_IN * INCH_TO_PIXELS)
CHART_HEIGHT_PX = int(CHART_HEIGHT_IN * INCH_TO_PIXELS)
CHARTS_SHEET_NAME = "Charts"
ROW_STRIDE = 27  # heuristic rows to leave between chart anchors
COL_STRIDE = 15  # heuristic columns to leave between chart anchors
TEST_LABEL_COLUMN = 1
FIRST_CHART_COLUMN = 2
HEADER_ROW = 3
DEFAULT_MAX_LOTS = 0  # 0 means include every lot discovered


@dataclass
class LotData:
    lot: str
    measurements: List[float]
    low_limits: List[float]
    high_limits: List[float]

    def representative_limits(self) -> Tuple[Optional[float], Optional[float]]:
        return _first_finite(self.low_limits), _first_finite(self.high_limits)


@dataclass
class TestData:
    name: str
    number: Optional[str]
    unit: Optional[str]
    lots: "OrderedDict[str, LotData]"

    @property
    def title(self) -> str:
        if self.number and self.number.strip():
            return f"{self.name} (Test {self.number})"
        return self.name


def _first_finite(values: Iterable[float]) -> Optional[float]:
    for value in values:
        if math.isfinite(value):
            return value
    return None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add histogram charts for each Test Name/Lot pairing in the combined workbook."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=Path("combined_workbook.xlsx"),
        help="Path to the workbook that already contains Measurements sheets (default: combined_workbook.xlsx).",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Optional path to save the augmented workbook. Defaults to overwriting the input.",
    )
    parser.add_argument(
        "--max-lots",
        type=int,
        default=DEFAULT_MAX_LOTS,
        help="Maximum number of Lot columns to render (0 means all lots).",
    )
    return parser.parse_args()


def load_measurement_tests(workbook_path: Path) -> Dict[Tuple[str, Optional[str]], TestData]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        measurement_sheet_names = [name for name in wb.sheetnames if name.startswith(MEASUREMENTS_PREFIX)]
        if not measurement_sheet_names:
            raise ValueError("No Measurements sheets were found in the workbook.")

        tests: Dict[Tuple[str, Optional[str]], TestData] = OrderedDict()
        for sheet_name in measurement_sheet_names:
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            header_map = {str(value).strip(): idx for idx, value in enumerate(header) if value is not None}
            required_columns = ["Test Name", "Measurement", "Lot"]
            for column in required_columns:
                if column not in header_map:
                    raise ValueError(f"Column '{column}' was not found in sheet '{sheet_name}'.")

            idx_test_name = header_map["Test Name"]
            idx_measurement = header_map["Measurement"]
            idx_lot = header_map["Lot"]
            idx_test_number = header_map.get("Test Number")
            idx_unit = header_map.get("Test Unit")
            idx_low_limit = header_map.get("Low Limit")
            idx_high_limit = header_map.get("High Limit")

            for row in rows:
                test_name = _safe_string(row, idx_test_name)
                if not test_name:
                    continue
                measurement_raw = row[idx_measurement] if idx_measurement < len(row) else None
                measurement = _coerce_float(measurement_raw)
                if measurement is None:
                    continue

                lot_label = _safe_string(row, idx_lot) or "Unknown Lot"
                test_number = _safe_string(row, idx_test_number)
                unit_text = _safe_string(row, idx_unit)

                key = (test_name, test_number)
                entry = tests.get(key)
                if entry is None:
                    entry = TestData(
                        name=test_name,
                        number=test_number,
                        unit=unit_text,
                        lots=OrderedDict(),
                    )
                    tests[key] = entry
                elif not entry.unit and unit_text:
                    entry.unit = unit_text

                lot_entry = entry.lots.get(lot_label)
                if lot_entry is None:
                    lot_entry = LotData(lot=lot_label, measurements=[], low_limits=[], high_limits=[])
                    entry.lots[lot_label] = lot_entry

                lot_entry.measurements.append(measurement)
                if idx_low_limit is not None and idx_low_limit < len(row):
                    low_value = _coerce_float(row[idx_low_limit])
                    if low_value is not None:
                        lot_entry.low_limits.append(low_value)
                if idx_high_limit is not None and idx_high_limit < len(row):
                    high_value = _coerce_float(row[idx_high_limit])
                    if high_value is not None:
                        lot_entry.high_limits.append(high_value)
        if not tests:
            raise ValueError("No measurement rows were found across the Measurements sheets.")
        return tests
    finally:
        wb.close()


def _coerce_float(value: Optional[object]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        value_float = float(value)
        return value_float if math.isfinite(value_float) else None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def _safe_string(row: Tuple[object, ...], index: Optional[int]) -> Optional[str]:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    return str(value).strip()


def determine_axis_limits(values: List[float], declared_low: Optional[float], declared_high: Optional[float]) -> Tuple[float, float]:
    if not values:
        raise ValueError("determine_axis_limits called with empty values")
    min_value = min(values)
    max_value = max(values)
    low = declared_low if declared_low is not None else min_value
    high = declared_high if declared_high is not None else max_value
    low = min(low, min_value)
    high = max(high, max_value)
    if math.isclose(low, high):
        span = max(abs(low) * 0.1, 1.0)
        low -= span
        high += span
    return low, high


def choose_bin_count(values: List[float], low: float, high: float) -> int:
    if not values:
        return 10
    span = high - low
    if span <= 0:
        span = max(abs(low) * 0.1, 1.0)
        high = low + span
    n = len(values)
    bins = max(10, min(60, int(round(math.sqrt(n) * 1.5))))
    return bins


def _next_charts_sheet_name(existing_names: Iterable[str]) -> str:
    if CHARTS_SHEET_NAME not in existing_names:
        return CHARTS_SHEET_NAME
    index = 1
    while True:
        candidate = f"{CHARTS_SHEET_NAME}_{index}"
        if candidate not in existing_names:
            return candidate
        index += 1


def build_chart_image(test: TestData, lot_data: LotData) -> Optional[BytesIO]:
    if not lot_data.measurements:
        return None
    low_declared, high_declared = lot_data.representative_limits()
    axis_low_raw, axis_high_raw = determine_axis_limits(lot_data.measurements, low_declared, high_declared)
    bin_count = choose_bin_count(lot_data.measurements, axis_low_raw, axis_high_raw)
    bin_count = max(1, bin_count)

    span = axis_high_raw - axis_low_raw
    bin_width = span / bin_count if bin_count else 0.0
    padding = max(bin_width / 2.0, abs(span) * 0.05, 1e-6)
    if not math.isfinite(padding) or padding <= 0:
        padding = 0.5
    axis_low = axis_low_raw - padding
    axis_high = axis_high_raw + padding
    if not math.isfinite(axis_low):
        axis_low = axis_low_raw if math.isfinite(axis_low_raw) else -0.5
    if not math.isfinite(axis_high):
        axis_high = axis_high_raw if math.isfinite(axis_high_raw) else 0.5
    if axis_low >= axis_high:
        axis_low -= 0.5
        axis_high += 0.5

    bin_edges = np.linspace(axis_low, axis_high, bin_count + 1)

    fig, ax = plt.subplots(figsize=(CHART_WIDTH_IN, CHART_HEIGHT_IN))
    ax.hist(lot_data.measurements, bins=bin_edges, color="#4F81BD", edgecolor="#1F497D", alpha=0.8)
    ax.set_title(f"{lot_data.lot} – {test.title}")
    axis_label = "Measurement"
    if test.unit:
        axis_label += f" ({test.unit})"
    ax.set_xlabel(axis_label)
    ax.set_ylabel("Count")
    ax.set_xlim(axis_low, axis_high)

    legend_entries: List[str] = []
    if low_declared is not None:
        ax.axvline(low_declared, color="#C0504D", linestyle="--", linewidth=2)
        legend_entries.append("Low Limit")
    if high_declared is not None:
        ax.axvline(high_declared, color="#9BBB59", linestyle="--", linewidth=2)
        legend_entries.append("High Limit")
    if legend_entries:
        ax.legend(legend_entries, loc="upper right")

    ax.grid(axis="y", linestyle=":", linewidth=0.6, alpha=0.7)
    fig.tight_layout()

    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=150)
    plt.close(fig)
    buffer.seek(0)
    return buffer


def add_charts_to_workbook(
    workbook_path: Path,
    output_path: Path,
    tests: Dict[Tuple[str, Optional[str]], TestData],
    max_lots: int,
) -> None:
    wb = load_workbook(workbook_path)
    sheet_name = _next_charts_sheet_name(wb.sheetnames)
    charts_ws = wb.create_sheet(sheet_name)
    charts_ws["A1"] = f"Charts generated by addCharts.py ({sheet_name})"
    charts_ws.freeze_panes = "B4"

    sorted_tests = [tests[key] for key in tests]

    lot_order: "OrderedDict[str, None]" = OrderedDict()
    for test in sorted_tests:
        for lot_name in test.lots:
            lot_order.setdefault(lot_name, None)
    lot_names = list(lot_order.keys())
    if max_lots > 0:
        lot_names = lot_names[:max_lots]

    # Header row that labels each lot column
    for col_idx, lot_name in enumerate(lot_names):
        header_column = FIRST_CHART_COLUMN + col_idx * COL_STRIDE
        charts_ws.cell(row=HEADER_ROW, column=header_column, value=lot_name)

    for test_idx, test in enumerate(sorted_tests):
        anchor_row = HEADER_ROW + 1 + test_idx * ROW_STRIDE
        charts_ws.cell(row=anchor_row, column=TEST_LABEL_COLUMN, value=test.title)

        for col_idx, lot_name in enumerate(lot_names):
            lot_data = test.lots.get(lot_name)
            if lot_data is None or not lot_data.measurements:
                continue
            image_stream = build_chart_image(test, lot_data)
            if image_stream is None:
                continue
            img = XLImage(image_stream)
            img.width = CHART_WIDTH_PX
            img.height = CHART_HEIGHT_PX

            anchor_col = FIRST_CHART_COLUMN + col_idx * COL_STRIDE
            anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"
            charts_ws.add_image(img, anchor_cell)

    wb.save(output_path)
    wb.close()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    tests = load_measurement_tests(workbook_path)
    output_path = (args.output or workbook_path).expanduser().resolve()
    add_charts_to_workbook(workbook_path, output_path, tests, max_lots=max(args.max_lots, 0))
    print(
        f"Added Charts sheet with {sum(len(t.lots) for t in tests.values())} histogram(s) across "
        f"{len(tests)} test(s) to {output_path}"
    )


if __name__ == "__main__":
    main()

