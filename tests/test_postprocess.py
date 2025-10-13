from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import pandas as pd
import pytest
from openpyxl import load_workbook, Workbook

from cpkanalysis import postprocess
from cpkanalysis.models import AnalysisInputs, OutlierOptions
from cpkanalysis.postprocess import actions
from cpkanalysis.postprocess.io_adapters import CliIO
from cpkanalysis.postprocess import sheet_utils
from cpkanalysis.workbook_builder import SUMMARY_COLUMNS, TEST_LIMIT_COLUMNS, MEAS_COLUMNS


def _build_test_workbook(
    tmp_path: Path,
    *,
    test_name: str = "Voltage",
    test_number: int | str = "100",
) -> tuple[Path, Path, AnalysisInputs]:
    workbook_path = tmp_path / "cpk_test.xlsx"
    metadata_path = workbook_path.with_suffix(".json")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(SUMMARY_COLUMNS)
    summary_values = {
        "File": "lot1",
        "Test Name": test_name,
        "Test Number": test_number,
        "Unit": "V",
        "MEAN": 1.0,
        "STDEV": 0.1,
        "LL_2CPK": 0.7,
        "UL_2CPK": 1.3,
        "LL_3IQR": 0.6,
        "UL_3IQR": 1.4,
        "CPK": 1.5,
    }
    summary_row = [summary_values.get(column, 0) for column in SUMMARY_COLUMNS]
    summary_ws.append(summary_row)

    # Measurements sheet
    measurements_ws = wb.create_sheet("Measurements")
    measurement_headers = [target for _, target in MEAS_COLUMNS]
    measurements_ws.append(measurement_headers)
    measurement_data = [
        ("lot1", "device-1", test_name, test_number, 1.05, "V", 1),
        ("lot1", "device-2", test_name, test_number, 0.95, "V", 2),
        ("lot1", "device-3", test_name, test_number, 1.10, "V", 3),
    ]
    for row in measurement_data:
        measurements_ws.append(row)

    # Test List and Limits sheet
    limits_ws = wb.create_sheet("Test List and Limits")
    limit_headers = [target for _, target in TEST_LIMIT_COLUMNS]
    limits_ws.append(limit_headers)
    limit_row_values = {
        "Test name": test_name,
        "Test number": test_number,
        "STDF Lower Limit": 0.7,
        "STDF Upper Limit": 1.3,
        "Spec Lower Limit": 0.65,
        "Spec Upper Limit": 1.35,
        "User What-If Lower Limit": 0.6,
        "User What-If Upper Limit": 1.4,
    }
    limits_row = [limit_row_values.get(source, "") for source, _ in TEST_LIMIT_COLUMNS]
    limits_ws.append(limits_row)

    # Template sheet with expected columns
    template_ws = wb.create_sheet("Template")
    template_headers = [
        "TEST NAME",
        "TEST NUM",
        "LL_ATE",
        "UL_ATE",
        "Spec Lower",
        "Spec Upper",
        "What-if Lower",
        "What-if Upper",
        "LL_PROP",
        "UL_PROP",
        "CPK_PROP",
        "%YLD LOSS_PROP",
    ]
    template_ws.append(["Cpk Report"])
    template_ws.append(template_headers)
    template_ws.append(
        [
            test_name,
            test_number,
            0.7,
            1.3,
            0.65,
            1.35,
            0.6,
            1.4,
            "",
            "",
            "",
            "",
        ]
    )

    wb.save(workbook_path)

    metadata_payload = {
        "template_sheet": "Template",
        "output": str(workbook_path),
        "summary_counts": {"rows": 1, "tests": 1},
    }
    metadata_path.write_text(json.dumps(metadata_payload, indent=2), encoding="utf-8")

    analysis_inputs = AnalysisInputs(
        sources=[],
        output=workbook_path,
        template=None,
        template_sheet="Template",
        outliers=OutlierOptions(),
        generate_histogram=True,
        generate_cdf=True,
        generate_time_series=True,
        plugins=[],
    )
    return workbook_path, metadata_path, analysis_inputs


def _template_row(workbook_path: Path) -> tuple[int, dict[str, int], any]:
    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    header_row, header_map = sheet_utils.build_header_map(template_ws)
    return header_row, header_map, template_ws


def _axis_bounds(workbook_path: Path) -> tuple[float, float]:
    wb = load_workbook(workbook_path)
    if "_PlotAxisRanges" not in wb.sheetnames:
        raise AssertionError("Axis ranges sheet not created.")
    ws = wb["_PlotAxisRanges"]
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    axis_min = ws.cell(row=2, column=headers["Axis Min"]).value
    axis_max = ws.cell(row=2, column=headers["Axis Max"]).value
    return float(axis_min), float(axis_max)


def test_update_stdf_limits(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=[])
    descriptor_key = "lot1|Voltage|100"

    # Clear existing template values to ensure the action repopulates them.
    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    header_row, header_map = sheet_utils.build_header_map(template_ws)
    ll_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    template_ws.cell(row=header_row + 1, column=ll_col, value=None)
    template_ws.cell(row=header_row + 1, column=ul_col, value=None)
    wb.save(workbook_path)

    result = actions.update_stdf_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key},
    )
    context.mark_dirty()
    audit_entry = dict(result.get("audit", {}))
    audit_entry.setdefault("action", "update_stdf_limits")
    context.add_audit_entry(audit_entry)
    context.save()

    header_row, header_map, template_ws = _template_row(workbook_path)
    ll_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    ll_value = template_ws.cell(row=header_row + 1, column=ll_col).value
    ul_value = template_ws.cell(row=header_row + 1, column=ul_col).value

    assert ll_value == pytest.approx(0.7, rel=1e-4)
    assert ul_value == pytest.approx(1.3, rel=1e-4)

    # Ensure metadata audit log recorded the action
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    runs = metadata.get("post_processing", {}).get("runs", [])
    assert runs
    assert runs[-1]["action"] == "update_stdf_limits"

    wb = load_workbook(workbook_path)
    assert any(name.startswith("Histogram") for name in wb.sheetnames)


def test_apply_spec_limits(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=[])
    descriptor_key = "lot1|Voltage|100"

    # Invoke action with a target CPK to recompute spec/what-if values
    result = actions.apply_spec_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 1.2},
    )
    context.mark_dirty()
    audit_entry = dict(result.get("audit", {}))
    audit_entry.setdefault("action", "apply_spec_limits")
    context.add_audit_entry(audit_entry)
    context.save()

    header_row, header_map, template_ws = _template_row(workbook_path)
    spec_lower_col = header_map[sheet_utils.normalize_header("Spec Lower")]
    spec_upper_col = header_map[sheet_utils.normalize_header("Spec Upper")]
    spec_lower = template_ws.cell(row=header_row + 1, column=spec_lower_col).value
    spec_upper = template_ws.cell(row=header_row + 1, column=spec_upper_col).value
    expected_spec_lower = 1.0 - (1.2 * 3 * 0.1)
    expected_spec_upper = 1.0 + (1.2 * 3 * 0.1)
    assert spec_lower == pytest.approx(expected_spec_lower, rel=1e-4)
    assert spec_upper == pytest.approx(expected_spec_upper, rel=1e-4)

    wb = load_workbook(workbook_path)
    assert any(name.startswith("CDF") for name in wb.sheetnames)


def test_calculate_proposed_limits(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=[])
    descriptor_key = "lot1|Voltage|100"

    result = actions.calculate_proposed_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 1.1},
    )
    context.mark_dirty()
    audit_entry = dict(result.get("audit", {}))
    audit_entry.setdefault("action", "calculate_proposed_limits")
    context.add_audit_entry(audit_entry)
    context.save()

    header_row, header_map, template_ws = _template_row(workbook_path)
    ll_prop_col = header_map[sheet_utils.normalize_header("LL_PROP")]
    ul_prop_col = header_map[sheet_utils.normalize_header("UL_PROP")]
    cpk_prop_col = header_map[sheet_utils.normalize_header("CPK_PROP")]
    ll_prop = template_ws.cell(row=header_row + 1, column=ll_prop_col).value
    ul_prop = template_ws.cell(row=header_row + 1, column=ul_prop_col).value
    cpk_prop = template_ws.cell(row=header_row + 1, column=cpk_prop_col).value

    expected_ll = 1.0 - (1.1 * 3 * 0.1)
    expected_ul = 1.0 + (1.1 * 3 * 0.1)
    assert pytest.approx(ll_prop, rel=1e-4) == expected_ll
    assert pytest.approx(ul_prop, rel=1e-4) == expected_ul
    assert pytest.approx(cpk_prop, rel=1e-4) == 1.1

    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    runs = metadata.get("post_processing", {}).get("runs", [])
    assert runs[-1]["action"] == "calculate_proposed_limits"


def test_reject_invalid_cpk_param(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=["1.4"])
    descriptor_key = "lot1|Voltage|100"

    result = actions.calculate_proposed_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 0},
    )

    assert result["audit"]["parameters"]["target_cpk"] == pytest.approx(1.4)


def test_update_limits_with_zero_test_number(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path, test_number=0)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=[])
    descriptor_key = "lot1|Voltage|0"

    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    header_row, header_map = sheet_utils.build_header_map(template_ws)
    ll_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    template_ws.cell(row=header_row + 1, column=ll_col, value=None)
    template_ws.cell(row=header_row + 1, column=ul_col, value=None)
    wb.save(workbook_path)

    result = actions.update_stdf_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key},
    )
    context.mark_dirty()
    audit_entry = dict(result.get("audit", {}))
    audit_entry.setdefault("action", "update_stdf_limits_zero")
    context.add_audit_entry(audit_entry)
    context.save()

    header_row, header_map, template_ws = _template_row(workbook_path)
    ll_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    ll_value = template_ws.cell(row=header_row + 1, column=ll_col).value
    ul_value = template_ws.cell(row=header_row + 1, column=ul_col).value

    assert ll_value == pytest.approx(0.7, rel=1e-4)
    assert ul_value == pytest.approx(1.3, rel=1e-4)


def test_apply_spec_limits_extends_axis(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=[])
    descriptor_key = "lot1|Voltage|100"

    actions.apply_spec_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 5},
    )
    context.save()

    header_row, header_map, template_ws = _template_row(workbook_path)
    spec_lower_col = header_map[sheet_utils.normalize_header("Spec Lower")]
    spec_upper_col = header_map[sheet_utils.normalize_header("Spec Upper")]
    spec_lower = float(template_ws.cell(row=header_row + 1, column=spec_lower_col).value)
    spec_upper = float(template_ws.cell(row=header_row + 1, column=spec_upper_col).value)
    axis_min, axis_max = _axis_bounds(workbook_path)

    assert axis_min <= spec_lower
    assert axis_max >= spec_upper


def test_calculate_proposed_limits_extends_axis(tmp_path):
    workbook_path, metadata_path, analysis_inputs = _build_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )
    io = CliIO(scripted_choices=[])
    descriptor_key = "lot1|Voltage|100"

    actions.calculate_proposed_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 5},
    )
    context.save()

    header_row, header_map, template_ws = _template_row(workbook_path)
    ll_prop_col = header_map[sheet_utils.normalize_header("LL_PROP")]
    ul_prop_col = header_map[sheet_utils.normalize_header("UL_PROP")]
    proposed_lower = float(template_ws.cell(row=header_row + 1, column=ll_prop_col).value)
    proposed_upper = float(template_ws.cell(row=header_row + 1, column=ul_prop_col).value)
    axis_min, axis_max = _axis_bounds(workbook_path)

    assert axis_min <= proposed_lower
    assert axis_max >= proposed_upper
