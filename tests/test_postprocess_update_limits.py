from __future__ import annotations

from pathlib import Path
import sys

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cpkanalysis.postprocess import actions, sheet_utils
from cpkanalysis.postprocess.context import PostProcessContext
from cpkanalysis.models import AnalysisInputs


class DummyIO:
    def __init__(self) -> None:
        self.infos: list[str] = []
        self.warnings: list[str] = []

    def print(self, *args, **kwargs) -> None:  # pragma: no cover - not used in test
        return None

    def info(self, message: str) -> None:
        self.infos.append(message)

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    def prompt_choice(self, prompt: str, options, show_options: bool = True) -> int:  # pragma: no cover
        return 0

    def prompt(self, prompt: str, default: str | None = None) -> str:  # pragma: no cover
        return default or ""

    def confirm(self, prompt: str, default: bool = True) -> bool:  # pragma: no cover
        return default


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["File", "Test Name", "Test Number", "Unit", "MEAN", "STDEV", "LL_2CPK", "UL_2CPK", "LL_3IQR", "UL_3IQR"])
    summary.append(["lot1", "TestA", "1", "V", 1.5, 0.4, 0.8, 1.2, 0.7, 1.3])

    limits = wb.create_sheet("Test List and Limits")
    limits.append(["Test Name", "Test Number", "STDF Lower Limit", "STDF Upper Limit"])
    limits.append(["TestA", "1", None, None])

    template = wb.create_sheet("Template")
    template.append(["Test Name", "Test Number", "LL_ATE", "UL_ATE"])
    template.append(["TestA", "1", None, None])

    wb.save(path)


def test_update_stdf_limits_writes_limits(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    metadata_path = tmp_path / "report.json"

    inputs = AnalysisInputs(
        sources=[],
        output=tmp_path / "analysis.xlsx",
        template_sheet="Template",
        generate_histogram=False,
        generate_cdf=False,
        generate_time_series=False,
    )

    context = PostProcessContext(inputs, workbook_path, metadata_path, metadata={})
    io = DummyIO()
    refreshed: dict[str, list] = {}
    monkeypatch.setattr(actions.charts, "refresh_tests", lambda ctx, tests, include_spec=True, include_proposed=False: refreshed.setdefault("tests", list(tests)))

    result = actions.update_stdf_limits(context, io, {"scope": "all"})

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    ll_column = headers[sheet_utils.normalize_header("LL_ATE")]
    ul_column = headers[sheet_utils.normalize_header("UL_ATE")]
    row_idx = header_row + 1

    assert template_ws.cell(row=row_idx, column=ll_column).value == pytest.approx(0.8)
    assert template_ws.cell(row=row_idx, column=ul_column).value == pytest.approx(1.2)

    limits_ws = context.workbook()["Test List and Limits"]
    limits_header_row, limits_headers = sheet_utils.build_header_map(limits_ws)
    stdf_ll_col = limits_headers[sheet_utils.normalize_header("STDF Lower Limit")]
    stdf_ul_col = limits_headers[sheet_utils.normalize_header("STDF Upper Limit")]
    data_row = limits_header_row + 1

    assert limits_ws.cell(row=data_row, column=stdf_ll_col).value == pytest.approx(0.8)
    assert limits_ws.cell(row=data_row, column=stdf_ul_col).value == pytest.approx(1.2)

    assert refreshed["tests"][0].test_name == "TestA"
    assert result["summary"] == "Updated STDF limits for 1 test(s)."
    assert result["mark_dirty"] is True
    assert result["audit"]["tests"] == ["lot1|TestA|1"]
    assert context._limits_df is None
