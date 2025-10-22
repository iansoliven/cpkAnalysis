from __future__ import annotations

import cpkanalysis.workbook_builder as workbook_builder
import pandas as pd
import pytest
from openpyxl import Workbook
from cpkanalysis.workbook_builder import (
    COL_STRIDE,
    _create_plot_sheets,
    _format_site_label,
    _normalise_site_value,
    _populate_cpk_report,
    _site_block_columns,
    _write_yield_pareto_sheet,
)


def test_site_block_columns_stride() -> None:
    label0, image0 = _site_block_columns(0)
    label1, image1 = _site_block_columns(1)
    label5, image5 = _site_block_columns(5)
    assert label0 == 1
    assert image0 == 2
    assert label1 - label0 == COL_STRIDE
    assert image1 - image0 == COL_STRIDE
    assert label5 - label0 == COL_STRIDE * 5
    assert image5 - image0 == COL_STRIDE * 5


def test_normalise_and_format_site_values() -> None:
    assert _normalise_site_value(None) is None
    assert _normalise_site_value(float("nan")) is None
    assert _normalise_site_value(5.0) == 5
    assert _normalise_site_value(5.5) == 5.5
    assert _format_site_label(None) == "Unknown"
    assert _format_site_label(float("nan")) == "Unknown"
    assert _format_site_label(3) == "3"
    assert _format_site_label("siteA") == "siteA"


def test_cpk_report_includes_site_column() -> None:
    wb = Workbook()
    summary = pd.DataFrame(
        [
            {
                "File": "lotA.stdf",
                "Test Name": "VDD",
                "Test Number": "1",
                "Unit": "V",
                "COUNT": 2,
            }
        ]
    )
    site_summary = pd.DataFrame(
        [
            {
                "File": "lotA.stdf",
                "Site": 1,
                "Test Name": "VDD",
                "Test Number": "1",
                "Unit": "V",
            }
        ]
    )
    test_limits = pd.DataFrame(
        [
            {"test_name": "VDD", "test_number": "1", "unit": "V"},
        ]
    )
    _populate_cpk_report(
        wb,
        summary,
        test_limits,
        plot_links={},
        site_summary=site_summary,
        site_limit_sources=None,
        site_plot_links={("lotA.stdf", 1, "VDD", "1"): "#'Histogram_lotA.stdf'!B2"},
        include_site_rows=True,
    )
    sheet = wb["CPK Report"]
    header = [cell.value for cell in sheet[1]]
    assert "Site" in header
    site_col_index = header.index("Site") + 1
    column_values = next(
        sheet.iter_cols(min_col=site_col_index, max_col=site_col_index, min_row=2, values_only=True)
    )
    assert any(str(value) == "1" for value in column_values if value is not None)


def test_yield_pareto_sheet_includes_site_sections() -> None:
    wb = Workbook()
    yield_summary = pd.DataFrame(
        [
            {"file": "lotA.stdf", "devices_total": 2, "devices_pass": 1, "devices_fail": 1, "yield_percent": 0.5},
        ]
    )
    site_yield = pd.DataFrame(
        [
            {"file": "lotA.stdf", "site": 1, "devices_total": 1, "devices_pass": 1, "devices_fail": 0, "yield_percent": 1.0},
            {"file": "lotA.stdf", "site": 2, "devices_total": 1, "devices_pass": 0, "devices_fail": 1, "yield_percent": 0.0},
        ]
    )
    site_pareto = pd.DataFrame(
        [
            {"file": "lotA.stdf", "site": 2, "test_name": "VDD", "test_number": "1", "devices_fail": 1, "fail_rate_percent": 1.0, "cumulative_percent": 1.0, "lower_limit": 0.5, "upper_limit": 1.5}
        ]
    )
    _write_yield_pareto_sheet(
        wb,
        yield_summary,
        pd.DataFrame(),
        site_yield_summary=site_yield,
        site_pareto_summary=site_pareto,
    )
    sheet = wb["Yield and Pareto"]
    labels = [
        value
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    ]
    assert any("Site 1" in label for label in labels)
    assert any("Site 2" in label for label in labels)


def test_site_plots_share_axis_ranges(monkeypatch: pytest.MonkeyPatch) -> None:
    wb = Workbook()
    measurements = pd.DataFrame(
        [
            {"file": "lotA.stdf", "test_name": "VDD", "test_number": "1", "value": 0.5, "site": 1, "device_sequence": 1},
            {"file": "lotA.stdf", "test_name": "VDD", "test_number": "1", "value": 0.7, "site": 1, "device_sequence": 2},
            {"file": "lotA.stdf", "test_name": "VDD", "test_number": "1", "value": 1.2, "site": 2, "device_sequence": 3},
            {"file": "lotA.stdf", "test_name": "VDD", "test_number": "1", "value": 1.4, "site": 2, "device_sequence": 4},
        ]
    )
    summary = pd.DataFrame(
        [
            {
                "File": "lotA.stdf",
                "Test Name": "VDD",
                "Test Number": "1",
                "CPK": 1.0,
            }
        ]
    )
    site_summary = pd.DataFrame(
        [
            {"File": "lotA.stdf", "Site": 1, "Test Name": "VDD", "Test Number": "1", "CPK": 1.5},
            {"File": "lotA.stdf", "Site": 2, "Test Name": "VDD", "Test Number": "1", "CPK": 0.8},
        ]
    )

    captured_tasks: list[dict[str, object]] = []

    def fake_render(
        task: dict[str, object],
        include_histogram: bool,
        include_cdf: bool,
        include_time_series: bool,
        include_rug: bool,
    ) -> dict[str, None]:
        captured_tasks.append(task.copy())
        return {"histogram": None, "cdf": None, "time_series": None}

    monkeypatch.setattr(workbook_builder, "_render_plot_images", fake_render)

    _create_plot_sheets(
        wb,
        measurements=measurements,
        test_limits=pd.DataFrame(),
        summary=summary,
        include_histogram=True,
        include_cdf=True,
        include_time_series=True,
        timings=None,
        include_rug=False,
        max_render_processes=1,
        site_summary=site_summary,
        site_enabled=True,
    )

    aggregate_task = next(task for task in captured_tasks if task.get("block_index") == 0)
    site_tasks = [task for task in captured_tasks if (task.get("block_index") or 0) > 0]
    assert site_tasks, "Expected per-site plot tasks to be generated."

    for site_task in site_tasks:
        assert site_task["x_range"] == aggregate_task["x_range"]
        assert site_task["axis_min"] == aggregate_task["axis_min"]
        assert site_task["axis_max"] == aggregate_task["axis_max"]
        assert site_task["site_data_min"] is not None
        assert site_task["site_data_max"] is not None
