from __future__ import annotations

import json
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pandas as pd
import types
from openpyxl import Workbook

from cpkanalysis import postprocess
from cpkanalysis.models import AnalysisInputs, OutlierOptions
from cpkanalysis.postprocess import charts, actions
from cpkanalysis.workbook_builder import (
    AXIS_META_SHEET,
    MEAS_COLUMNS,
    SUMMARY_COLUMNS,
    TEST_LIMIT_COLUMNS,
    _safe_sheet_name,
)


def _build_two_test_workbook(tmp_path: Path) -> tuple[Path, Path, AnalysisInputs]:
    workbook_path = tmp_path / "two_test_workbook.xlsx"
    metadata_path = workbook_path.with_suffix(".json")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(SUMMARY_COLUMNS)
    summary_rows = [
        {
            "File": "lot1",
            "Test Name": "TestA",
            "Test Number": "1",
            "Unit": "V",
            "MEAN": 1.0,
            "STDEV": 0.1,
            "LL_2CPK": 0.7,
            "UL_2CPK": 1.3,
            "LL_3IQR": 0.6,
            "UL_3IQR": 1.4,
            "CPK": 1.5,
        },
        {
            "File": "lot1",
            "Test Name": "TestB",
            "Test Number": "2",
            "Unit": "V",
            "MEAN": 2.0,
            "STDEV": 0.2,
            "LL_2CPK": 1.4,
            "UL_2CPK": 2.6,
            "LL_3IQR": 1.3,
            "UL_3IQR": 2.7,
            "CPK": 1.5,
        },
    ]
    for row in summary_rows:
        summary_ws.append([row.get(column, "") for column in SUMMARY_COLUMNS])

    measurements_ws = wb.create_sheet("Measurements")
    measurement_headers = [target for _, target in MEAS_COLUMNS]
    measurements_ws.append(measurement_headers)
    measurement_data = [
        ("lot1", "device-1", "TestA", "1", 1.05, "V", 1),
        ("lot1", "device-2", "TestA", "1", 0.95, "V", 2),
        ("lot1", "device-3", "TestA", "1", 1.10, "V", 3),
        ("lot1", "device-1", "TestB", "2", 2.05, "V", 1),
        ("lot1", "device-2", "TestB", "2", 1.95, "V", 2),
        ("lot1", "device-3", "TestB", "2", 2.08, "V", 3),
    ]
    for row in measurement_data:
        measurements_ws.append(row)

    limits_ws = wb.create_sheet("Test List and Limits")
    limit_headers = [target for _, target in TEST_LIMIT_COLUMNS]
    limits_ws.append(limit_headers)
    limit_rows = [
        {
            "Test name": "TestA",
            "Test number": "1",
            "STDF Lower Limit": 0.7,
            "STDF Upper Limit": 1.3,
            "Spec Lower Limit": 0.65,
            "Spec Upper Limit": 1.35,
            "User What-If Lower Limit": 0.6,
            "User What-If Upper Limit": 1.4,
        },
        {
            "Test name": "TestB",
            "Test number": "2",
            "STDF Lower Limit": 1.4,
            "STDF Upper Limit": 2.6,
            "Spec Lower Limit": 1.35,
            "Spec Upper Limit": 2.65,
            "User What-If Lower Limit": 1.3,
            "User What-If Upper Limit": 2.7,
        },
    ]
    for row in limit_rows:
        limits_ws.append([row.get(source, "") for source, _ in TEST_LIMIT_COLUMNS])

    template_ws = wb.create_sheet("Template")
    template_headers = [
        "TEST NAME",
        "TEST NUM",
        "LL_ATE",
        "UL_ATE",
        "Spec Lower",
        "Spec Upper",
        "What-if Lower",
        "What-if Upper",
        "LL_PROP",
        "UL_PROP",
        "CPK_PROP",
        "%YLD LOSS_PROP",
    ]
    template_ws.append(["Cpk Report"])
    template_ws.append(template_headers)
    template_ws.append(
        [
            "TestA",
            "1",
            0.7,
            1.3,
            0.65,
            1.35,
            0.6,
            1.4,
            "",
            "",
            "",
            "",
        ]
    )
    template_ws.append(
        [
            "TestB",
            "2",
            1.4,
            2.6,
            1.35,
            2.65,
            1.3,
            2.7,
            "",
            "",
            "",
            "",
        ]
    )

    wb.save(workbook_path)

    metadata_path.write_text(
        json.dumps(
            {
                "template_sheet": "Template",
                "output": str(workbook_path),
                "summary_counts": {"rows": 2, "tests": 2},
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    analysis_inputs = AnalysisInputs(
        sources=[],
        output=workbook_path,
        template=None,
        template_sheet="Template",
        outliers=OutlierOptions(),
        generate_histogram=True,
        generate_cdf=True,
        generate_time_series=True,
        plugins=[],
    )
    return workbook_path, metadata_path, analysis_inputs


def _read_axis_sheet(workbook) -> dict[tuple[str, str, str], tuple[float | None, ...]]:
    if AXIS_META_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[AXIS_META_SHEET]
    data: dict[tuple[str, str, str], tuple[float | None, ...]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue
        key = (row[0], row[1], row[2])
        data[key] = tuple(row[3:])
    return data


def test_refresh_tests_partial_update_preserves_other_charts(tmp_path: Path) -> None:
    workbook_path, metadata_path, analysis_inputs = _build_two_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )

    # Initial full refresh to populate charts and metadata
    charts.refresh_tests(context, [])
    workbook = context.workbook()
    hist_sheet_name = _safe_sheet_name("Histogram_lot1")
    assert hist_sheet_name in workbook.sheetnames
    histogram_sheet = workbook[hist_sheet_name]
    assert len(histogram_sheet._images) == 2
    previous_anchor = histogram_sheet._images[1].anchor

    axis_before = _read_axis_sheet(workbook)
    assert ("lot1", "TestA", "1") in axis_before
    assert ("lot1", "TestB", "2") in axis_before

    descriptor = actions.TestDescriptor(
        file="lot1",
        test_name="TestA",
        test_number="1",
        unit="V",
        mean=1.0,
        stdev=0.1,
        cpk=1.5,
    )

    charts.refresh_tests(context, [descriptor], include_spec=True, include_proposed=True)
    workbook = context.workbook()
    histogram_sheet = workbook[hist_sheet_name]
    assert len(histogram_sheet._images) == 2
    assert histogram_sheet._images[1].anchor == previous_anchor

    chart_state = context.metadata.get("post_processing_state", {}).get("chart_positions", {})
    hist_positions = chart_state.get("Histogram", {}).get("lot1", {})
    assert hist_positions["lot1|TestA|1"] == 0
    assert hist_positions["lot1|TestB|2"] == 1

    axis_after = _read_axis_sheet(workbook)
    assert axis_after.keys() == axis_before.keys()
    assert axis_after[("lot1", "TestB", "2")] == axis_before[("lot1", "TestB", "2")]


def test_refresh_tests_subset_on_virgin_metadata_triggers_full_refresh(monkeypatch, tmp_path: Path) -> None:
    workbook_path, metadata_path, analysis_inputs = _build_two_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )

    # Initial full refresh to build the chart sheets, then drop metadata to emulate
    # a workbook produced by the pipeline where chart positions were never recorded.
    charts.refresh_tests(context, [])
    context.metadata.pop("post_processing_state", None)

    descriptor = actions.TestDescriptor(
        file="lot1",
        test_name="TestB",
        test_number="2",
        unit="V",
        mean=2.0,
        stdev=0.2,
        cpk=1.5,
    )

    call_counter = {"count": 0}
    original = charts._refresh_all_tests

    def _wrapper(*args, **kwargs):
        call_counter["count"] += 1
        return original(*args, **kwargs)

    monkeypatch.setattr(charts, "_refresh_all_tests", _wrapper)

    charts.refresh_tests(context, [descriptor], include_spec=True, include_proposed=True)

    assert call_counter["count"] == 1

    workbook = context.workbook()
    hist_sheet = workbook[_safe_sheet_name("Histogram_lot1")]
    assert len(hist_sheet._images) == 2

    chart_state = context.metadata.get("post_processing_state", {}).get("chart_positions", {})
    hist_positions = chart_state.get("Histogram", {}).get("lot1", {})
    assert hist_positions["lot1|TestA|1"] == 0
    assert hist_positions["lot1|TestB|2"] == 1


def test_refresh_tests_handles_empty_measurements_subset(monkeypatch, tmp_path: Path) -> None:
    workbook_path, metadata_path, analysis_inputs = _build_two_test_workbook(tmp_path)
    context = postprocess.create_context(
        workbook_path=workbook_path,
        metadata_path=metadata_path,
        analysis_inputs=analysis_inputs,
    )

    charts.refresh_tests(context, [])
    workbook = context.workbook()
    axis_before = charts._load_axis_ranges(workbook)

    descriptor = actions.TestDescriptor(
        file="lot1",
        test_name="TestA",
        test_number="1",
        unit="V",
        mean=1.0,
        stdev=0.1,
        cpk=1.5,
    )

    def _empty_measurements(self, *, refresh: bool = False):
        return pd.DataFrame(columns=["File", "Test Name", "Test Number", "Value"])

    monkeypatch.setattr(context, "measurements_frame", types.MethodType(_empty_measurements, context))

    charts.refresh_tests(context, [descriptor], include_spec=True, include_proposed=True)

    axis_after = charts._load_axis_ranges(context.workbook())
    assert axis_after == axis_before
