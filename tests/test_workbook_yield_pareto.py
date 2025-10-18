from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from cpkanalysis.workbook_builder import build_workbook
from cpkanalysis import workbook_builder
from cpkanalysis.stats import SUMMARY_COLUMNS


def _make_summary_rows() -> pd.DataFrame:
    base = {column: 0 for column in SUMMARY_COLUMNS}
    rows = []
    row1 = base.copy()
    row1.update({"File": "lot1", "Test Name": "T1", "Test Number": "1", "Unit": "V", "COUNT": 3})
    row2 = base.copy()
    row2.update({"File": "lot2", "Test Name": "T1", "Test Number": "1", "Unit": "V", "COUNT": 1})
    rows.extend([row1, row2])
    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def test_build_workbook_creates_yield_pareto_sheet(tmp_path: Path) -> None:
    summary_df = _make_summary_rows()

    measurements_df = pd.DataFrame(
        [
            {"file": "lot1", "device_id": "D1", "test_name": "T1", "test_number": "1", "value": 0.5, "units": "V", "timestamp": 0.0},
        ]
    )

    test_limits_df = pd.DataFrame(
        [
            {
                "test_name": "T1",
                "test_number": "1",
                "unit": "V",
                "stdf_lower": 0.0,
                "stdf_upper": 1.0,
                "spec_lower": None,
                "spec_upper": None,
                "what_if_lower": None,
                "what_if_upper": None,
            }
        ]
    )

    yield_summary = pd.DataFrame(
        [
            {"file": "lot1", "devices_total": 3, "devices_pass": 2, "devices_fail": 1, "yield_percent": 2 / 3},
            {"file": "lot2", "devices_total": 1, "devices_pass": 1, "devices_fail": 0, "yield_percent": 1.0},
        ]
    )

    pareto_summary = pd.DataFrame(
        [
            {
                "file": "lot1",
                "test_name": "T1",
                "test_number": "1",
                "devices_fail": 1,
                "fail_rate_percent": 1 / 3,
                "cumulative_percent": 1 / 3,
                "lower_limit": 0.0,
                "upper_limit": 1.0,
            }
        ]
    )

    output_path = tmp_path / "cpk.xlsx"

    build_workbook(
        summary=summary_df,
        measurements=measurements_df,
        test_limits=test_limits_df,
        limit_sources={},
        outlier_summary={"removed": 0},
        per_file_stats=[],
        output_path=output_path,
        template_path=None,
        include_histogram=False,
        include_cdf=False,
        include_time_series=False,
        include_yield_pareto=True,
        yield_summary=yield_summary,
        pareto_summary=pareto_summary,
        fallback_decimals=None,
        temp_dir=tmp_path,
    )

    workbook = load_workbook(output_path)
    assert "Yield and Pareto" in workbook.sheetnames
    sheet = workbook["Yield and Pareto"]

    assert sheet["A1"].value == "File"
    assert sheet["A2"].value == "lot1"

    assert sheet.cell(row=2, column=5).value == pytest.approx(2 / 3, rel=1e-3)

    found_file_header = any(cell.value == "File: lot1" for cell in sheet[
        "A"
    ])
    assert found_file_header

    table_names = {table.displayName for table in sheet._tables.values()}
    assert "YieldSummaryTable" in table_names
    assert any(name.startswith("YieldTable") for name in table_names)
    assert any(name.startswith("ParetoTable") for name in table_names)


def test_build_workbook_handles_empty_yield_summary(tmp_path: Path) -> None:
    summary_df = _make_summary_rows()
    measurements_df = pd.DataFrame(
        [
            {"file": "lot1", "device_id": "D1", "test_name": "T1", "test_number": "1", "value": 0.5, "units": "V", "timestamp": 0.0},
        ]
    )
    limits_df = pd.DataFrame()
    output_path = tmp_path / "empty.xlsx"

    build_workbook(
        summary=summary_df,
        measurements=measurements_df,
        test_limits=limits_df,
        limit_sources={},
        outlier_summary={"removed": 0},
        per_file_stats=[],
        output_path=output_path,
        template_path=None,
        include_histogram=False,
        include_cdf=False,
        include_time_series=False,
        include_yield_pareto=True,
        yield_summary=pd.DataFrame(columns=["file"]),
        pareto_summary=pd.DataFrame(),
        fallback_decimals=None,
        temp_dir=tmp_path,
    )

    wb = load_workbook(output_path)
    sheet = wb["Yield and Pareto"]
    assert sheet.cell(row=1, column=1).value == "No yield data available."


def test_build_workbook_handles_no_pareto_failures(tmp_path: Path) -> None:
    summary_df = _make_summary_rows().iloc[0:1]
    measurements_df = pd.DataFrame(
        [
            {"file": "lot1", "device_id": "D1", "test_name": "T1", "test_number": "1", "value": 0.5, "units": "V", "timestamp": 0.0},
        ]
    )
    limits_df = pd.DataFrame()
    yield_summary = pd.DataFrame(
        [
            {"file": "lot1", "devices_total": 1, "devices_pass": 1, "devices_fail": 0, "yield_percent": 1.0},
        ]
    )
    output_path = tmp_path / "no_fail.xlsx"

    build_workbook(
        summary=summary_df,
        measurements=measurements_df,
        test_limits=limits_df,
        limit_sources={},
        outlier_summary={"removed": 0},
        per_file_stats=[],
        output_path=output_path,
        template_path=None,
        include_histogram=False,
        include_cdf=False,
        include_time_series=False,
        include_yield_pareto=True,
        yield_summary=yield_summary,
        pareto_summary=pd.DataFrame(),
        fallback_decimals=None,
        temp_dir=tmp_path,
    )

    wb = load_workbook(output_path)
    sheet = wb["Yield and Pareto"]
    assert any(cell.value == "No failing tests" for cell in sheet["A"])


def test_build_workbook_chart_embedding_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    summary_df = _make_summary_rows().iloc[0:1]
    measurements_df = pd.DataFrame(
        [
            {"file": "lot1", "device_id": "D1", "test_name": "T1", "test_number": "1", "value": 0.5, "units": "V", "timestamp": 0.0},
        ]
    )
    limits_df = pd.DataFrame()
    yield_summary = pd.DataFrame(
        [
            {"file": "lot1", "devices_total": 1, "devices_pass": 1, "devices_fail": 0, "yield_percent": 1.0},
        ]
    )

    class FailingImage:
        def __init__(self, *_args, **_kwargs) -> None:
            raise ValueError("image error")

    monkeypatch.setattr(workbook_builder, "XLImage", FailingImage)

    with pytest.raises(ValueError):
        build_workbook(
            summary=summary_df,
            measurements=measurements_df,
            test_limits=limits_df,
            limit_sources={},
            outlier_summary={"removed": 0},
            per_file_stats=[],
            output_path=tmp_path / "fail.xlsx",
            template_path=None,
            include_histogram=False,
            include_cdf=False,
            include_time_series=False,
            include_yield_pareto=True,
            yield_summary=yield_summary,
            pareto_summary=pd.DataFrame(),
            fallback_decimals=None,
            temp_dir=tmp_path,
        )
