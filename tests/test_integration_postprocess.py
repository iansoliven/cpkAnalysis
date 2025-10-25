from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cpkanalysis.postprocess import actions, sheet_utils
from tests import test_postprocess_actions


class MinimalIO(test_postprocess_actions.DummyIO):
    def prompt(self, prompt: str, default: str | None = None) -> str:
        return default or ""


def test_calculate_proposed_limits_grr_roundtrip(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "integration.xlsx"
    test_postprocess_actions._build_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    header_row, header_map = sheet_utils.build_header_map(template_ws)
    ll_ate_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_ate_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    template_ws.cell(row=header_row + 1, column=ll_ate_col, value=-1.6)
    template_ws.cell(row=header_row + 1, column=ul_ate_col, value=1.6)
    wb.save(workbook_path)
    wb.close()

    metadata_path = tmp_path / "meta.json"
    metadata_path.write_text("{}")

    grr_dir = tmp_path / "grr"
    grr_dir.mkdir()
    (grr_dir / "Total_GRR.xlsx").write_text("dummy")

    from cpkanalysis.postprocess.context import PostProcessContext

    inputs = test_postprocess_actions.AnalysisInputs(
        sources=[],
        output=tmp_path / "analysis.xlsx",
        template_sheet="Template",
        generate_histogram=False,
        generate_cdf=False,
        generate_time_series=False,
    )
    context = PostProcessContext(inputs, workbook_path, metadata_path, metadata={})

    io = MinimalIO()

    record = actions.proposed_limits_grr.GRRRecord(
        test_number="1",
        test_name="TestA",
        unit_raw="V",
        unit_normalized="VOLTS",
        spec_lower=-0.5,
        spec_upper=0.5,
        guardband_full=0.05,
    )

    class DummyGRRTable:
        def __init__(self):
            self._records = [record]

        def find(self, test_number: str, test_name: str):
            return record

        def records(self):
            return list(self._records)

    original_loader = actions.proposed_limits_grr.load_grr_table
    actions.proposed_limits_grr.load_grr_table = lambda path: DummyGRRTable()  # type: ignore
    monkeypatch.setattr(actions.charts, "refresh_tests", lambda *args, **kwargs: None)
    try:
        _ = actions.calculate_proposed_limits_grr(
            context,
            io,
            {
                "scope": "single",
                "test_key": "lot1|TestA|1",
                "grr_path": str(grr_dir),
                "grr_available": True,
                "cpk_min": 1.0,
                "cpk_max": 2.0,
            },
        )
    finally:
        actions.proposed_limits_grr.load_grr_table = original_loader  # type: ignore

    save_path = tmp_path / "integration_saved.xlsx"
    context.workbook().save(save_path)
    reloaded = load_workbook(save_path)
    template_reloaded = reloaded["Template"]
    header_row_reloaded, headers_reloaded = sheet_utils.build_header_map(template_reloaded)
    data_row = header_row_reloaded + 1
    ll_prop_col = headers_reloaded[sheet_utils.normalize_header("LL_PROP")]
    ul_prop_col = headers_reloaded[sheet_utils.normalize_header("UL_PROP")]

    assert template_reloaded.cell(row=data_row, column=ll_prop_col).value == pytest.approx(-1.6)
    assert template_reloaded.cell(row=data_row, column=ul_prop_col).value == pytest.approx(1.6)
    state = context.metadata["post_processing_state"]["proposed_limits"]["lot1|TestA|1"]
    assert state["skipped"] is True
    assert state.get("required_guardband") == pytest.approx(0.05)

    lower_pct_col = headers_reloaded[sheet_utils.normalize_header("Lower Guardband Percent")]
    upper_pct_col = headers_reloaded[sheet_utils.normalize_header("Upper Guardband Percent")]
    lower_pct_formula = template_reloaded.cell(row=data_row, column=lower_pct_col).value
    upper_pct_formula = template_reloaded.cell(row=data_row, column=upper_pct_col).value
    assert isinstance(lower_pct_formula, str) and "_GRR_reference" in lower_pct_formula and "SUMIFS" in lower_pct_formula
    assert isinstance(upper_pct_formula, str) and "_GRR_reference" in upper_pct_formula and "SUMIFS" in upper_pct_formula
