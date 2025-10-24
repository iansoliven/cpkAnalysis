from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cpkanalysis.postprocess import actions, sheet_utils
from cpkanalysis.tools import update_cpk_formulas
from tests import test_postprocess_actions


def test_update_cpk_formulas_roundtrip(tmp_path: Path) -> None:
    workbook_path = tmp_path / "roundtrip.xlsx"
    test_postprocess_actions._build_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    update_cpk_formulas.apply_formulas(wb, template_ws)

    save_path = tmp_path / "roundtrip_saved.xlsx"
    wb.save(save_path)
    reloaded = load_workbook(save_path)
    template_reloaded = reloaded["Template"]
    header_row, headers = sheet_utils.build_header_map(template_reloaded)
    data_row = header_row + 1
    cpk_col = headers[sheet_utils.normalize_header("CPK_PROP")]
    cell = template_reloaded.cell(row=data_row, column=cpk_col)
    assert cell.data_type == "f"


def test_spec_difference_cf_roundtrip(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cf_roundtrip.xlsx"
    test_postprocess_actions._build_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    update_cpk_formulas.apply_formulas(wb, template_ws)

    header_row, headers = sheet_utils.build_header_map(template_ws)

    ref_ws = wb.create_sheet(actions.GRR_REFERENCE_SHEET)
    ref_ws.append(["Key", "Test Name", "Test Number", "Spec Lower", "Spec Upper"])
    ref_ws.append(["TestA|1", "TestA", "1", 0.5, 1.5])

    actions._apply_spec_difference_cf(template_ws, headers, header_row)

    save_path = tmp_path / "cf_roundtrip_saved.xlsx"
    wb.save(save_path)
    reloaded = load_workbook(save_path)
    template_reloaded = reloaded["Template"]
    cf_rules = list(template_reloaded.conditional_formatting)
    assert cf_rules
