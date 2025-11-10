from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cpkanalysis.workbook_builder import (
    build_workbook,
    ROW_STRIDE,
    ROW_HEIGHT_POINTS,
)


def _make_simple_inputs(with_fail: bool = True) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    # Minimal summary/measurements/limits for a single test and two devices
    summary_df = pd.DataFrame(
        [
            {
                "File": "lot1",
                "Test Name": "T1",
                "Test Number": "1",
                "Unit": "V",
                "COUNT": 2,
            }
        ]
    )
    measurements_df = pd.DataFrame(
        [
            {"file": "lot1", "device_id": "D1", "test_name": "T1", "test_number": "1", "value": 0.5, "units": "V", "timestamp": 0.0},
            {"file": "lot1", "device_id": "D2", "test_name": "T1", "test_number": "1", "value": 0.8, "units": "V", "timestamp": 1.0},
        ]
    )
    limits_df = pd.DataFrame(
        [
            {
                "test_name": "T1",
                "test_number": "1",
                "unit": "V",
                "stdf_lower": 0.0,
                "stdf_upper": 1.0,
                "spec_lower": None,
                "spec_upper": None,
                "what_if_lower": None,
                "what_if_upper": None,
            }
        ]
    )
    if with_fail:
        yield_summary = pd.DataFrame(
            [
                {"file": "lot1", "devices_total": 2, "devices_pass": 1, "devices_fail": 1, "yield_percent": 0.5},
            ]
        )
        pareto_summary = pd.DataFrame(
            [
                {
                    "file": "lot1",
                    "test_name": "T1",
                    "test_number": "1",
                    "devices_fail": 1,
                    "fail_rate_percent": 0.5,
                    "cumulative_percent": 0.5,
                    "lower_limit": 0.0,
                    "upper_limit": 1.0,
                }
            ]
        )
    else:
        yield_summary = pd.DataFrame(
            [
                {"file": "lot1", "devices_total": 2, "devices_pass": 2, "devices_fail": 0, "yield_percent": 1.0},
            ]
        )
        pareto_summary = pd.DataFrame()
    return summary_df, measurements_df, limits_df, (yield_summary, pareto_summary)


def test_aggregate_yield_pareto_anchors_non_overlapping(tmp_path: Path) -> None:
    summary_df, measurements_df, limits_df, (yield_summary, pareto_summary) = _make_simple_inputs(with_fail=True)

    output_path = tmp_path / "anchors.xlsx"
    build_workbook(
        summary=summary_df,
        measurements=measurements_df,
        test_limits=limits_df,
        limit_sources={},
        outlier_summary={"removed": 0},
        per_file_stats=[],
        output_path=output_path,
        template_path=None,
        include_histogram=False,
        include_cdf=False,
        include_time_series=False,
        include_yield_pareto=True,
        yield_summary=yield_summary,
        pareto_summary=pareto_summary,
        fallback_decimals=None,
        temp_dir=tmp_path,
    )

    wb = load_workbook(output_path)
    ws = wb["Yield and Pareto"]

    # Locate the first file section
    file_header_row = None
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.startswith("File: lot1"):
            file_header_row = cell.row
            break
    assert file_header_row is not None

    # Yield header should be the next row with 'Outcome', 'Units', 'Percent'
    yield_header_row = file_header_row + 1
    headers = [ws.cell(row=yield_header_row, column=col).value for col in (1, 2, 3)]
    assert headers == ["Outcome", "Units", "Percent"]

    # Two rows of data (Pass, Fail) follow
    yield_end_row = yield_header_row + 2
    expected_yield_chart_anchor = yield_end_row + 2
    assert (ws.row_dimensions[expected_yield_chart_anchor].height or 0) >= ROW_HEIGHT_POINTS - 1

    # Pareto section header should be at or beyond yield chart anchor + ROW_STRIDE
    # Find the next row containing 'Pareto' in column A
    pareto_header_row = None
    for r in range(expected_yield_chart_anchor + 1, expected_yield_chart_anchor + ROW_STRIDE + 50):
        if ws.cell(row=r, column=1).value == "Pareto":
            pareto_header_row = r
            break
    assert pareto_header_row is not None
    assert pareto_header_row >= expected_yield_chart_anchor + ROW_STRIDE

    # Pareto table has one data row after its header row (header row is pareto_header_row + 1)
    pareto_end_row = (pareto_header_row + 1) + 1
    expected_pareto_chart_anchor = pareto_end_row + 2
    assert (ws.row_dimensions[expected_pareto_chart_anchor].height or 0) >= ROW_HEIGHT_POINTS - 1


def test_site_yield_pareto_anchors_non_overlapping(tmp_path: Path) -> None:
    # Single file with two sites so we can exercise per-site sections
    summary_df = pd.DataFrame([
        {"File": "lotA.stdf", "Test Name": "VDD", "Test Number": "1", "Unit": "V", "COUNT": 2}
    ])
    measurements_df = pd.DataFrame([
        {"file": "lotA.stdf", "device_id": "D1", "test_name": "VDD", "test_number": "1", "value": 0.5, "units": "V", "site": 1, "timestamp": 0.0},
        {"file": "lotA.stdf", "device_id": "D2", "test_name": "VDD", "test_number": "1", "value": 1.5, "units": "V", "site": 2, "timestamp": 1.0},
    ])
    limits_df = pd.DataFrame([{"test_name": "VDD", "test_number": "1", "unit": "V"}])
    yield_summary = pd.DataFrame([
        {"file": "lotA.stdf", "devices_total": 2, "devices_pass": 1, "devices_fail": 1, "yield_percent": 0.5}
    ])
    site_yield = pd.DataFrame([
        {"file": "lotA.stdf", "site": 1, "devices_total": 1, "devices_pass": 1, "devices_fail": 0, "yield_percent": 1.0},
        {"file": "lotA.stdf", "site": 2, "devices_total": 1, "devices_pass": 0, "devices_fail": 1, "yield_percent": 0.0},
    ])
    site_pareto = pd.DataFrame([
        {"file": "lotA.stdf", "site": 2, "test_name": "VDD", "test_number": "1", "devices_fail": 1, "fail_rate_percent": 1.0, "cumulative_percent": 1.0, "lower_limit": 0.0, "upper_limit": 2.0}
    ])

    output_path = tmp_path / "site_anchors.xlsx"
    build_workbook(
        summary=summary_df,
        measurements=measurements_df,
        test_limits=limits_df,
        limit_sources={},
        outlier_summary={"removed": 0},
        per_file_stats=[],
        output_path=output_path,
        template_path=None,
        include_histogram=False,
        include_cdf=False,
        include_time_series=False,
        include_yield_pareto=True,
        yield_summary=yield_summary,
        pareto_summary=pd.DataFrame(),
        fallback_decimals=None,
        temp_dir=tmp_path,
        site_yield_summary=site_yield,
        site_pareto_summary=site_pareto,
        site_enabled=True,
    )

    wb = load_workbook(output_path)
    ws = wb["Yield and Pareto"]

    # Locate site section header row for Site 1
    site1_header_row = None
    for row in ws.iter_rows(min_row=1, values_only=False):
        for cell in row[:20]:  # search first 20 columns
            if isinstance(cell.value, str) and cell.value.startswith("File: lotA.stdf (Site 1)"):
                site1_header_row = cell.row
                break
        if site1_header_row:
            break
    assert site1_header_row is not None

    # Yield header for site is next row; two data rows follow
    yield_header_row = site1_header_row + 1
    row_values = [ws.cell(row=yield_header_row, column=c).value for c in range(1, 10)]
    assert "Outcome" in row_values and "Units" in row_values and "Percent" in row_values

    yield_end_row = yield_header_row + 2
    site_yield_chart_anchor = yield_end_row + 2
    assert (ws.row_dimensions[site_yield_chart_anchor].height or 0) >= ROW_HEIGHT_POINTS - 1

    # Pareto header should be at or beyond yield chart anchor + ROW_STRIDE
    pareto_header_row = None
    for r in range(site_yield_chart_anchor + 1, site_yield_chart_anchor + ROW_STRIDE + 60):
        # look for any 'Pareto' label in this band (site sections write headers at their base column)
        if any((ws.cell(row=r, column=c).value == "Pareto") for c in range(1, 15)):
            pareto_header_row = r
            break
    assert pareto_header_row is not None
    assert pareto_header_row >= site_yield_chart_anchor + ROW_STRIDE
