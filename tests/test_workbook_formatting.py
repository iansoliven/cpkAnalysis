from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook

from cpkanalysis.workbook_builder import (
    COL_STRIDE,
    _create_plot_sheets,
    _excel_number_format,
    _fixed_decimal_format,
    _row_number_format,
    set_fallback_decimals,
)


@pytest.fixture(autouse=True)
def reset_fallback_decimals() -> None:
    set_fallback_decimals(None)
    yield
    set_fallback_decimals(None)


def test_excel_number_format_from_printf_tokens() -> None:
    assert _excel_number_format("%8.3f", None) == "0.000"
    assert _excel_number_format("%10.2E", None) == "0.00E+00"
    assert _excel_number_format("%6.3g", None) == "0.###"
    assert _excel_number_format("%5d", None) == "0"


def test_excel_number_format_uses_scale_when_missing_format() -> None:
    assert _excel_number_format(None, -3) == "0.###"
    assert _excel_number_format(None, 0) == "0"


def test_fallback_decimals_override_applies_to_missing_hints() -> None:
    set_fallback_decimals(2)
    assert _excel_number_format(None, None) == "0.##"
    set_fallback_decimals(5)
    assert _excel_number_format(None, None) == "0.#####"


def test_row_number_format_prefers_specific_field() -> None:
    row = {
        "stdf_lower_format": "%7.1f",
        "stdf_lower_scale": None,
        "stdf_result_scale": -2,
    }
    assert _row_number_format(
        row,
        format_fields=("stdf_lower_format", "stdf_result_format"),
        scale_fields=("stdf_lower_scale", "stdf_result_scale"),
    ) == "0.0"


def test_fixed_decimal_format_matches_current_fallback() -> None:
    set_fallback_decimals(4)
    assert _fixed_decimal_format(4) == "0.0000"
    set_fallback_decimals(0)
    assert _fixed_decimal_format(0) == "0"


def test_create_plot_sheets_adds_site_plots() -> None:
    workbook = Workbook()
    measurements = pd.DataFrame(
        [
            {
                "file": "first.stdf",
                "file_path": "first.stdf",
                "device_id": "D1",
                "device_sequence": 1,
                "site": 1,
                "test_name": "VDD",
                "test_number": "1",
                "units": "V",
                "value": 1.0,
                "timestamp": 0.0,
                "measurement_index": 1,
            },
            {
                "file": "first.stdf",
                "file_path": "first.stdf",
                "device_id": "D2",
                "device_sequence": 2,
                "site": 2,
                "test_name": "VDD",
                "test_number": "1",
                "units": "V",
                "value": 2.0,
                "timestamp": 1.0,
                "measurement_index": 1,
            },
        ]
    )
    test_limits = pd.DataFrame(
        [
            {
                "test_name": "VDD",
                "test_number": "1",
                "unit": "V",
                "stdf_lower": 0.5,
                "stdf_upper": 1.5,
                "spec_lower": None,
                "spec_upper": None,
                "what_if_lower": None,
                "what_if_upper": None,
            }
        ]
    )
    summary = pd.DataFrame(
        [
            {
                "File": "first.stdf",
                "Test Name": "VDD",
                "Test Number": "1",
                "Unit": "V",
                "COUNT": 2,
                "MEAN": 1.5,
                "MEDIAN": 1.5,
                "STDEV": 0.5,
                "IQR": 0.5,
                "CPL": 0.0,
                "CPU": 0.0,
                "CPK": 0.0,
                "%YLD LOSS": 0.5,
                "LL_2CPK": 0.0,
                "UL_2CPK": 0.0,
                "CPK_2.0": 0.0,
                "%YLD LOSS_2.0": 0.5,
                "LL_3IQR": 0.0,
                "UL_3IQR": 0.0,
                "CPK_3IQR": 0.0,
                "%YLD LOSS_3IQR": 0.5,
            }
        ]
    )
    site_summary = pd.DataFrame(
        [
            summary.iloc[0].to_dict() | {"Site": 1, "COUNT": 1, "MEAN": 1.0, "MEDIAN": 1.0, "%YLD LOSS": 0.0},
            summary.iloc[0].to_dict() | {"Site": 2, "COUNT": 1, "MEAN": 2.0, "MEDIAN": 2.0, "%YLD LOSS": 1.0},
        ]
    )

    plot_links, site_plot_links = _create_plot_sheets(
        workbook,
        measurements,
        test_limits,
        summary,
        include_histogram=True,
        include_cdf=False,
        include_time_series=False,
        site_summary=site_summary,
        site_enabled=True,
    )

    assert plot_links
    assert site_plot_links

    sheet_name = "Histogram_first.stdf"
    sheet = workbook[sheet_name]
    base_label = sheet.cell(row=2, column=1).value
    site_label = sheet.cell(row=2, column=1 + COL_STRIDE).value
    assert base_label is not None and "VDD" in base_label
    assert site_label is not None and "Site" in site_label
