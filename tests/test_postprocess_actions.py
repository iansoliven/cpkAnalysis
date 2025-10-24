from __future__ import annotations

from pathlib import Path
import sys
import math

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cpkanalysis.postprocess import actions, sheet_utils, proposed_limits_grr
from cpkanalysis.postprocess.context import PostProcessContext
from cpkanalysis.models import AnalysisInputs


class DummyIO:
    def __init__(self) -> None:
        self.infos: list[str] = []
        self.warnings: list[str] = []

    def print(self, *args, **kwargs) -> None:  # pragma: no cover
        return None

    def info(self, message: str) -> None:
        self.infos.append(message)

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    def prompt_choice(self, prompt: str, options, show_options: bool = True) -> int:  # pragma: no cover
        return 0

    def prompt(self, prompt: str, default: str | None = None) -> str:
        return default or ""

    def confirm(self, prompt: str, default: bool = True) -> bool:  # pragma: no cover
        return default


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(
        ["File", "Test Name", "Test Number", "Unit", "MEAN", "STDEV", "LL_2CPK", "UL_2CPK", "LL_3IQR", "UL_3IQR", "CPK"]
    )
    summary.append(["lot1", "TestA", "1", "V", 1.0, 0.2, 0.7, 1.3, 0.6, 1.4, 1.5])

    limits = wb.create_sheet("Test List and Limits")
    limits.append(
        [
            "Test Name",
            "Test Number",
            "Spec Lower Limit",
            "Spec Upper Limit",
            "User What-If Lower Limit",
            "User What-If Upper Limit",
        ]
    )
    limits.append(["TestA", "1", None, None, None, None])

    template = wb.create_sheet("Template")
    template.append(
        [
            "Test Name",
            "Test Number",
            "Spec Lower",
            "Spec Upper",
            "What-if Lower",
            "What-if Upper",
            "LL_ATE",
            "UL_ATE",
            "MEAN",
            "MEDIAN",
            "STDEV",
            "IQR",
            "CPL",
            "CPU",
            "CPK",
            "LL_2CPK",
            "UL_2CPK",
            "CPK_2.0",
            "LL_3IQR",
            "UL_3IQR",
            "CPK_3IQR",
            "LL_PROP",
            "UL_PROP",
            "CPK_PROP",
            "%YLD LOSS_PROP",
        ]
    )
    template.append(
        [
            "TestA",
            "1",
            None,
            None,
            None,
            None,
            0.7,
            1.3,
            1.0,
            1.0,
            0.2,
            0.25,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )

    measurements = wb.create_sheet("Measurements")
    measurements.append(["Test Name", "Test Number", "Value"])
    measurements.append(["TestA", "1", 0.2])
    measurements.append(["TestA", "1", 1.8])
    measurements.append(["TestA", "1", 1.0])
    wb.save(path)


def test_prompt_scope_reprompts_until_target_valid(monkeypatch: pytest.MonkeyPatch) -> None:
    class PromptingIO(DummyIO):
        def __init__(self) -> None:
            super().__init__()
            self._responses = iter(["", "1.5"])

        def prompt(self, prompt: str, default: str | None = None) -> str:
            return next(self._responses, "1.5")

    io = PromptingIO()
    result = actions._prompt_scope(io, {"scope": "all", "target_cpk": "invalid"}, allow_single=True, require_target=True)
    assert result["target_cpk"] == pytest.approx(1.5)
    assert any("numeric" in warning.lower() for warning in io.warnings)


def test_prompt_scope_optional_invalid_value(monkeypatch: pytest.MonkeyPatch) -> None:
    io = DummyIO()
    result = actions._prompt_scope(
        io,
        {"scope": "all", "target_cpk": "not-anumber"},
        allow_single=True,
        require_target=False,
        prompt_target=False,
    )
    assert result["target_cpk"] is None
    assert any("numeric" in warning.lower() for warning in io.warnings)


def test_should_skip_grr_rejects_reversed_spec_limits() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Test Name", "Test Number", "LL_ATE", "UL_ATE"])
    ws.append(["TestA", "1", -1.2, 1.2])
    header_row, header_map = sheet_utils.build_header_map(ws)
    template_rows = [header_row + 1]

    descriptor = actions.TestDescriptor(
        file="lot",
        test_name="TestA",
        test_number="1",
        unit="V",
        mean=0.0,
        stdev=0.2,
        cpk=1.4,
    )

    skip, info = actions._should_skip_grr(
        descriptor,
        ws,
        template_rows,
        header_map,
        spec_lower=0.6,
        spec_upper=-0.6,
        min_cpk=1.0,
        max_cpk=2.0,
        guardband_full=0.05,
    )
    assert skip is False
    assert info.get("reason") == "invalid_spec_range"
    assert "lower_guardband_percent" not in info


def test_should_skip_grr_when_guardband_met() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Test Name", "Test Number", "LL_ATE", "UL_ATE"])
    ws.append(["TestA", "1", -0.55, 0.55])
    header_row, header_map = sheet_utils.build_header_map(ws)
    template_rows = [header_row + 1]

    descriptor = actions.TestDescriptor(
        file="lot",
        test_name="TestA",
        test_number="1",
        unit="V",
        mean=0.0,
        stdev=0.2,
        cpk=1.5,
    )

    skip, info = actions._should_skip_grr(
        descriptor,
        ws,
        template_rows,
        header_map,
        spec_lower=-0.5,
        spec_upper=0.5,
        min_cpk=1.0,
        max_cpk=2.0,
        guardband_full=0.05,
    )
    assert skip is True
    assert info.get("reason") == "guardband_satisfied"
    assert info.get("lower_guardband_width") == pytest.approx(0.05)
    assert info.get("upper_guardband_width") == pytest.approx(0.05)


def test_should_skip_grr_when_guardband_insufficient() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Test Name", "Test Number", "LL_ATE", "UL_ATE"])
    ws.append(["TestA", "1", -0.52, 0.52])
    header_row, header_map = sheet_utils.build_header_map(ws)
    template_rows = [header_row + 1]

    descriptor = actions.TestDescriptor(
        file="lot",
        test_name="TestA",
        test_number="1",
        unit="V",
        mean=0.0,
        stdev=0.2,
        cpk=1.5,
    )

    skip, info = actions._should_skip_grr(
        descriptor,
        ws,
        template_rows,
        header_map,
        spec_lower=-0.5,
        spec_upper=0.5,
        min_cpk=1.0,
        max_cpk=2.0,
        guardband_full=0.05,
    )
    assert skip is False
    assert info.get("reason") == "guardband_insufficient"
    assert info.get("lower_guardband_width") == pytest.approx(0.02)
    assert info.get("upper_guardband_width") == pytest.approx(0.02)


def test_compute_yield_loss_returns_nan_without_limits() -> None:
    measurements = pd.DataFrame(
        {
            "Test Name": ["TestA", "TestA"],
            "Test Number": ["1", "1"],
            "Value": [0.1, 0.2],
        }
    )
    result = actions._compute_yield_loss(measurements, "TestA", "1", None, None)
    assert math.isnan(result)


def _load_context(workbook_path: Path, tmp_path: Path) -> PostProcessContext:
    inputs = AnalysisInputs(
        sources=[],
        output=tmp_path / "analysis.xlsx",
        template_sheet="Template",
        generate_histogram=False,
        generate_cdf=False,
        generate_time_series=False,
    )
    metadata_path = tmp_path / "meta.json"
    metadata_path.write_text("{}", encoding="utf-8")
    return PostProcessContext(inputs, workbook_path, metadata_path, metadata={})


def _column_ref(header_map: dict[str, int], header: str, row: int) -> str:
    column = header_map[sheet_utils.normalize_header(header)]
    return f"${get_column_letter(column)}{row}"


def _expected_cpk_prop_formula(header_map: dict[str, int], row: int) -> str:
    mean = _column_ref(header_map, "MEAN", row)
    stdev = _column_ref(header_map, "STDEV", row)
    ll_prop = _column_ref(header_map, "LL_PROP", row)
    ul_prop = _column_ref(header_map, "UL_PROP", row)
    return (
        f"=IF(AND({stdev}>0,{ll_prop}<>\"\",{ul_prop}<>\"\"),"
        f"MIN(({mean}-{ll_prop})/(3*{stdev}),({ul_prop}-{mean})/(3*{stdev})),\"\")"
    )


def _expected_cpk_proposed_spec_formula(header_map: dict[str, int], row: int) -> str:
    mean = _column_ref(header_map, "MEAN", row)
    stdev = _column_ref(header_map, "STDEV", row)
    spec_ll = _column_ref(header_map, "Proposed Spec Lower", row)
    spec_ul = _column_ref(header_map, "Proposed Spec Upper", row)
    return (
        f"=IF(AND({stdev}>0,{spec_ll}<>\"\",{spec_ul}<>\"\"),"
        f"MIN(({mean}-{spec_ll})/(3*{stdev}),({spec_ul}-{mean})/(3*{stdev})),\"\")"
    )







def _expected_lower_guardband_pct_formula(header_map: dict[str, int], row: int) -> str:
    spec_ll = _column_ref(header_map, 'Proposed Spec Lower', row)
    spec_ul = _column_ref(header_map, 'Proposed Spec Upper', row)
    ll_prop = _column_ref(header_map, 'LL_PROP', row)
    spec_width = f'({spec_ul}-{spec_ll})'
    return (
        f'=IF(OR({spec_width}=0,{spec_ll}="",{spec_ul}="",{ll_prop}=""),"",'
        f'ABS({ll_prop}-{spec_ll})/{spec_width})'
    )


def _expected_upper_guardband_pct_formula(header_map: dict[str, int], row: int) -> str:
    spec_ll = _column_ref(header_map, 'Proposed Spec Lower', row)
    spec_ul = _column_ref(header_map, 'Proposed Spec Upper', row)
    ul_prop = _column_ref(header_map, 'UL_PROP', row)
    spec_width = f'({spec_ul}-{spec_ll})'
    return (
        f'=IF(OR({spec_width}=0,{spec_ll}="",{spec_ul}="",{ul_prop}=""),"",'
        f'ABS({spec_ul}-{ul_prop})/{spec_width})'
    )


def _expected_lower_guardband_formula(header_map: dict[str, int], row: int) -> str:
    spec_ll = _column_ref(header_map, 'Proposed Spec Lower', row)
    ll_prop = _column_ref(header_map, 'LL_PROP', row)
    return f'=IF(OR({ll_prop}="",{spec_ll}=""),"",ABS({ll_prop}-{spec_ll}))'


def _expected_upper_guardband_formula(header_map: dict[str, int], row: int) -> str:
    spec_ul = _column_ref(header_map, 'Proposed Spec Upper', row)
    ul_prop = _column_ref(header_map, 'UL_PROP', row)
    return f'=IF(OR({ul_prop}="",{spec_ul}=""),"",ABS({spec_ul}-{ul_prop}))'


def _compute_cpk(mean: float, stdev: float, lower: float | None, upper: float | None) -> float | None:
    if stdev is None or stdev <= 0:
        return None
    candidates: list[float] = []
    if upper is not None:
        candidates.append((upper - mean) / (3 * stdev))
    if lower is not None:
        candidates.append((mean - lower) / (3 * stdev))
    return min(candidates) if candidates else None


def test_apply_spec_limits_updates_template_and_limits(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    refreshed: dict[str, pd.DataFrame] = {}
    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: refreshed.setdefault("tests", list(tests)),
    )

    params = {"scope": "single", "test_key": "lot1|TestA|1", "target_cpk": 1.0}
    result = actions.apply_spec_limits(context, io, params)

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    spec_lower_col = headers[sheet_utils.normalize_header("Spec Lower")]
    spec_upper_col = headers[sheet_utils.normalize_header("Spec Upper")]
    what_lower_col = headers[sheet_utils.normalize_header("What-if Lower")]
    what_upper_col = headers[sheet_utils.normalize_header("What-if Upper")]
    row_idx = header_row + 1

    expected_lower = pytest.approx(1.0 - (3.0 * 1.0 * 0.2))
    expected_upper = pytest.approx(1.0 + (3.0 * 1.0 * 0.2))
    assert template_ws.cell(row=row_idx, column=spec_lower_col).value == expected_lower
    assert template_ws.cell(row=row_idx, column=spec_upper_col).value == expected_upper
    assert template_ws.cell(row=row_idx, column=what_lower_col).value == expected_lower
    assert template_ws.cell(row=row_idx, column=what_upper_col).value == expected_upper

    limits_ws = context.workbook()["Test List and Limits"]
    limits_header_row, limits_headers = sheet_utils.build_header_map(limits_ws)
    spec_ll = limits_headers[sheet_utils.normalize_header("Spec Lower Limit")]
    spec_ul = limits_headers[sheet_utils.normalize_header("Spec Upper Limit")]
    what_ll = limits_headers[sheet_utils.normalize_header("User What-If Lower Limit")]
    what_ul = limits_headers[sheet_utils.normalize_header("User What-If Upper Limit")]
    data_row = limits_header_row + 1

    assert limits_ws.cell(row=data_row, column=spec_ll).value == expected_lower
    assert limits_ws.cell(row=data_row, column=spec_ul).value == expected_upper
    assert limits_ws.cell(row=data_row, column=what_ll).value == expected_lower
    assert limits_ws.cell(row=data_row, column=what_ul).value == expected_upper

    assert refreshed["tests"]
    assert result["summary"] == "Applied Spec/What-If limits for 1 test(s)."
    assert result["audit"]["parameters"]["target_cpk"] == 1.0


def test_apply_spec_limits_missing_columns_cancels(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    wb = load_workbook(workbook_path)
    template = wb["Template"]
    template.delete_cols(3, 6)  # remove spec/what-if and proposal columns
    template.cell(row=1, column=3, value="Other")
    wb.save(workbook_path)
    wb.close()

    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    with pytest.raises(actions.ActionCancelled):
        actions.apply_spec_limits(context, io, {"scope": "all", "target_cpk": 1.0})


def test_calculate_proposed_limits_populates_proposal_columns(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    result = actions.calculate_proposed_limits(context, io, {"scope": "all", "target_cpk": 1.0})

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    ll_prop = headers[sheet_utils.normalize_header("LL_PROP")]
    ul_prop = headers[sheet_utils.normalize_header("UL_PROP")]
    cpk_prop = headers[sheet_utils.normalize_header("CPK_PROP")]
    yld_prop = headers[sheet_utils.normalize_header("%YLD LOSS_PROP")]
    row_idx = header_row + 1

    width = 3.0 * 1.0 * 0.2
    expected_ll_value = 1.0 - width
    expected_ul_value = 1.0 + width

    assert template_ws.cell(row=row_idx, column=ll_prop).value == pytest.approx(expected_ll_value)
    assert template_ws.cell(row=row_idx, column=ul_prop).value == pytest.approx(expected_ul_value)

    cpk_cell = template_ws.cell(row=row_idx, column=cpk_prop)
    assert cpk_cell.data_type == "f"
    expected_form = _expected_cpk_prop_formula(headers, row_idx)
    assert cpk_cell.value == expected_form
    lower_pct_col = headers[sheet_utils.normalize_header("Lower Guardband Percent")]
    upper_pct_col = headers[sheet_utils.normalize_header("Upper Guardband Percent")]
    lower_guard_col = headers[sheet_utils.normalize_header("Lower Guardband")]
    upper_guard_col = headers[sheet_utils.normalize_header("Upper Guardband")]

    lower_pct_cell = template_ws.cell(row=row_idx, column=lower_pct_col)
    upper_pct_cell = template_ws.cell(row=row_idx, column=upper_pct_col)
    lower_guard_cell = template_ws.cell(row=row_idx, column=lower_guard_col)
    upper_guard_cell = template_ws.cell(row=row_idx, column=upper_guard_col)

    assert lower_pct_cell.data_type == "f"
    assert lower_pct_cell.value == _expected_lower_guardband_pct_formula(headers, row_idx)
    assert lower_pct_cell.number_format == "0.0%"
    assert upper_pct_cell.data_type == "f"
    assert upper_pct_cell.value == _expected_upper_guardband_pct_formula(headers, row_idx)
    assert upper_pct_cell.number_format == "0.0%"

    assert lower_guard_cell.data_type == "f"
    assert lower_guard_cell.value == _expected_lower_guardband_formula(headers, row_idx)
    assert lower_guard_cell.number_format == "0.000"
    assert upper_guard_cell.data_type == "f"
    assert upper_guard_cell.value == _expected_upper_guardband_formula(headers, row_idx)
    assert upper_guard_cell.number_format == "0.000"

    mean = template_ws.cell(row=row_idx, column=headers[sheet_utils.normalize_header("MEAN")]).value
    stdev = template_ws.cell(row=row_idx, column=headers[sheet_utils.normalize_header("STDEV")]).value
    computed_cpk = _compute_cpk(mean, stdev, expected_ll_value, expected_ul_value)
    assert computed_cpk == pytest.approx(1.0)

    assert template_ws.cell(row=row_idx, column=yld_prop).value == pytest.approx(2 / 3, rel=1e-3)

    assert result["summary"] == "Updated proposed limits for 1 test(s). Recomputed metrics for 1 test(s)."

    audit_params = result["audit"]["parameters"]
    assert audit_params["target_cpk"] == pytest.approx(1.0)
    assert len(audit_params["per_test"]) == 1
    per_test = audit_params["per_test"][0]
    assert per_test["lower_origin"] == "computed"
    assert per_test["upper_origin"] == "computed"
    assert per_test["cpk_updated"] is True
    assert per_test["yield_updated"] is True

    proposal_state = context.metadata["post_processing_state"]["proposed_limits"]
    entry = proposal_state["lot1|TestA|1"]
    assert entry["ll"] == pytest.approx(expected_ll_value)
    assert entry["ul"] == pytest.approx(expected_ul_value)


def test_calculate_proposed_limits_respects_user_proposals(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    descriptor_key = "lot1|TestA|1"
    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    ll_prop = headers[sheet_utils.normalize_header("LL_PROP")]
    ul_prop = headers[sheet_utils.normalize_header("UL_PROP")]
    cpk_prop = headers[sheet_utils.normalize_header("CPK_PROP")]
    yld_prop = headers[sheet_utils.normalize_header("%YLD LOSS_PROP")]
    data_row = header_row + 1

    user_lower = 0.5
    user_upper = 1.6
    template_ws.cell(row=data_row, column=ll_prop, value=user_lower)
    template_ws.cell(row=data_row, column=ul_prop, value=user_upper)
    template_ws.cell(row=data_row, column=cpk_prop, value=None)
    template_ws.cell(row=data_row, column=yld_prop, value=None)

    context.metadata["post_processing_state"] = {
        "proposed_limits": {descriptor_key: {"ll": 0.7, "ul": 1.3, "timestamp": "prev"}},
    }

    result = actions.calculate_proposed_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 1.2},
    )

    assert template_ws.cell(row=data_row, column=ll_prop).value == pytest.approx(user_lower)
    assert template_ws.cell(row=data_row, column=ul_prop).value == pytest.approx(user_upper)

    cpk_cell = template_ws.cell(row=data_row, column=cpk_prop)
    yld_value = template_ws.cell(row=data_row, column=yld_prop).value
    assert cpk_cell.data_type == "f"
    assert cpk_cell.value == _expected_cpk_prop_formula(headers, data_row)
    mean = template_ws.cell(row=data_row, column=headers[sheet_utils.normalize_header("MEAN")]).value
    stdev = template_ws.cell(row=data_row, column=headers[sheet_utils.normalize_header("STDEV")]).value
    expected_cpk = _compute_cpk(mean, stdev, user_lower, user_upper)
    assert expected_cpk == pytest.approx(0.833333, rel=1e-3)
    assert yld_value == pytest.approx(2 / 3, rel=1e-3)

    per_test = result["audit"]["parameters"]["per_test"][0]
    assert per_test["lower_origin"] == "user"
    assert per_test["upper_origin"] == "user"
    assert per_test["cpk_updated"] is True
    assert per_test["yield_updated"] is True

    state = context.metadata["post_processing_state"]["proposed_limits"][descriptor_key]
    assert state["ll"] == pytest.approx(user_lower)
    assert state["ul"] == pytest.approx(user_upper)


def test_calculate_proposed_limits_computes_blank_side(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    descriptor_key = "lot1|TestA|1"
    context.metadata["post_processing_state"] = {
        "proposed_limits": {descriptor_key: {"ll": 0.4, "ul": 1.5, "timestamp": "prev"}},
    }

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    ll_prop = headers[sheet_utils.normalize_header("LL_PROP")]
    ul_prop = headers[sheet_utils.normalize_header("UL_PROP")]
    cpk_prop = headers[sheet_utils.normalize_header("CPK_PROP")]
    yld_prop = headers[sheet_utils.normalize_header("%YLD LOSS_PROP")]
    data_row = header_row + 1

    template_ws.cell(row=data_row, column=ll_prop, value=0.4)
    template_ws.cell(row=data_row, column=ul_prop, value=None)
    template_ws.cell(row=data_row, column=cpk_prop, value=None)
    template_ws.cell(row=data_row, column=yld_prop, value=None)

    result = actions.calculate_proposed_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 1.0},
    )

    assert template_ws.cell(row=data_row, column=ll_prop).value == pytest.approx(0.4)
    assert template_ws.cell(row=data_row, column=ul_prop).value == pytest.approx(1.6)

    cpk_cell = template_ws.cell(row=data_row, column=cpk_prop)
    assert cpk_cell.data_type == "f"
    assert cpk_cell.value == _expected_cpk_prop_formula(headers, data_row)
    mean = template_ws.cell(row=data_row, column=headers[sheet_utils.normalize_header("MEAN")]).value
    stdev = template_ws.cell(row=data_row, column=headers[sheet_utils.normalize_header("STDEV")]).value
    expected_cpk = _compute_cpk(mean, stdev, 0.4, 1.6)
    assert expected_cpk == pytest.approx(1.0)

    per_test = result["audit"]["parameters"]["per_test"][0]
    assert per_test["lower_origin"] == "unchanged"
    assert per_test["upper_origin"] == "computed"


def test_calculate_proposed_limits_skips_when_no_changes(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    descriptor_key = "lot1|TestA|1"
    context.metadata["post_processing_state"] = {
        "proposed_limits": {descriptor_key: {"ll": 0.7, "ul": 1.3, "timestamp": "prev"}},
    }

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    ll_prop = headers[sheet_utils.normalize_header("LL_PROP")]
    ul_prop = headers[sheet_utils.normalize_header("UL_PROP")]
    cpk_prop = headers[sheet_utils.normalize_header("CPK_PROP")]
    yld_prop = headers[sheet_utils.normalize_header("%YLD LOSS_PROP")]
    data_row = header_row + 1
    template_ws.cell(row=data_row, column=ll_prop, value=0.7)
    template_ws.cell(row=data_row, column=ul_prop, value=1.3)
    template_ws.cell(row=data_row, column=cpk_prop, value=0.9)
    template_ws.cell(row=data_row, column=yld_prop, value=0.05)

    with pytest.raises(actions.ActionCancelled, match="No tests updated."):
        actions.calculate_proposed_limits(
            context,
            io,
            {"scope": "single", "test_key": descriptor_key, "target_cpk": 1.0},
        )


def test_calculate_proposed_limits_recomputes_metrics_when_blank(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    descriptor_key = "lot1|TestA|1"
    context.metadata["post_processing_state"] = {
        "proposed_limits": {descriptor_key: {"ll": 0.7, "ul": 1.3, "timestamp": "prev"}},
    }

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    ll_prop = headers[sheet_utils.normalize_header("LL_PROP")]
    ul_prop = headers[sheet_utils.normalize_header("UL_PROP")]
    cpk_prop = headers[sheet_utils.normalize_header("CPK_PROP")]
    yld_prop = headers[sheet_utils.normalize_header("%YLD LOSS_PROP")]
    data_row = header_row + 1
    template_ws.cell(row=data_row, column=ll_prop, value=0.7)
    template_ws.cell(row=data_row, column=ul_prop, value=1.3)
    template_ws.cell(row=data_row, column=cpk_prop, value=None)
    template_ws.cell(row=data_row, column=yld_prop, value=None)

    result = actions.calculate_proposed_limits(
        context,
        io,
        {"scope": "single", "test_key": descriptor_key, "target_cpk": 1.0},
    )

    cpk_cell = template_ws.cell(row=data_row, column=cpk_prop)
    assert cpk_cell.data_type == "f"
    assert cpk_cell.value == _expected_cpk_prop_formula(headers, data_row)
    mean = template_ws.cell(row=data_row, column=headers[sheet_utils.normalize_header("MEAN")]).value
    stdev = template_ws.cell(row=data_row, column=headers[sheet_utils.normalize_header("STDEV")]).value
    expected_cpk = _compute_cpk(mean, stdev, 0.7, 1.3)
    assert expected_cpk == pytest.approx(0.5)
    assert template_ws.cell(row=data_row, column=yld_prop).value == pytest.approx(2 / 3, rel=1e-3)

    per_test = result["audit"]["parameters"]["per_test"][0]
    assert per_test["lower_origin"] == "unchanged"
    assert per_test["upper_origin"] == "unchanged"
    assert per_test["cpk_updated"] is True
    assert per_test["yield_updated"] is True


def test_calculate_proposed_limits_grr_inserts_formula(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)
    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    header_row, header_map = sheet_utils.build_header_map(template_ws)
    ll_ate_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_ate_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    template_ws.cell(row=header_row + 1, column=ll_ate_col, value=0.66)
    template_ws.cell(row=header_row + 1, column=ul_ate_col, value=1.34)
    wb.save(workbook_path)
    wb.close()
    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    grr_dir = tmp_path / "grr"
    grr_dir.mkdir()
    (grr_dir / "Total_GRR.xlsx").write_bytes(b"dummy")

    record = proposed_limits_grr.GRRRecord(
        test_number="1",
        test_name="testa",
        unit_raw="V",
        unit_normalized="VOLTS",
        spec_lower=0.64,
        spec_upper=1.36,
        guardband_full=0.05,
    )

    class DummyGRRTable:
        def __init__(self):
            self._records = [record]

        def find(self, test_number: str, test_name: str) -> proposed_limits_grr.GRRRecord | None:
            return record

        def records(self):
            return list(self._records)

    monkeypatch.setattr(actions.proposed_limits_grr, "load_grr_table", lambda path: DummyGRRTable())

    comp_result = proposed_limits_grr.ComputationResult(
        guardband_label="100% GRR",
        guardband_value=0.05,
        guardband_cpk=None,
        ft_lower=0.7,
        ft_upper=1.3,
        ft_cpk=1.1,
        spec_lower=0.62,
        spec_upper=1.38,
        spec_cpk=(1.38 - 1.0) / (3 * 0.2),
        spec_widened=False,
        notes=[],
    )
    monkeypatch.setattr(
        actions.proposed_limits_grr,
        "compute_proposed_limits",
        lambda **kwargs: comp_result,
    )

    descriptor_key = "lot1|TestA|1"
    params = {
        "scope": "single",
        "test_key": descriptor_key,
        "grr_path": str(grr_dir),
        "grr_available": True,
        "cpk_min": 1.0,
        "cpk_max": 2.0,
    }

    result = actions.calculate_proposed_limits_grr(context, io, params)

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    row_idx = header_row + 1
    spec_lower_col = headers[sheet_utils.normalize_header("Proposed Spec Lower")]
    spec_upper_col = headers[sheet_utils.normalize_header("Proposed Spec Upper")]
    spec_cpk_col = headers[sheet_utils.normalize_header("CPK Proposed Spec")]

    spec_lower = template_ws.cell(row=row_idx, column=spec_lower_col).value
    spec_upper = template_ws.cell(row=row_idx, column=spec_upper_col).value
    assert spec_lower == pytest.approx(comp_result.spec_lower)
    assert spec_upper == pytest.approx(comp_result.spec_upper)

    spec_cpk_cell = template_ws.cell(row=row_idx, column=spec_cpk_col)
    assert spec_cpk_cell.data_type == "f"
    assert spec_cpk_cell.value == _expected_cpk_proposed_spec_formula(headers, row_idx)

    mean = template_ws.cell(row=row_idx, column=headers[sheet_utils.normalize_header("MEAN")]).value
    stdev = template_ws.cell(row=row_idx, column=headers[sheet_utils.normalize_header("STDEV")]).value
    expected_cpk = _compute_cpk(mean, stdev, spec_lower, spec_upper)
    assert expected_cpk == pytest.approx(comp_result.spec_cpk)

    assert "_GRR_reference" in context.workbook().sheetnames
    assert result["summary"] == "Calculated GRR-based proposed limits for 1 test(s). Updated metrics for 2 test(s)."



def test_calculate_proposed_limits_grr_skips_when_guardband_met(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    _build_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    template_ws = wb["Template"]
    header_row, header_map = sheet_utils.build_header_map(template_ws)
    ll_ate_col = header_map[sheet_utils.normalize_header("LL_ATE")]
    ul_ate_col = header_map[sheet_utils.normalize_header("UL_ATE")]
    template_ws.cell(row=header_row + 1, column=ll_ate_col, value=-1.6)
    template_ws.cell(row=header_row + 1, column=ul_ate_col, value=1.6)

    summary_ws = wb["Summary"]
    summary_header_row, summary_headers = sheet_utils.build_header_map(summary_ws)
    cpk_col = summary_headers[sheet_utils.normalize_header("CPK")]
    summary_ws.cell(row=summary_header_row + 1, column=cpk_col, value=1.2)
    wb.save(workbook_path)
    wb.close()

    context = _load_context(workbook_path, tmp_path)
    io = DummyIO()

    monkeypatch.setattr(
        actions.charts,
        "refresh_tests",
        lambda ctx, tests, **kwargs: None,
    )

    grr_dir = tmp_path / "grr"
    grr_dir.mkdir()
    (grr_dir / "Total_GRR.xlsx").write_bytes(b"dummy")

    record = proposed_limits_grr.GRRRecord(
        test_number="1",
        test_name="TestA",
        unit_raw="V",
        unit_normalized="VOLTS",
        spec_lower=-0.5,
        spec_upper=0.5,
        guardband_full=0.05,
    )

    class SkipDummyGRRTable:
        def __init__(self):
            self._records = [record]

        def find(self, test_number: str, test_name: str) -> proposed_limits_grr.GRRRecord | None:
            return record

        def records(self):
            return list(self._records)

    monkeypatch.setattr(actions.proposed_limits_grr, "load_grr_table", lambda path: SkipDummyGRRTable())

    result = actions.calculate_proposed_limits_grr(
        context,
        io,
        {
            "scope": "single",
            "test_key": "lot1|TestA|1",
            "grr_path": str(grr_dir),
            "grr_available": True,
            "cpk_min": 1.0,
            "cpk_max": 2.0,
        },
    )

    template_ws = context.template_sheet()
    header_row, headers = sheet_utils.build_header_map(template_ws)
    data_row = header_row + 1
    ll_prop_col = headers[sheet_utils.normalize_header("LL_PROP")]
    ul_prop_col = headers[sheet_utils.normalize_header("UL_PROP")]
    spec_lower_col = headers[sheet_utils.normalize_header("Proposed Spec Lower")]
    spec_upper_col = headers[sheet_utils.normalize_header("Proposed Spec Upper")]

    assert template_ws.cell(row=data_row, column=ll_prop_col).value == pytest.approx(-1.6)
    assert template_ws.cell(row=data_row, column=ul_prop_col).value == pytest.approx(1.6)
    assert template_ws.cell(row=data_row, column=spec_lower_col).value == pytest.approx(record.spec_lower)
    assert template_ws.cell(row=data_row, column=spec_upper_col).value == pytest.approx(record.spec_upper)

    assert "Skipped 1 test(s) (guardband already satisfied)." in result["summary"]

    state = context.metadata["post_processing_state"]["proposed_limits"]["lot1|TestA|1"]
    assert state["skipped"] is True
    assert state.get("required_guardband") == pytest.approx(record.guardband_full)

    per_test = result["audit"]["parameters"]["per_test"][0]
    assert per_test.get("skipped") is True
    assert per_test.get("existing_lower_guardband_width") == pytest.approx(abs(-1.6 - record.spec_lower))
    assert per_test.get("existing_upper_guardband_width") == pytest.approx(abs(record.spec_upper - 1.6))
    assert per_test.get("required_guardband") == pytest.approx(record.guardband_full)

    lower_guardband_col = headers[sheet_utils.normalize_header("Lower Guardband")]
    upper_guardband_col = headers[sheet_utils.normalize_header("Upper Guardband")]
    lower_guard_cell = template_ws.cell(row=data_row, column=lower_guardband_col)
    upper_guard_cell = template_ws.cell(row=data_row, column=upper_guardband_col)
    assert lower_guard_cell.data_type == "f"
    assert lower_guard_cell.value == _expected_lower_guardband_formula(headers, data_row)
    assert upper_guard_cell.data_type == "f"
    assert upper_guard_cell.value == _expected_upper_guardband_formula(headers, data_row)

