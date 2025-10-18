from __future__ import annotations

import inspect
import sys
from pathlib import Path
from typing import Any, Callable, Iterator

import pandas as pd
import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cpkanalysis import pipeline, workbook_builder
from cpkanalysis.models import AnalysisInputs, SourceFile, PluginConfig, IngestResult
from cpkanalysis.postprocess import menu


# ---------------------------------------------------------------------------
# Workbook builder tests
# ---------------------------------------------------------------------------


def test_write_measurements_splits_large_frames(monkeypatch: pytest.MonkeyPatch) -> None:
    original = workbook_builder._write_measurements
    source = inspect.getsource(original)
    patched_source = source.replace("max_rows = 1_048_576", "max_rows = 5")
    namespace: dict[str, Any] = {}
    exec(patched_source, workbook_builder.__dict__, namespace)
    monkeypatch.setattr(workbook_builder, "_write_measurements", namespace["_write_measurements"])

    wb = Workbook()
    measurements = pd.DataFrame(
        {
            "file": [f"lot{i}" for i in range(8)],
            "device_id": [f"D{i}" for i in range(8)],
            "test_name": ["T"] * 8,
            "test_number": ["1"] * 8,
            "value": list(range(8)),
            "units": ["V"] * 8,
            "timestamp": list(range(8)),
        }
    )

    workbook_builder._write_measurements(wb, measurements)

    assert "Measurements" in wb.sheetnames
    assert "Measurements_2" in wb.sheetnames
    first = wb["Measurements"]
    second = wb["Measurements_2"]
    assert first.max_row == 5  # header + 4 rows with chunk size 4
    assert second.max_row == 5


# ---------------------------------------------------------------------------
# Pipeline failure handling tests
# ---------------------------------------------------------------------------


def _make_analysis_inputs(tmp_path: Path) -> AnalysisInputs:
    source_path = tmp_path / "input.stdf"
    source_path.write_text("stdf")
    return AnalysisInputs(
        sources=[SourceFile(source_path)],
        output=tmp_path / "out.xlsx",
        generate_histogram=False,
        generate_cdf=False,
        generate_time_series=False,
    )


def test_pipeline_cleans_session_when_stage_fails(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    config = _make_analysis_inputs(tmp_path)

    def failing_ingest(sources, temp_dir):
        raise RuntimeError("ingest boom")

    cleanup_calls: list[Path] = []

    monkeypatch.setattr(pipeline.ingest, "ingest_sources", failing_ingest)
    monkeypatch.setattr(pipeline, "_cleanup_session_dir", lambda path: cleanup_calls.append(path))

    pipe = pipeline.Pipeline(config)

    with pytest.raises(RuntimeError):
        pipe.run()

    assert cleanup_calls and cleanup_calls[0] == pipe._session_dir


def test_pipeline_metadata_failure_triggers_cleanup(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    config = _make_analysis_inputs(tmp_path)

    frame = pd.DataFrame({"file": ["lot"], "device_id": ["D"], "test_name": ["T"], "test_number": ["1"], "value": [1.0], "units": ["V"], "timestamp": [0.0]})
    limits = pd.DataFrame({"test_name": ["T"], "test_number": ["1"], "unit": ["V"]})
    ingest_result = IngestResult(frame=frame, test_catalog=limits, per_file_stats=[{"file": "lot"}], raw_store_path=tmp_path / "raw.parquet")

    monkeypatch.setattr(pipeline.ingest, "ingest_sources", lambda sources, temp_dir: ingest_result)
    monkeypatch.setattr(pipeline.outliers, "apply_outlier_filter", lambda frame, method, k: (frame, {"removed": 0}))
    summary_df = pd.DataFrame({"File": ["lot"], "Test Name": ["T"], "Test Number": ["1"]})
    monkeypatch.setattr(pipeline.stats, "compute_summary", lambda measurements, limits: (summary_df, {}))
    monkeypatch.setattr(pipeline.stats, "compute_yield_pareto", lambda measurements, limits: (pd.DataFrame(), pd.DataFrame()))

    class DummyWorkbook:
        def save(self, path: Path) -> None:
            pass

        def close(self) -> None:
            pass

    monkeypatch.setattr(workbook_builder, "build_workbook", lambda **kwargs: DummyWorkbook())

    cleanup_calls: list[Path] = []
    monkeypatch.setattr(pipeline, "_cleanup_session_dir", lambda path: cleanup_calls.append(path))

    original_write_text = pipeline.Path.write_text

    def failing_write_text(self, data: str, encoding: str = "utf-8", errors: str | None = None) -> int:
        if self.suffix == ".json":
            raise OSError("disk full")
        return original_write_text(self, data, encoding=encoding, errors=errors)

    monkeypatch.setattr(pipeline.Path, "write_text", failing_write_text, raising=False)

    pipe = pipeline.Pipeline(config)

    with pytest.raises(OSError):
        pipe.run()

    assert cleanup_calls and cleanup_calls[0] == pipe._session_dir


# ---------------------------------------------------------------------------
# Post-process menu tests
# ---------------------------------------------------------------------------


class DummyIO:
    def __init__(self, choices: Iterator[int], confirm_response: bool = True) -> None:
        self._choices = iter(choices)
        self.confirm_response = confirm_response
        self.printed: list[str] = []
        self.infos: list[str] = []
        self.warnings: list[str] = []

    def print(self, message: str = "") -> None:
        self.printed.append(message)

    def info(self, message: str) -> None:
        self.infos.append(message)

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    def prompt_choice(self, prompt: str, options: list[str], show_options: bool = True) -> int:
        return next(self._choices)

    def confirm(self, prompt: str, default: bool = True) -> bool:
        return self.confirm_response


class DummyContext:
    def __init__(self) -> None:
        self.dirty = False
        self.saved = False
        self.reloaded = False
        self.mark_dirty_calls = 0
        self.audit_log: list[dict[str, Any]] = []
        self.metadata = {}

    def mark_dirty(self) -> None:
        self.dirty = True
        self.mark_dirty_calls += 1

    def reload(self) -> None:
        self.reloaded = True

    def save(self) -> None:
        self.saved = True

    def add_audit_entry(self, entry: dict[str, Any]) -> None:
        self.audit_log.append(entry)


def test_postprocess_menu_runs_actions_and_exit(monkeypatch: pytest.MonkeyPatch) -> None:
    context = DummyContext()
    action_calls: list[dict[str, Any]] = []

    def fake_action(ctx, io, params):
        action_calls.append({"ctx": ctx, "params": params})
        return {
            "summary": "done",
            "warnings": [],
            "audit": {"scope": "all", "tests": []},
            "replay_params": {"foo": "bar"},
            "mark_dirty": True,
        }

    action_def = menu.ActionDefinition("demo", "Demo Action", fake_action)
    monkeypatch.setattr(menu, "ACTION_DEFINITIONS", (action_def,))

    io = DummyIO(iter([0, 1, 4]))

    menu.loop(context, io=io)  # type: ignore[arg-type]

    assert len(action_calls) == 2  # original + rerun
    assert context.saved is True
    assert context.mark_dirty_calls >= 1
    assert any("Exiting post-processing menu." in msg for msg in io.infos)
