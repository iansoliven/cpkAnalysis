from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cpkanalysis import move_to_template


def _build_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    report = wb.active
    report.title = "CPK Report"
    report.append(["Test Name", "Value"])
    report.append(["T1", 1.23])
    report.append(["T2", 4.56])

    template = wb.create_sheet("Template")
    template.append(["Cpk Report", ""])
    template.append(["Test Name", "Value"])
    template.append(["", ""])
    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    wb.close()
    return path


def test_apply_template_copies_matching_headers(tmp_path: Path) -> None:
    path = _build_workbook(tmp_path)
    wb = load_workbook(path)
    try:
        target_sheet = move_to_template.apply_template(wb, sheet_name="Template")
        assert target_sheet == "Template"
        template_ws = wb["Template"]
        assert template_ws["A3"].value == "T1"
        assert template_ws["B4"].value == pytest.approx(4.56)
    finally:
        wb.close()


def test_run_writes_changes_to_disk(tmp_path: Path) -> None:
    path = _build_workbook(tmp_path)
    result_sheet = move_to_template.run(path)
    assert result_sheet == "Template"
    wb = load_workbook(path)
    try:
        template_ws = wb["Template"]
        assert template_ws["A3"].value == "T1"
        assert template_ws["A4"].value == "T2"
    finally:
        wb.close()
