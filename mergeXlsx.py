#!/usr/bin/env python3
"""
Combine worksheets from multiple Excel workbooks into a single workbook.

The script copies cell values and, by default, basic formatting (fonts, fills, borders,
column widths, row heights, merged cells, and frozen panes). Advanced Excel-specific
features such as pivot tables, slicers, or macros are not preserved.
"""

from __future__ import annotations

import argparse
import os
import sys
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

EXCEL_SHEET_NAME_LIMIT = 31
INVALID_SHEET_CHARS = set('[]:*?/\\')

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Combine worksheets from multiple Excel workbooks into a single file."
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        type=Path,
        help="Optional paths to the source .xlsx files. Defaults to all .xlsx files in the directory."
    )
    parser.add_argument(
        "-d",
        "--directory",
        type=Path,
        default=Path.cwd(),
        help="Directory to search when no explicit input files are provided."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("combined_workbook.xlsx"),
        help="Path for the merged workbook (default: combined_workbook.xlsx)."
    )
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Include hidden worksheets from the source workbooks."
    )
    parser.add_argument(
        "--values-only",
        action="store_true",
        help="Copy only cell values and skip formatting to speed up the merge."
    )
    return parser.parse_args()

def sanitize_sheet_title(name: str) -> str:
    sanitized = ''.join('_' if ch in INVALID_SHEET_CHARS else ch for ch in name)
    sanitized = sanitized.rstrip("'").strip()
    return sanitized or "Sheet"

def build_unique_sheet_title(source_file: Path, sheet_name: str, used_titles: Set[str]) -> str:
    base = sanitize_sheet_title(f"{source_file.stem}_{sheet_name}")
    base = base[:EXCEL_SHEET_NAME_LIMIT] or "Sheet"
    candidate = base
    index = 1
    while candidate in used_titles:
        suffix = f"_{index}"
        allowed_length = EXCEL_SHEET_NAME_LIMIT - len(suffix)
        trimmed = base[:allowed_length] if allowed_length > 0 else base[:EXCEL_SHEET_NAME_LIMIT]
        candidate = (trimmed or "Sheet") + suffix
        index += 1
    used_titles.add(candidate)
    return candidate

@dataclass
class SummaryEntry:
    sheet_title: str
    lot: str
    event: str
    intensity: str
    file_name: str
    file_link: str
    pass_count: int
    fail_count: int


@dataclass
class MeasurementRecord:
    lot: str
    event: str
    intensity: str
    sn: str
    test_number: str
    test_name: str
    test_unit: str
    low_limit: Optional[float]
    high_limit: Optional[float]
    measurement: Any


def parse_filename_metadata(stem: str) -> Tuple[str, str, str]:
    parts = stem.split('_')
    lot = parts[1] if len(parts) >= 2 else ''
    event = parts[2] if len(parts) >= 3 else ''
    intensity = '_'.join(parts[3:]) if len(parts) >= 4 else ''
    if not lot and stem:
        lot = stem[0]
    return lot, event, intensity


def build_file_link(source_file: Path, output_dir: Path) -> str:
    source_resolved = source_file.resolve()
    output_resolved = output_dir.resolve()
    try:
        rel_path = source_resolved.relative_to(output_resolved)
        return rel_path.as_posix()
    except ValueError:
        try:
            rel_string = os.path.relpath(source_resolved, output_resolved)
        except ValueError:
            return source_resolved.as_posix()
        return Path(rel_string).as_posix()


def analyze_unit_results(sheet: Worksheet) -> Tuple[int, int]:
    header_row_index: Optional[int] = None
    header_values: List[str] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        for cell_value in row:
            if isinstance(cell_value, str) and cell_value.strip().upper() == 'UNIT SN':
                header_row_index = idx
                header_values = [
                    value.strip() if isinstance(value, str) else ''
                    for value in row
                ]
                break
        if header_row_index is not None:
            break
    if header_row_index is None:
        return 0, 0
    try:
        sn_col = next(
            index for index, value in enumerate(header_values)
            if isinstance(value, str) and value.strip().upper() == 'UNIT SN'
        )
    except StopIteration:
        return 0, 0
    try:
        status_col = next(
            index for index, value in enumerate(header_values)
            if isinstance(value, str) and value.strip().upper() == 'PASS/FAIL'
        )
    except StopIteration:
        return 0, 0

    outcomes: Dict[str, str] = {}
    mode: Optional[str] = None
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if sn_col >= len(row):
            continue
        sn_cell = row[sn_col]
        if isinstance(sn_cell, str):
            label = sn_cell.strip()
            upper_label = label.upper()
            if 'FAIL UNITS' in upper_label:
                mode = 'fail'
                continue
            if 'PASS UNITS' in upper_label:
                mode = 'pass'
                continue
            if not label:
                continue
        if sn_cell is None:
            continue
        sn_value = str(sn_cell).strip()
        if not sn_value:
            continue
        status_cell = row[status_col] if status_col < len(row) else None
        status_text = ''
        if isinstance(status_cell, str):
            status_text = status_cell.strip().upper()
        elif status_cell is not None:
            status_text = str(status_cell).strip().upper()
        if status_text in {'P', 'PASS'}:
            outcomes[sn_value] = 'pass'
        elif status_text:
            outcomes[sn_value] = 'fail'
        else:
            if mode == 'pass':
                outcomes[sn_value] = 'pass'
            elif mode == 'fail':
                outcomes[sn_value] = 'fail'
    pass_total = sum(1 for result in outcomes.values() if result == 'pass')
    fail_total = sum(1 for result in outcomes.values() if result == 'fail')
    return pass_total, fail_total


def _find_column_index(row: Sequence[Any], target_label: str) -> Optional[int]:
    target = target_label.strip().upper()
    for idx, value in enumerate(row):
        if isinstance(value, str) and value.strip().upper() == target:
            return idx
    return None


def _get_cell(row: Optional[Sequence[Any]], index: int) -> Any:
    if row is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        upper = stripped.upper()
        if upper in {'.', 'X', 'N/A'}:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None




def _normalize_serial_value(value: Any) -> Any:
    numeric = _coerce_float(value)
    if numeric is None:
        if isinstance(value, str):
            stripped = value.strip()
            if stripped.isdigit():
                try:
                    return int(stripped)
                except ValueError:
                    return value
        return value
    return int(numeric) if isinstance(numeric, float) and numeric.is_integer() else numeric

def _normalize_measurement(value: Any) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        upper = stripped.upper()
        if upper in {'.', 'X', 'N/A', 'PASS', 'FAIL'}:
            return None
        try:
            return float(stripped)
        except ValueError:
            return stripped
    return value


def _clean_test_name(raw: Any) -> str:
    if raw is None:
        return ''
    text = str(raw).strip()

    if not text:
        return ''
    if text.startswith('"') and text.endswith('"') and len(text) > 1:
        text = text[1:-1]
    if '",' in text:
        text = text.split('",', 1)[0]
    return text


def _build_test_columns(header_row: Sequence[Any], metadata: Dict[str, Sequence[Any]], status_col: int) -> List[Dict[str, Any]]:
    number_row = metadata.get('number')
    name_row = metadata.get('name')
    unit_row = metadata.get('unit')
    low_row = metadata.get('low')
    high_row = metadata.get('high')
    if not number_row or not name_row or not unit_row:
        return []

    start_index: Optional[int] = None
    for idx, value in enumerate(header_row):
        if isinstance(value, str) and value.strip().upper() == 'RETEST ANALYSIS':
            start_index = idx + 1
            break
    if start_index is None:
        start_index = status_col + 1

    max_len = max(len(number_row), len(name_row), len(unit_row), len(low_row or []), len(high_row or []))
    columns: List[Dict[str, Any]] = []
    for idx in range(start_index, max_len):
        test_number = _get_cell(number_row, idx)
        test_name = _get_cell(name_row, idx)
        test_unit = _get_cell(unit_row, idx)
        if test_number is None or test_name is None or test_unit is None:
            continue
        number_text = str(test_number).strip()
        if not number_text:
            continue
        name_text = _clean_test_name(test_name)
        if not name_text:
            continue
        unit_text = str(test_unit).strip()
        if not unit_text:
            continue
        columns.append(
            {
                'index': idx,
                'test_number': number_text,
                'test_name': name_text,
                'test_unit': unit_text,
                'low_limit': _coerce_float(_get_cell(low_row, idx)),
                'high_limit': _coerce_float(_get_cell(high_row, idx)),
            }
        )
    return columns


def extract_pass_measurements(sheet: Worksheet, lot: str, event: str, intensity: str) -> List[MeasurementRecord]:
    metadata_rows: Dict[str, Sequence[Any]] = {}
    header_row: Optional[Sequence[Any]] = None
    header_row_index: Optional[int] = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        first_cell = row[0]
        if not isinstance(first_cell, str):
            continue
        label = first_cell.strip().upper()
        if label == 'TEST NUMBER':
            metadata_rows['number'] = list(row)
        elif label == 'TEST NAME':
            metadata_rows['name'] = list(row)
        elif label == 'TEST UNIT':
            metadata_rows['unit'] = list(row)
        elif label == 'LOW LIMIT':
            metadata_rows['low'] = list(row)
        elif label == 'HIGH LIMIT':
            metadata_rows['high'] = list(row)
        elif label == 'UNIT SN':
            header_row = list(row)
            header_row_index = idx
            break

    if header_row is None or header_row_index is None:
        return []

    sn_col = _find_column_index(header_row, 'UNIT SN')
    status_col = _find_column_index(header_row, 'PASS/FAIL')
    if sn_col is None or status_col is None:
        return []

    test_columns = _build_test_columns(header_row, metadata_rows, status_col)
    if not test_columns:
        return []

    pass_rows_start: Optional[int] = None
    for idx, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
        if not row:
            continue
        first_cell = row[0]
        if isinstance(first_cell, str) and 'PASS UNITS' in first_cell.strip().upper():
            pass_rows_start = idx + 1
            break

    if pass_rows_start is None:
        return []

    records: List[MeasurementRecord] = []
    for row in sheet.iter_rows(min_row=pass_rows_start, values_only=True):
        if not row:
            break
        sn_value = _get_cell(row, sn_col)
        status_value = _get_cell(row, status_col)
        if sn_value is None and status_value is None:
            if all(cell in (None, '') for cell in row):
                break
            continue
        status_text = ''
        if isinstance(status_value, str):
            status_text = status_value.strip().upper()
        elif status_value is not None:
            status_text = str(status_value).strip().upper()
        if status_text not in {'P', 'PASS'}:
            continue
        sn_text = str(sn_value).strip() if sn_value is not None else ''
        if not sn_text:
            continue
        for column in test_columns:
            measurement_raw = _get_cell(row, column['index'])
            measurement_value = _normalize_measurement(measurement_raw)
            if measurement_value is None:
                continue
            records.append(
                MeasurementRecord(
                    lot=lot,
                    event=event,
                    intensity=intensity,
                    sn=sn_text,
                    test_number=column['test_number'],
                    test_name=column['test_name'],
                    test_unit=column['test_unit'],
                    low_limit=column['low_limit'],
                    high_limit=column['high_limit'],
                    measurement=measurement_value,
                )
            )
    return records


def create_summary_sheet(workbook: Workbook, entries: Sequence[SummaryEntry]) -> None:
    if not entries:
        return
    summary_ws = workbook.create_sheet(title='Summary', index=0)
    headers = ['Sheet Name', 'Lot', 'Event', 'Int', 'File Name', 'Passing Units', 'Failing Units']
    summary_ws.append(headers)
    for entry in entries:
        summary_ws.append([
            entry.sheet_title,
            entry.lot,
            entry.event,
            entry.intensity,
            entry.file_name,
            entry.pass_count,
            entry.fail_count,
        ])
    for row_idx, entry in enumerate(entries, start=2):
        sheet_cell = summary_ws.cell(row=row_idx, column=1)
        sheet_cell.hyperlink = f"#{quote_sheetname(entry.sheet_title)}!A1"
        sheet_cell.style = 'Hyperlink'
        file_cell = summary_ws.cell(row=row_idx, column=5)
        file_cell.hyperlink = entry.file_link
        file_cell.style = 'Hyperlink'
    last_row = len(entries) + 1
    table = Table(displayName='SummaryTable', ref=f'A1:G{last_row}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showRowStripes=True,
        showColumnStripes=False,
    )
    summary_ws.add_table(table)
    summary_ws.freeze_panes = 'A2'
    for index, width in ((1, 32), (2, 14), (3, 18), (4, 16), (5, 48), (6, 18), (7, 18)):
        summary_ws.column_dimensions[get_column_letter(index)].width = width


def create_measurements_sheet(workbook: Workbook, records: Sequence[MeasurementRecord]) -> None:
    if not records:
        return
    max_rows = 1048576
    data_capacity = max_rows - 1
    base_index = 1 if 'Summary' in workbook.sheetnames else 0
    headers = ['Lot', 'Event', 'Int', 'SN', 'Test Number', 'Test Name', 'Test Unit', 'Low Limit', 'High Limit', 'Measurement']
    widths = [16, 18, 12, 12, 18, 42, 12, 14, 14, 16]
    total = len(records)
    chunks = (total + data_capacity - 1) // data_capacity
    for chunk_idx in range(chunks):
        start_row = chunk_idx * data_capacity
        end_row = min(start_row + data_capacity, total)
        chunk = records[start_row:end_row]
        if not chunk:
            continue
        suffix = '' if chunk_idx == 0 else f'_{chunk_idx + 1}'
        sheet_title = 'Measurements' if chunk_idx == 0 else f'Measurements{suffix}'
        worksheet_index = base_index + chunk_idx
        detail_ws = workbook.create_sheet(title=sheet_title, index=worksheet_index)
        detail_ws.append(headers)
        for record in chunk:
            sn_value = _normalize_serial_value(record.sn)
            detail_ws.append([
                record.lot,
                record.event,
                record.intensity,
                sn_value,
                record.test_number,
                record.test_name,
                record.test_unit,
                record.low_limit,
                record.high_limit,
                record.measurement,
            ])
        last_row = len(chunk) + 1
        table_name = 'MeasurementTable' if chunk_idx == 0 else f'MeasurementTable{chunk_idx + 1}'
        table = Table(displayName=table_name, ref=f'A1:J{last_row}')
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium9',
            showRowStripes=True,
            showColumnStripes=False,
        )
        detail_ws.add_table(table)
        detail_ws.freeze_panes = 'A2'
        for idx, width in enumerate(widths, start=1):
            detail_ws.column_dimensions[get_column_letter(idx)].width = width


def gather_sources(inputs: Sequence[Path], directory: Path, output_path: Path) -> List[Path]:
    if inputs:
        source_candidates: List[Path] = []
        missing: List[str] = []
        for path in inputs:
            expanded = path.expanduser()
            if not expanded.exists():
                missing.append(str(expanded))
                continue
            if expanded.is_dir():
                raise ValueError(f"Input path '{expanded}' is a directory. Provide .xlsx files instead.")
            source_candidates.append(expanded.resolve())
        if missing:
            raise FileNotFoundError(f"Source file(s) not found: {', '.join(missing)}")
    else:
        search_dir = directory.expanduser()
        if not search_dir.exists():
            raise FileNotFoundError(f"Directory not found: {search_dir}")
        if not search_dir.is_dir():
            raise NotADirectoryError(f"Not a directory: {search_dir}")
        source_candidates = sorted(search_dir.glob("*.xlsx"))

    output_resolved = output_path.expanduser().resolve()
    filtered: List[Path] = []
    seen = set()
    for candidate in source_candidates:
        if candidate.name.startswith("~$"):
            continue
        if candidate.suffix.lower() != ".xlsx":
            continue
        resolved = candidate.resolve()
        if resolved == output_resolved:
            continue
        if resolved in seen:
            continue
        seen.add(resolved)
        filtered.append(resolved)

    if not filtered:
        raise FileNotFoundError("No .xlsx source files were found to merge.")
    return filtered

def copy_sheet_contents(source_ws, target_ws, copy_formatting: bool) -> None:
    if copy_formatting:
        if source_ws.sheet_properties.tabColor is not None:
            target_ws.sheet_properties.tabColor = copy(source_ws.sheet_properties.tabColor)
        if source_ws.freeze_panes:
            target_ws.freeze_panes = source_ws.freeze_panes

        for idx, row_dimension in source_ws.row_dimensions.items():
            target_dimension = target_ws.row_dimensions[idx]
            if row_dimension.height is not None:
                target_dimension.height = row_dimension.height
            if row_dimension.hidden is not None:
                target_dimension.hidden = row_dimension.hidden
            if row_dimension.outlineLevel:
                target_dimension.outlineLevel = row_dimension.outlineLevel

        for key, column_dimension in source_ws.column_dimensions.items():
            target_dimension = target_ws.column_dimensions[key]
            if column_dimension.width is not None:
                target_dimension.width = column_dimension.width
            if column_dimension.hidden is not None:
                target_dimension.hidden = column_dimension.hidden
            if column_dimension.outlineLevel:
                target_dimension.outlineLevel = column_dimension.outlineLevel
            if column_dimension.bestFit is not None:
                target_dimension.bestFit = column_dimension.bestFit

        if source_ws.auto_filter and source_ws.auto_filter.ref:
            target_ws.auto_filter.ref = source_ws.auto_filter.ref

    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if copy_formatting and cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)
            if cell.hyperlink:
                target_cell.hyperlink = cell.hyperlink.target
            if copy_formatting and cell.comment:
                target_cell.comment = copy(cell.comment)

    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))

def main() -> int:
    args = parse_args()
    try:
        sources = gather_sources(args.inputs, args.directory, args.output)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    destination_path = args.output.expanduser().resolve()
    destination_dir = destination_path.parent
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: Set[str] = set()
    copied_sheets = 0
    skipped_hidden = 0
    failures: List[str] = []
    summary_entries: List[SummaryEntry] = []
    measurement_records: List[MeasurementRecord] = []

    for source_file in sources:
        print(f"Processing {source_file}...")
        lot, event, intensity = parse_filename_metadata(source_file.stem)
        file_link = build_file_link(source_file, destination_dir)
        file_display_name = source_file.name
        try:
            source_workbook = load_workbook(source_file, data_only=False, read_only=False)
        except Exception as exc:
            failures.append(f"{source_file}: {exc}")
            continue

        try:
            for sheet in source_workbook.worksheets:
                if not args.include_hidden and sheet.sheet_state != "visible":
                    skipped_hidden += 1
                    continue
                title = build_unique_sheet_title(source_file, sheet.title, used_titles)
                pass_count, fail_count = analyze_unit_results(sheet)
                target_sheet = workbook.create_sheet(title=title)
                target_sheet.sheet_state = sheet.sheet_state
                copy_sheet_contents(sheet, target_sheet, copy_formatting=not args.values_only)
                measurement_records.extend(
                    extract_pass_measurements(
                        sheet,
                        lot=lot,
                        event=event,
                        intensity=intensity,
                    )
                )
                summary_entries.append(
                    SummaryEntry(
                        sheet_title=title,
                        lot=lot,
                        event=event,
                        intensity=intensity,
                        file_name=file_display_name,
                        file_link=file_link,
                        pass_count=pass_count,
                        fail_count=fail_count,
                    )
                )
                copied_sheets += 1
        finally:
            source_workbook.close()

    if copied_sheets == 0:
        print("No worksheets were copied. Nothing to merge.", file=sys.stderr)
        if failures:
            for failure in failures:
                print(f"  {failure}", file=sys.stderr)
        return 1

    create_summary_sheet(workbook, summary_entries)
    create_measurements_sheet(workbook, measurement_records)

    try:
        destination_dir.mkdir(parents=True, exist_ok=True)
        workbook.save(destination_path)
    except PermissionError:
        print(f"Permission denied when writing {destination_path}. Close the file if it is open and retry.", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Failed to save workbook: {exc}", file=sys.stderr)
        return 1

    print(f"Merged {copied_sheets} worksheet(s) into {destination_path}")
    if skipped_hidden and not args.include_hidden:
        print(f"Skipped {skipped_hidden} hidden worksheet(s). Use --include-hidden to copy them.", file=sys.stderr)
    if failures:
        print("Some workbooks could not be processed:", file=sys.stderr)
        for failure in failures:
            print(f"  {failure}", file=sys.stderr)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
