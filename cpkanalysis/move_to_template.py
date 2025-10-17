"""Stage that copies CPK Report data into the template sheet.

The sheet provided by the template is expected to have a header row whose
labels match those found on the CPK Report sheet generated earlier in the
pipeline. Matching headers are copied column-by-column, bringing over values,
number formats, and hyperlinks while appending rows as needed.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Dict, Iterable, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

SOURCE_SHEET = "CPK Report"


def apply_template(workbook: Workbook, sheet_name: Optional[str] = None) -> str:
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Source sheet '{SOURCE_SHEET}' not found in workbook.")
    source_ws = workbook[SOURCE_SHEET]

    target_ws = _select_target_sheet(workbook, sheet_name)
    target_sheetname = target_ws.title

    source_headers, source_header_row = _header_map(source_ws)
    target_headers, target_header_row = _header_map(target_ws)
    shared_headers = [header for header in source_headers if header in target_headers]

    if not shared_headers:
        raise ValueError("No matching headers between CPK Report and target sheet.")

    max_row = source_ws.max_row
    for header in shared_headers:
        src_col = source_headers[header]
        tgt_col = target_headers[header]
        _copy_column(source_ws, target_ws, src_col, tgt_col, max_row, source_header_row, target_header_row)

    return target_sheetname


def run(workbook_path: Path, sheet_name: Optional[str] = None) -> str:
    """Copy CPK Report contents into the template sheet and return the target sheet name."""
    resolved = workbook_path.expanduser().resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"Workbook not found: {resolved}")

    workbook = load_workbook(resolved)
    try:
        target_sheetname = apply_template(workbook, sheet_name)
        workbook.save(resolved)
    finally:
        workbook.close()
    return target_sheetname


def _select_target_sheet(workbook, sheet_name: Optional[str]) -> Worksheet:
    # Excluded sheets that should never be used as template targets
    excluded_sheets = {"Summary", "Measurements", "CPK Report", "Test List and Limits"}

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Target sheet '{sheet_name}' not found in workbook.")
        if sheet_name in excluded_sheets:
            raise ValueError(
                f"Sheet '{sheet_name}' is a generated sheet, not a template sheet. "
                f"Please specify the actual template sheet name (look for 'Cpk Report' in row 1), "
                f"or leave blank for auto-detection."
            )
        return workbook[sheet_name]

    # Auto-detect template sheet by looking for "Cpk Report" in first row
    for candidate_name in workbook.sheetnames:
        if candidate_name == SOURCE_SHEET:
            continue
        sheet = workbook[candidate_name]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if any(
            isinstance(value, str) and value.strip().lower() == "cpk report"
            for value in first_row
        ):
            return sheet

    # Fallback: use first non-CPK Report sheet
    for candidate in workbook.sheetnames:
        if candidate != SOURCE_SHEET:
            return workbook[candidate]
    raise ValueError("No suitable target sheet found (workbook only contains CPK Report).")


def _header_map(sheet: Worksheet) -> tuple[Dict[str, int], int]:
    """Find headers in the sheet by searching the first several rows.
    
    Returns:
        tuple: (headers_dict, header_row_number)
    """
    best_headers: Dict[str, int] = {}
    header_row = 1
    
    # Search the first 20 rows for headers
    for row_num in range(1, min(21, sheet.max_row + 1)):
        row_headers = {}
        for cell in sheet[row_num]:
            value = str(cell.value).strip() if cell.value is not None else ""
            if value:
                row_headers[value] = cell.column
        
        # If this row has more potential headers than our current best, use it
        # This heuristic assumes the header row will have the most filled cells
        if len(row_headers) > len(best_headers):
            best_headers = row_headers
            header_row = row_num
    
    return best_headers, header_row


def _copy_column(
    source: Worksheet,
    target: Worksheet,
    source_column: int,
    target_column: int,
    source_max_row: int,
    source_header_row: int,
    target_header_row: int,
) -> None:
    data_rows = list(
        source.iter_rows(
            min_row=source_header_row + 1,
            max_row=source_max_row,
            min_col=source_column,
            max_col=source_column,
        )
    )
    if not data_rows:
        return

    start_row = target_header_row + 1

    # First pass: assign values
    for offset, (src_cell,) in enumerate(data_rows):
        tgt_cell = target.cell(row=start_row + offset, column=target_column)
        tgt_cell.value = src_cell.value

    # Second pass: reapply number formats and hyperlinks
    for offset, (src_cell,) in enumerate(data_rows):
        tgt_cell = target.cell(row=start_row + offset, column=target_column)
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format
        if src_cell.hyperlink:
            tgt_cell.hyperlink = copy(src_cell.hyperlink)
