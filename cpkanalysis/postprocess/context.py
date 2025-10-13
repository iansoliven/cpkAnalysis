"""Context management for post-processing workflow."""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import AnalysisInputs

__all__ = [
    "PostProcessContext",
    "load_metadata",
]


def load_metadata(path: Path) -> dict:
    """Return metadata JSON if present; otherwise an empty structure."""
    if not path.exists():
        return {}
    try:
        text = path.read_text(encoding="utf-8")
        return json.loads(text) if text else {}
    except Exception:
        return {}


@dataclass
class PostProcessContext:
    """Aggregate mutable state shared across post-processing actions."""

    analysis_inputs: AnalysisInputs
    workbook_path: Path
    metadata_path: Path
    metadata: Dict[str, Any] = field(default_factory=dict)

    _workbook: Workbook | None = field(default=None, init=False, repr=False)
    _summary_df: pd.DataFrame | None = field(default=None, init=False, repr=False)
    _limits_df: pd.DataFrame | None = field(default=None, init=False, repr=False)
    _measurements_df: pd.DataFrame | None = field(default=None, init=False, repr=False)
    _template_sheet_name: str | None = field(default=None, init=False, repr=False)

    dirty: bool = field(default=False, init=False)
    audit_log: List[Dict[str, Any]] = field(default_factory=list, init=False)

    def workbook(self) -> Workbook:
        """Return the openpyxl workbook, loading it on first access."""
        if self._workbook is None:
            self._workbook = load_workbook(self.workbook_path, data_only=False, read_only=False, keep_vba=False)
        return self._workbook

    # ------------------------------------------------------------------
    # Metadata helpers
    # ------------------------------------------------------------------

    def add_audit_entry(self, entry: Dict[str, Any]) -> None:
        """Record an audit entry for inclusion in metadata."""
        stamped = dict(entry)
        stamped.setdefault("timestamp", datetime.now(timezone.utc).isoformat())
        self.audit_log.append(stamped)

    def save(self) -> None:
        """Persist workbook changes and metadata audit log."""
        if self._workbook is not None:
            self._workbook.save(self.workbook_path)

        if self.audit_log:
            post_processing = self.metadata.setdefault("post_processing", {})
            runs: List[Dict[str, Any]] = post_processing.setdefault("runs", [])
            runs.extend(self.audit_log)
            self.audit_log.clear()

        self.metadata_path.parent.mkdir(parents=True, exist_ok=True)
        self.metadata_path.write_text(json.dumps(self.metadata, indent=2, sort_keys=False), encoding="utf-8")
        self.dirty = False

    def reload(self) -> None:
        """Reload workbook and cached DataFrames from disk."""
        if self._workbook is not None:
            try:
                self._workbook.close()
            except Exception:
                pass
        self._workbook = None
        self._summary_df = None
        self._limits_df = None
        self._measurements_df = None
        self._template_sheet_name = None
        self.dirty = False
        # Reload metadata in case it changed externally
        self.metadata = load_metadata(self.metadata_path)

    # ------------------------------------------------------------------
    # Workbook data access
    # ------------------------------------------------------------------

    def summary_frame(self, *, refresh: bool = False) -> pd.DataFrame:
        if refresh or self._summary_df is None:
            sheet = self._get_sheet("Summary")
            self._summary_df = _sheet_to_dataframe(sheet)
        return self._summary_df.copy()

    def limits_frame(self, *, refresh: bool = False) -> pd.DataFrame:
        if refresh or self._limits_df is None:
            sheet = self._get_sheet("Test List and Limits")
            self._limits_df = _sheet_to_dataframe(sheet)
        return self._limits_df.copy()

    def measurements_frame(self, *, refresh: bool = False) -> pd.DataFrame:
        if refresh or self._measurements_df is None:
            self._measurements_df = self._build_measurements_frame()
        return self._measurements_df.copy()

    def template_sheet(self) -> Worksheet:
        workbook = self.workbook()
        sheet_name = self.template_sheet_name()
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Template sheet '{sheet_name}' not found in workbook.")
        return workbook[sheet_name]

    def template_sheet_name(self) -> str:
        if self._template_sheet_name:
            return self._template_sheet_name

        candidate = (
            self.analysis_inputs.template_sheet
            or self.metadata.get("template_sheet")
            or self.metadata.get("templateSheet")
        )
        workbook = self.workbook()
        if candidate and candidate in workbook.sheetnames:
            self._template_sheet_name = candidate
            return candidate

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            if any(
                isinstance(value, str) and value.strip().lower() == "cpk report"
                for value in first_row
            ):
                self._template_sheet_name = sheet_name
                return sheet_name

        message = "Unable to determine template sheet: no sheet contains 'Cpk Report' in the top row."
        print(message, file=sys.stderr)
        try:
            input("Press Enter to exit...")
        except EOFError:
            pass
        raise SystemExit(1)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def mark_dirty(self) -> None:
        self.dirty = True

    def invalidate_frames(self, *names: str) -> None:
        """Invalidate cached DataFrames by name."""
        if not names or "summary" in names:
            self._summary_df = None
        if not names or "limits" in names:
            self._limits_df = None
        if not names or "measurements" in names:
            self._measurements_df = None

    def _get_sheet(self, name: str) -> Worksheet:
        workbook = self.workbook()
        if name not in workbook.sheetnames:
            raise ValueError(f"Expected sheet '{name}' was not found in workbook {self.workbook_path}.")
        return workbook[name]

    def _build_measurements_frame(self) -> pd.DataFrame:
        workbook = self.workbook()
        frames: List[pd.DataFrame] = []
        for sheet_name in workbook.sheetnames:
            if not sheet_name.startswith("Measurements"):
                continue
            ws = workbook[sheet_name]
            df = _sheet_to_dataframe(ws)
            if not df.empty:
                frames.append(df)
        if not frames:
            return pd.DataFrame()
        combined = pd.concat(frames, ignore_index=True)
        return combined


# ----------------------------------------------------------------------
# Utility functions
# ----------------------------------------------------------------------

def _sheet_to_dataframe(sheet: Worksheet) -> pd.DataFrame:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = _normalise_header_row(rows[0])
    data = [row for row in rows[1:] if any(cell is not None for cell in row)]
    if not data:
        return pd.DataFrame(columns=header)
    trimmed = [
        list(row[: len(header)]) + [None] * max(0, len(header) - len(row))
        for row in data
    ]
    return pd.DataFrame(trimmed, columns=header)


def _normalise_header_row(row: Sequence[Any]) -> List[str]:
    header = []
    for cell in row:
        if cell is None:
            header.append("")
            continue
        header.append(str(cell).strip())
    # Remove trailing empty headers
    while header and not header[-1]:
        header.pop()
    return header
