"""Action implementations for post-processing menu."""

from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
import logging

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.worksheet import Worksheet

from .context import PostProcessContext
from .io_adapters import PostProcessIO
from . import sheet_utils
from . import charts
from . import proposed_limits_grr

__all__ = [
    "ActionCancelled",
    "update_stdf_limits",
    "apply_spec_limits",
    "calculate_proposed_limits",
    "calculate_proposed_limits_grr",
]

PROPOSAL_TOLERANCE = 1e-6
GRR_REFERENCE_SHEET = "_GRR_reference"
SPEC_LEGEND_TEXT = "Light yellow highlight indicates Proposed Spec limit differs from Total_GRR spec."
GUARD_LEGEND_TEXT = (
    "Red guardband percent indicates guardband width is below 50% of required GRR guardband."
)
SPEC_LEGEND_COLOR = "FFFFF2CC"
GUARD_LEGEND_COLOR = "FFFFC7CE"
SPEC_LEGEND_FILL = PatternFill(start_color=SPEC_LEGEND_COLOR, end_color=SPEC_LEGEND_COLOR, fill_type="solid")
GUARD_LEGEND_FILL = PatternFill(start_color=GUARD_LEGEND_COLOR, end_color=GUARD_LEGEND_COLOR, fill_type="solid")
LEGEND_FONT = Font(bold=True)
SHEET_VISIBLE = "visible"
SHEET_HIDDEN = "hidden"

logger = logging.getLogger(__name__)


def _reapply_cpk_formulas(context: PostProcessContext) -> None:
    """Ensure template CPK columns remain formula-driven after updates."""
    try:
        template_ws = context.template_sheet()
        from ..tools import update_cpk_formulas

        update_cpk_formulas.apply_formulas(template_ws.parent, template_ws)
    except ValueError as exc:
        logger.warning("Skipping CPK formula refresh: %s", exc)


def _ensure_grr_reference_sheet(workbook, grr_table) -> Optional[Worksheet]:
    try:
        records = list(grr_table.records())
    except AttributeError:
        records = []

    if not records:
        if GRR_REFERENCE_SHEET in workbook.sheetnames:
            workbook.remove(workbook[GRR_REFERENCE_SHEET])
        return None

    if GRR_REFERENCE_SHEET in workbook.sheetnames:
        ref_ws = workbook[GRR_REFERENCE_SHEET]
        ref_ws.sheet_state = SHEET_VISIBLE
        if ref_ws.max_row:
            ref_ws.delete_rows(1, ref_ws.max_row)
    else:
        ref_ws = workbook.create_sheet(GRR_REFERENCE_SHEET)

    ref_ws.append(
        [
            "Key",
            "Test Name",
            "Test Number",
            "Spec Lower",
            "Spec Upper",
            "Guardband Requirement",
        ]
    )
    seen: set[str] = set()
    for record in records:
        name = record.test_name or ""
        number = record.test_number or ""
        key = f"{name}|{number}".strip("|")
        if not key:
            continue
        if key in seen:
            continue
        seen.add(key)
        ref_ws.append(
            [
                key,
                record.test_name,
                record.test_number,
                record.spec_lower,
                record.spec_upper,
                record.guardband_full,
            ]
        )
    ref_ws.sheet_state = SHEET_HIDDEN
    return ref_ws


def _clear_conditional_formatting_range(ws, range_str: str) -> None:
    cf = ws.conditional_formatting
    try:
        cf.remove(range_str)  # type: ignore[attr-defined]
        return
    except (AttributeError, KeyError):
        pass
    rules = getattr(cf, "_cf_rules", None)
    if not rules:
        return
    if isinstance(rules, dict):
        keys_to_delete = [key for key in list(rules.keys()) if str(key) == range_str]
        for key in keys_to_delete:
            del rules[key]
    else:
        filtered = [rule for rule in rules if str(getattr(rule, "sqref", "")) != range_str]
        if len(filtered) != len(rules):
            setattr(cf, "_cf_rules", filtered)


def _write_fixed_legends(template_ws, spec_lower_col: int | None) -> None:
    if spec_lower_col is None:
        return

    def _set(row: int, text: str, fill: PatternFill) -> None:
        cell = template_ws.cell(row=row, column=spec_lower_col)
        if cell.value != text:
            cell.value = text
        cell.font = LEGEND_FONT
        cell.fill = fill

    _set(1, GUARD_LEGEND_TEXT, GUARD_LEGEND_FILL)
    _set(2, SPEC_LEGEND_TEXT, SPEC_LEGEND_FILL)


def _apply_spec_difference_cf(template_ws, header_map: Dict[str, int], header_row: int) -> None:
    workbook = template_ws.parent
    if GRR_REFERENCE_SHEET not in workbook.sheetnames:
        return

    spec_lower_col = header_map.get(sheet_utils.normalize_header("Proposed Spec Lower"))
    spec_upper_col = header_map.get(sheet_utils.normalize_header("Proposed Spec Upper"))
    if spec_lower_col is None or spec_upper_col is None:
        return

    _write_fixed_legends(template_ws, spec_lower_col)

    data_start = header_row + 1
    max_row = template_ws.max_row
    if data_start > max_row:
        return

    ref_ws = workbook[GRR_REFERENCE_SHEET]
    ref_header_row, ref_headers = sheet_utils.build_header_map(ref_ws)
    key_col = ref_headers.get(sheet_utils.normalize_header("Key"))
    name_col = ref_headers.get(sheet_utils.normalize_header("Test Name"))
    number_col = ref_headers.get(sheet_utils.normalize_header("Test Number"))
    lower_col = ref_headers.get(sheet_utils.normalize_header("Spec Lower"))
    upper_col = ref_headers.get(sheet_utils.normalize_header("Spec Upper"))
    if not all([key_col, name_col, number_col, lower_col, upper_col]):
        return

    ref_start_row = ref_header_row + 1
    if ref_start_row > ref_ws.max_row:
        return

    def _col_range(col_idx: int) -> str:
        letter = get_column_letter(col_idx)
        return f"'{GRR_REFERENCE_SHEET}'!${letter}${ref_start_row}:${letter}${ref_ws.max_row}"

    key_range = _col_range(key_col)
    lower_ref_range = _col_range(lower_col)
    upper_ref_range = _col_range(upper_col)
    # Build name/number ranges for INDEX/MATCH guardband lookups.
    test_name_col = _resolve_column(
        header_map,
        ["Test Name", "Test", "TEST NAME", "Test_Name"],
    )
    test_number_col = _resolve_column(
        header_map,
        ["Test Number", "Test Num", "Number", "TEST NUM", "Test_Num"],
    )
    if test_name_col is None or test_number_col is None:
        return

    name_letter = get_column_letter(test_name_col)
    number_letter = get_column_letter(test_number_col)

    lower_letter = get_column_letter(spec_lower_col)
    upper_letter = get_column_letter(spec_upper_col)
    lower_range = f"{lower_letter}{data_start}:{lower_letter}{max_row}"
    upper_range = f"{upper_letter}{data_start}:{upper_letter}{max_row}"
    key_expr = f"${name_letter}{data_start}&\"|\"&${number_letter}{data_start}"

    base_guard = "1E-6"
    highlight_fill = SPEC_LEGEND_FILL

    lower_formula = (
        f"IF(OR(${lower_letter}{data_start}=\"\",${number_letter}{data_start}=\"\",${name_letter}{data_start}=\"\"),FALSE,"
        f"IF(COUNTIF({key_range},{key_expr})=0,FALSE,"
        f"ABS(${lower_letter}{data_start}-SUMIFS({lower_ref_range},{key_range},{key_expr}))>{base_guard}))"
    )
    upper_formula = (
        f"IF(OR(${upper_letter}{data_start}=\"\",${number_letter}{data_start}=\"\",${name_letter}{data_start}=\"\"),FALSE,"
        f"IF(COUNTIF({key_range},{key_expr})=0,FALSE,"
        f"ABS(${upper_letter}{data_start}-SUMIFS({upper_ref_range},{key_range},{key_expr}))>{base_guard}))"
    )

    for target_range in (lower_range, upper_range):
        _clear_conditional_formatting_range(template_ws, target_range)

    template_ws.conditional_formatting.add(
        lower_range, FormulaRule(formula=[lower_formula], fill=highlight_fill)
    )
    template_ws.conditional_formatting.add(
        upper_range, FormulaRule(formula=[upper_formula], fill=highlight_fill)
    )


def _should_skip_grr(
    descriptor: TestDescriptor,
    template_ws,
    template_rows: List[int],
    template_headers: Dict[str, int],
    *,
    spec_lower: float,
    spec_upper: float,
    min_cpk: float,
    max_cpk: float,
    guardband_full: float,
) -> tuple[bool, Dict[str, float | None | str]]:
    info: Dict[str, float | None | str] = {}
    if not template_rows:
        return False, info

    info["spec_lower"] = spec_lower
    info["spec_upper"] = spec_upper

    if guardband_full <= PROPOSAL_TOLERANCE:
        info["reason"] = "invalid_guardband_requirement"
        return False, info

    if spec_upper <= spec_lower:
        info["reason"] = "invalid_spec_range"
        return False, info

    spec_width = spec_upper - spec_lower
    if spec_width <= PROPOSAL_TOLERANCE:
        info["reason"] = "insufficient_spec_width"
        return False, info

    ll_ate_col = _resolve_column(template_headers, ["LL_ATE", "LL ATE", "Lower ATE", "LL"])
    ul_ate_col = _resolve_column(template_headers, ["UL_ATE", "UL ATE", "Upper ATE", "UL"])
    ll_ate = _read_first_from_rows(template_ws, template_rows, ll_ate_col)
    ul_ate = _read_first_from_rows(template_ws, template_rows, ul_ate_col)
    info["ll_ate"] = ll_ate
    info["ul_ate"] = ul_ate

    if ll_ate is None or ul_ate is None:
        return False, info

    lower_guardband_width = abs(ll_ate - spec_lower)
    upper_guardband_width = abs(spec_upper - ul_ate)
    info["lower_guardband_width"] = lower_guardband_width
    info["upper_guardband_width"] = upper_guardband_width
    info["required_guardband"] = guardband_full

    if spec_width > PROPOSAL_TOLERANCE:
        info["lower_guardband_percent"] = lower_guardband_width / spec_width
        info["upper_guardband_percent"] = upper_guardband_width / spec_width

    lower_met = lower_guardband_width + PROPOSAL_TOLERANCE >= guardband_full
    upper_met = upper_guardband_width + PROPOSAL_TOLERANCE >= guardband_full
    if not (lower_met and upper_met):
        info["reason"] = "guardband_insufficient"
        return False, info

    current_cpk = descriptor.cpk
    if current_cpk is None:
        return False, info
    if current_cpk < min_cpk - PROPOSAL_TOLERANCE or current_cpk > max_cpk + PROPOSAL_TOLERANCE:
        return False, info

    info["reason"] = "guardband_satisfied"
    return True, info

class ActionCancelled(RuntimeError):
    """Raised when the user aborts an action."""


@dataclass(frozen=True)
class TestDescriptor:
    file: str
    test_name: str
    test_number: str
    unit: str
    mean: float | None
    stdev: float | None
    cpk: float | None

    def key(self) -> str:
        return f"{self.file}|{self.test_name}|{self.test_number}"

    def label(self) -> str:
        suffix = f" (Test {self.test_number})" if self.test_number else ""
        file_part = f" [{self.file}]" if self.file else ""
        return f"{self.test_name}{suffix}{file_part}"


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _build_test_descriptors(context: PostProcessContext) -> List[TestDescriptor]:
    summary = context.summary_frame()
    if summary.empty:
        return []
    descriptors: List[TestDescriptor] = []
    for _, row in summary.iterrows():
        descriptors.append(
            TestDescriptor(
                file=_safe_str(row.get("File")),
                test_name=_safe_str(row.get("Test Name")),
                test_number=_safe_str(row.get("Test Number")),
                unit=_safe_str(row.get("Unit")),
                mean=_safe_float(row.get("MEAN")),
                stdev=_safe_float(row.get("STDEV")),
                cpk=_safe_float(row.get("CPK")),
            )
        )
    return descriptors


def _safe_float(value) -> float | None:
    if value in (None, "", "nan"):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(result):
        return None
    return result


def _prompt_scope(
    io: PostProcessIO,
    params: Optional[dict],
    *,
    allow_single: bool = True,
    require_target: bool = False,
    default_target: float | None = None,
    prompt_target: bool = True,
) -> dict:
    params_dict: Dict[str, object] = dict(params or {})

    scope = params_dict.get("scope")
    if scope not in {"all", "single"}:
        options = ["All tests"]
        if allow_single:
            options.append("Single test")
        choice = io.prompt_choice("Select scope for this action:", options)
        scope = "all" if choice == 0 else "single"

    def _parse_target(value) -> tuple[float | None, str | None]:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None, "Target CPK must be numeric."
        if not math.isfinite(number) or number <= 0:
            return None, "Target CPK must be positive."
        return number, None

    target_cpk = None
    if "target_cpk" in params_dict:
        raw_target = params_dict.get("target_cpk")
        target_cpk, error = _parse_target(raw_target)
        if error:
            io.warn(error)

    needs_prompt = False
    if require_target and target_cpk is None:
        needs_prompt = True
    elif prompt_target and target_cpk is None:
        needs_prompt = True

    if needs_prompt:
        prompt_message = (
            "Enter target CPK (blank to keep existing limits):"
            if not require_target
            else "Enter target CPK (must be greater than zero):"
        )
        default_candidate: float | None = target_cpk if target_cpk is not None else default_target
        default_text = "" if default_candidate is None else str(default_candidate)

        while True:
            response = io.prompt(prompt_message, default=default_text) or ""
            value = response.strip()
            if not value:
                if require_target:
                    io.warn("Target CPK is required.")
                    continue
                target_cpk = None
                break
            parsed, error = _parse_target(value)
            if error:
                io.warn(error)
                continue
            target_cpk = parsed
            break

    if require_target and target_cpk is None:
        raise ActionCancelled("Target CPK is required for this action.")

    selection = params_dict.get("test_key")

    return {"scope": scope, "target_cpk": target_cpk, "test_key": selection}


def _resolve_tests(
    descriptors: Sequence[TestDescriptor],
    context: PostProcessContext,
    io: PostProcessIO,
    resolved: dict,
) -> List[TestDescriptor]:
    if resolved["scope"] == "all":
        return list(descriptors)
    if not descriptors:
        raise ActionCancelled("No tests available.")
    test_map = {descriptor.key(): descriptor for descriptor in descriptors}
    if resolved.get("test_key") in test_map:
        return [test_map[resolved["test_key"]]]

    io.info("Select the test you want to update:")
    labels = [desc.label() for desc in descriptors]
    choice = io.prompt_choice("Test selection:", labels)
    descriptor = descriptors[choice]
    resolved["test_key"] = descriptor.key()
    return [descriptor]


def _summaries_by_key(summary: pd.DataFrame) -> Dict[str, pd.Series]:
    mapping: Dict[str, pd.Series] = {}
    for _, row in summary.iterrows():
        key = "|".join(
            [
                _safe_str(row.get("File")),
                _safe_str(row.get("Test Name")),
                _safe_str(row.get("Test Number")),
            ]
        )
        mapping[key] = row
    return mapping


def _warn_if_missing(io: PostProcessIO, warnings: List[str], message: str) -> None:
    warnings.append(message)
    io.warn(message)


def _resolve_column(header_map: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        normalized = sheet_utils.normalize_header(alias)
        if normalized in header_map:
            return header_map[normalized]
    return None


def _tests_to_strings(tests: Iterable[TestDescriptor]) -> List[str]:
    return [test.key() for test in tests]


def _safe_str(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value)
    return text.strip() if text is not None else ""


def _first_not_none(*values):
    for value in values:
        if value is not None:
            return value
    return None


def _read_first_from_rows(sheet, rows: List[int], column: int | None) -> float | None:
    if column is None:
        return None
    for row_idx in rows:
        value = _safe_float(sheet_utils.get_cell(sheet, row_idx, column))
        if value is not None:
            return value
    return None


# ---------------------------------------------------------------------------
# Proposed limit helpers
# ---------------------------------------------------------------------------


def _almost_equal(a: float | None, b: float | None) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) <= PROPOSAL_TOLERANCE


def _ensure_proposal_state(metadata: Dict[str, object]) -> Dict[str, Dict[str, float | None]]:
    state = metadata.setdefault("post_processing_state", {})
    if not isinstance(state, dict):
        state = {}
        metadata["post_processing_state"] = state
    proposals = state.setdefault("proposed_limits", {})
    if not isinstance(proposals, dict):
        proposals = {}
        state["proposed_limits"] = proposals
    return proposals


def _ensure_proposed_spec_columns(
    template_ws,
    header_row: int,
    header_map: Dict[str, int],
) -> dict[str, int]:
    from ..tools import update_cpk_formulas

    update_cpk_formulas.apply_formulas(template_ws.parent, template_ws)

    updated_header_row, updated_header_map = sheet_utils.build_header_map(template_ws)
    header_map.clear()
    header_map.update(updated_header_map)
    header_row = updated_header_row

    spec_lower_col = _resolve_column(updated_header_map, ["Proposed Spec Lower"])
    spec_upper_col = _resolve_column(updated_header_map, ["Proposed Spec Upper"])
    spec_cpk_col = _resolve_column(updated_header_map, ["CPK Proposed Spec"])
    lower_pct_col = _resolve_column(updated_header_map, ["Lower Guardband Percent"])
    upper_pct_col = _resolve_column(updated_header_map, ["Upper Guardband Percent"])
    lower_guardband_col = _resolve_column(updated_header_map, ["Lower Guardband"])
    upper_guardband_col = _resolve_column(updated_header_map, ["Upper Guardband"])

    if spec_lower_col is None or spec_upper_col is None or spec_cpk_col is None:
        raise ActionCancelled("Template sheet missing Proposed Spec columns after formula injection.")

    return {
        "spec_lower": spec_lower_col,
        "spec_upper": spec_upper_col,
        "spec_cpk": spec_cpk_col,
        "lower_pct": lower_pct_col,
        "upper_pct": upper_pct_col,
        "lower_guardband": lower_guardband_col,
        "upper_guardband": upper_guardband_col,
        "header_row": header_row,
    }


def _resolve_unit_override(
    io: PostProcessIO,
    cache: Dict[Tuple[str, str], bool],
    descriptor_label: str,
    workbook_unit: str,
    grr_unit: str,
) -> bool:
    key = (workbook_unit or "", grr_unit or "")
    if key in cache:
        return cache[key]
    message = (
        f"Unit mismatch for {descriptor_label}: workbook '{workbook_unit or 'n/a'}' "
        f"vs GRR '{grr_unit or 'n/a'}'. Proceed using workbook unit?"
    )
    decision = io.confirm(message, default=False)
    cache[key] = decision
    return decision


def _coerce_positive_float(value) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number <= 0 or math.isnan(number) or math.isinf(number):
        return None
    return number


# ---------------------------------------------------------------------------
# Action implementations
# ---------------------------------------------------------------------------


def update_stdf_limits(context: PostProcessContext, io: PostProcessIO, params: Optional[dict]) -> dict:
    descriptors = _build_test_descriptors(context)
    if not descriptors:
        raise ActionCancelled("Summary sheet is empty.")

    resolved = _prompt_scope(io, params, allow_single=True, require_target=False, prompt_target=False)
    selected_tests = _resolve_tests(descriptors, context, io, resolved)

    summary_df = context.summary_frame()
    summary_lookup = _summaries_by_key(summary_df)

    template_ws = context.template_sheet()
    template_header_row, template_headers = sheet_utils.build_header_map(template_ws)
    ll_column = _resolve_column(template_headers, ["LL_ATE", "LL ATE", "Lower ATE", "LL"])
    ul_column = _resolve_column(template_headers, ["UL_ATE", "UL ATE", "Upper ATE", "UL"])
    if ll_column is None or ul_column is None:
        raise ActionCancelled("Template sheet missing LL_ATE or UL_ATE columns.")

    limits_ws = context.workbook()["Test List and Limits"]
    limits_header_row, limits_headers = sheet_utils.build_header_map(limits_ws)
    stdf_lower_col = _resolve_column(limits_headers, ["STDF Lower Limit", "STDF Lower"])
    stdf_upper_col = _resolve_column(limits_headers, ["STDF Upper Limit", "STDF Upper"])

    warnings: List[str] = []
    updated_tests: List[TestDescriptor] = []
    for descriptor in selected_tests:
        row = summary_lookup.get(descriptor.key())
        if row is None:
            _warn_if_missing(
                io,
                warnings,
                f"Summary entry not found for test {descriptor.label()} – skipping.",
            )
            continue

        mean = _safe_float(row.get("MEAN"))
        stdev = _safe_float(row.get("STDEV"))
        if stdev is None or stdev <= 0:
            _warn_if_missing(
                io,
                warnings,
                f"Cannot compute limits for {descriptor.label()} (non-positive STDEV).",
            )
            continue

        lower_limit = _first_not_none(_safe_float(row.get("LL_2CPK")), _safe_float(row.get("LL_3IQR")))
        upper_limit = _first_not_none(_safe_float(row.get("UL_2CPK")), _safe_float(row.get("UL_3IQR")))

        if lower_limit is None and upper_limit is None:
            _warn_if_missing(
                io,
                warnings,
                f"Unable to determine limits for {descriptor.label()} – skipping.",
            )
            continue

        template_rows = sheet_utils.find_rows_by_test(
            template_ws,
            template_header_row,
            template_headers,
            test_name=descriptor.test_name,
            test_number=descriptor.test_number,
        )
        if not template_rows:
            _warn_if_missing(
                io,
                warnings,
                f"Template row not found for {descriptor.label()} – skipping.",
            )
            continue

        for row_idx in template_rows:
            if lower_limit is not None:
                sheet_utils.set_cell(template_ws, row_idx, ll_column, lower_limit)
            if upper_limit is not None:
                sheet_utils.set_cell(template_ws, row_idx, ul_column, upper_limit)

        if stdf_lower_col is not None or stdf_upper_col is not None:
            limit_rows = sheet_utils.find_rows_by_test(
                limits_ws,
                limits_header_row,
                limits_headers,
                test_name=descriptor.test_name,
                test_number=descriptor.test_number,
            )
            for row_idx in limit_rows or []:
                if stdf_lower_col is not None and lower_limit is not None:
                    sheet_utils.set_cell(limits_ws, row_idx, stdf_lower_col, lower_limit)
                if stdf_upper_col is not None and upper_limit is not None:
                    sheet_utils.set_cell(limits_ws, row_idx, stdf_upper_col, upper_limit)

        updated_tests.append(descriptor)

    if not updated_tests:
        raise ActionCancelled("No tests updated.")

    context.invalidate_frames("limits")
    charts.refresh_tests(context, updated_tests)
    _reapply_cpk_formulas(context)

    summary_text = f"Updated STDF limits for {len(updated_tests)} test(s)."
    return {
        "summary": summary_text,
        "warnings": warnings,
        "audit": {
            "scope": resolved["scope"],
            "tests": _tests_to_strings(updated_tests),
        },
        "replay_params": resolved,
        "mark_dirty": True,
    }


def apply_spec_limits(context: PostProcessContext, io: PostProcessIO, params: Optional[dict]) -> dict:
    descriptors = _build_test_descriptors(context)
    if not descriptors:
        raise ActionCancelled("Summary sheet is empty.")

    resolved = _prompt_scope(io, params, allow_single=True, require_target=False)
    selected_tests = _resolve_tests(descriptors, context, io, resolved)

    template_ws = context.template_sheet()
    template_header_row, template_headers = sheet_utils.build_header_map(template_ws)
    spec_lower_col = _resolve_column(template_headers, ["Spec Lower", "Spec LL", "Spec Lower Limit"])
    spec_upper_col = _resolve_column(template_headers, ["Spec Upper", "Spec UL", "Spec Upper Limit"])
    what_lower_col = _resolve_column(template_headers, ["What-if Lower", "What If Lower", "What-If Lower"])
    what_upper_col = _resolve_column(template_headers, ["What-if Upper", "What If Upper", "What-If Upper"])

    limits_ws = context.workbook()["Test List and Limits"]
    limits_header_row, limits_headers = sheet_utils.build_header_map(limits_ws)
    limits_spec_lower_col = _resolve_column(limits_headers, ["Spec Lower Limit", "Spec Lower"])
    limits_spec_upper_col = _resolve_column(limits_headers, ["Spec Upper Limit", "Spec Upper"])
    limits_what_lower_col = _resolve_column(limits_headers, ["User What-If Lower Limit", "What-if Lower"])
    limits_what_upper_col = _resolve_column(limits_headers, ["User What-If Upper Limit", "What-if Upper"])

    if not any([spec_lower_col, spec_upper_col, what_lower_col, what_upper_col]):
        raise ActionCancelled("Template sheet missing Spec / What-If columns.")

    summary_df = context.summary_frame()
    summary_lookup = _summaries_by_key(summary_df)

    warnings: List[str] = []
    updated_tests: List[TestDescriptor] = []

    for descriptor in selected_tests:
        template_rows = sheet_utils.find_rows_by_test(
            template_ws,
            template_header_row,
            template_headers,
            test_name=descriptor.test_name,
            test_number=descriptor.test_number,
        )
        if not template_rows:
            _warn_if_missing(io, warnings, f"Template row not found for {descriptor.label()} – skipping.")
            continue

        row = summary_lookup.get(descriptor.key())
        mean = _safe_float(row.get("MEAN")) if row is not None else None
        stdev = _safe_float(row.get("STDEV")) if row is not None else None

        spec_lower = spec_upper = what_lower = what_upper = None
        if resolved.get("target_cpk"):
            target = resolved["target_cpk"]
            if target and target > 0 and stdev and stdev > 0 and mean is not None:
                width = 3.0 * target * stdev
                spec_lower = mean - width
                spec_upper = mean + width
                what_lower = spec_lower
                what_upper = spec_upper
            else:
                _warn_if_missing(io, warnings, f"Cannot compute target-based limits for {descriptor.label()}.")

        for row_idx in template_rows:
            if spec_lower_col:
                spec_lower = spec_lower if spec_lower is not None else _safe_float(
                    sheet_utils.get_cell(template_ws, row_idx, spec_lower_col)
                )
            if spec_upper_col:
                spec_upper = spec_upper if spec_upper is not None else _safe_float(
                    sheet_utils.get_cell(template_ws, row_idx, spec_upper_col)
                )
            if what_lower_col:
                what_lower = what_lower if what_lower is not None else _safe_float(
                    sheet_utils.get_cell(template_ws, row_idx, what_lower_col)
                )
            if what_upper_col:
                what_upper = what_upper if what_upper is not None else _safe_float(
                    sheet_utils.get_cell(template_ws, row_idx, what_upper_col)
                )

        def _write_if_available(column, value):
            if column and value is not None:
                for row_idx in template_rows:
                    sheet_utils.set_cell(template_ws, row_idx, column, value)

        _write_if_available(spec_lower_col, spec_lower)
        _write_if_available(spec_upper_col, spec_upper)
        _write_if_available(what_lower_col, what_lower)
        _write_if_available(what_upper_col, what_upper)

        limit_rows = sheet_utils.find_rows_by_test(
            limits_ws,
            limits_header_row,
            limits_headers,
            test_name=descriptor.test_name,
            test_number=descriptor.test_number,
        )
        for row_idx in limit_rows or []:
            if limits_spec_lower_col and spec_lower is not None:
                sheet_utils.set_cell(limits_ws, row_idx, limits_spec_lower_col, spec_lower)
            if limits_spec_upper_col and spec_upper is not None:
                sheet_utils.set_cell(limits_ws, row_idx, limits_spec_upper_col, spec_upper)
            if limits_what_lower_col and what_lower is not None:
                sheet_utils.set_cell(limits_ws, row_idx, limits_what_lower_col, what_lower)
            if limits_what_upper_col and what_upper is not None:
                sheet_utils.set_cell(limits_ws, row_idx, limits_what_upper_col, what_upper)

        updated_tests.append(descriptor)

    if not updated_tests:
        raise ActionCancelled("No tests updated.")

    context.invalidate_frames("limits")
    charts.refresh_tests(context, updated_tests, include_spec=True)
    _reapply_cpk_formulas(context)

    summary_text = f"Applied Spec/What-If limits for {len(updated_tests)} test(s)."
    return {
        "summary": summary_text,
        "warnings": warnings,
        "audit": {
            "scope": resolved["scope"],
            "tests": _tests_to_strings(updated_tests),
            "parameters": {"target_cpk": resolved.get("target_cpk")},
        },
        "replay_params": resolved,
        "mark_dirty": True,
    }


def calculate_proposed_limits(context: PostProcessContext, io: PostProcessIO, params: Optional[dict]) -> dict:
    descriptors = _build_test_descriptors(context)
    if not descriptors:
        raise ActionCancelled("Summary sheet is empty.")

    resolved = _prompt_scope(io, params, allow_single=True, require_target=True)
    selected_tests = _resolve_tests(descriptors, context, io, resolved)
    target_cpk = resolved.get("target_cpk") or 0.0

    template_ws = context.template_sheet()
    template_header_row, template_headers = sheet_utils.build_header_map(template_ws)
    ll_prop_col = _resolve_column(template_headers, ["LL_PROP", "LL Proposal", "LL Proposed"])
    ul_prop_col = _resolve_column(template_headers, ["UL_PROP", "UL Proposal", "UL Proposed"])
    cpk_prop_col = _resolve_column(template_headers, ["CPK_PROP", "CPK Proposed"])
    yld_prop_col = _resolve_column(template_headers, ["%YLD LOSS_PROP", "%YLD Loss Proposed"])

    if not all([ll_prop_col, ul_prop_col, cpk_prop_col, yld_prop_col]):
        missing = []
        if not ll_prop_col:
            missing.append("LL_PROP")
        if not ul_prop_col:
            missing.append("UL_PROP")
        if not cpk_prop_col:
            missing.append("CPK_PROP")
        if not yld_prop_col:
            missing.append("%YLD LOSS_PROP")
        raise ActionCancelled(
            f"Template sheet missing proposed-limit columns: {', '.join(missing)}. "
            "Please add these columns to your template sheet to use this feature."
        )

    summary_df = context.summary_frame()
    summary_lookup = _summaries_by_key(summary_df)
    measurements_df = context.measurements_frame()
    metadata_proposals = _ensure_proposal_state(context.metadata)
    timestamp = datetime.now(timezone.utc).isoformat()

    warnings: List[str] = []
    updated_tests: List[TestDescriptor] = []
    metrics_updates = 0
    audit_details: List[dict] = []
    mark_dirty = False

    def _read_first(rows: List[int], column: int | None) -> float | None:
        if column is None:
            return None
        for row_idx in rows:
            value = _safe_float(sheet_utils.get_cell(template_ws, row_idx, column))
            if value is not None:
                return value
        return None

    for descriptor in selected_tests:
        row = summary_lookup.get(descriptor.key())
        if row is None:
            _warn_if_missing(io, warnings, f"Summary entry not found for {descriptor.label()} – skipping.")
            continue

        mean = _safe_float(row.get("MEAN"))
        stdev = _safe_float(row.get("STDEV"))
        width = None
        if mean is not None and stdev is not None and stdev > 0:
            width = 3.0 * target_cpk * stdev

        template_rows = sheet_utils.find_rows_by_test(
            template_ws,
            template_header_row,
            template_headers,
            test_name=descriptor.test_name,
            test_number=descriptor.test_number,
        )
        if not template_rows:
            _warn_if_missing(io, warnings, f"Template row not found for {descriptor.label()} – skipping.")
            continue

        prior_entry = metadata_proposals.get(descriptor.key(), {})
        prev_lower = _safe_float(prior_entry.get("ll")) if isinstance(prior_entry, dict) else None
        prev_upper = _safe_float(prior_entry.get("ul")) if isinstance(prior_entry, dict) else None

        existing_lower = _read_first(template_rows, ll_prop_col)
        existing_upper = _read_first(template_rows, ul_prop_col)
        current_cpk = _read_first(template_rows, cpk_prop_col)
        current_yield = _read_first(template_rows, yld_prop_col)

        lower_origin = "unchanged"
        upper_origin = "unchanged"
        proposal_changed = False

        final_lower = existing_lower
        if existing_lower is not None:
            if prev_lower is None or not _almost_equal(existing_lower, prev_lower):
                lower_origin = "user"
                proposal_changed = True
        else:
            if width is None or mean is None:
                _warn_if_missing(io, warnings, f"Cannot compute lower proposal for {descriptor.label()} – missing stats.")
            else:
                final_lower = mean - width
                lower_origin = "computed"
                proposal_changed = True

        final_upper = existing_upper
        if existing_upper is not None:
            if prev_upper is None or not _almost_equal(existing_upper, prev_upper):
                upper_origin = "user"
                proposal_changed = True
        else:
            if width is None or mean is None:
                _warn_if_missing(io, warnings, f"Cannot compute upper proposal for {descriptor.label()} – missing stats.")
            else:
                final_upper = mean + width
                upper_origin = "computed"
                proposal_changed = True

        if final_lower is None and final_upper is None:
            _warn_if_missing(io, warnings, f"No proposed limits determined for {descriptor.label()} – skipping.")
            continue

        metrics_blank = current_cpk is None or current_yield is None
        if not proposal_changed and not metrics_blank:
            continue

        for row_idx in template_rows:
            if final_lower is not None and lower_origin in {"user", "computed"}:
                sheet_utils.set_cell(template_ws, row_idx, ll_prop_col, final_lower)
                mark_dirty = True
            if final_upper is not None and upper_origin in {"user", "computed"}:
                sheet_utils.set_cell(template_ws, row_idx, ul_prop_col, final_upper)
                mark_dirty = True

        cpk_updated = False
        yield_updated = False

        if proposal_changed or current_cpk is None:
            if stdev is None or stdev <= 0 or mean is None:
                _warn_if_missing(io, warnings, f"Cannot compute CPK for {descriptor.label()} – missing stats.")
            else:
                candidates: List[float] = []
                if final_upper is not None:
                    candidates.append((final_upper - mean) / (3.0 * stdev))
                if final_lower is not None:
                    candidates.append((mean - final_lower) / (3.0 * stdev))
                new_cpk_value = min(candidates) if candidates else None
                if new_cpk_value is not None and not np.isfinite(new_cpk_value):
                    new_cpk_value = None
                if new_cpk_value is not None:
                    if current_cpk is None or not _almost_equal(new_cpk_value, current_cpk):
                        for row_idx in template_rows:
                            sheet_utils.set_cell(template_ws, row_idx, cpk_prop_col, new_cpk_value)
                        cpk_updated = True
                        mark_dirty = True
                elif current_cpk is None:
                    pass  # Nothing to write

        if proposal_changed or current_yield is None:
            if measurements_df.empty:
                _warn_if_missing(io, warnings, f"No measurements available to compute yield for {descriptor.label()}.")
            else:
                new_yield_value = _compute_yield_loss(
                    measurements_df,
                    descriptor.test_name,
                    descriptor.test_number,
                    final_lower,
                    final_upper,
                )
                if new_yield_value is not None and not np.isfinite(new_yield_value):
                    new_yield_value = None
                if new_yield_value is not None:
                    if current_yield is None or not _almost_equal(new_yield_value, current_yield):
                        for row_idx in template_rows:
                            sheet_utils.set_cell(template_ws, row_idx, yld_prop_col, new_yield_value)
                        yield_updated = True
                        mark_dirty = True
                elif current_yield is None:
                    pass

        test_changed = proposal_changed or cpk_updated or yield_updated
        if not test_changed:
            continue

        metadata_proposals[descriptor.key()] = {"ll": final_lower, "ul": final_upper, "timestamp": timestamp}
        mark_dirty = True

        updated_tests.append(descriptor)
        if cpk_updated or yield_updated:
            metrics_updates += 1
        audit_details.append(
            {
                "test": descriptor.key(),
                "lower_origin": lower_origin,
                "upper_origin": upper_origin,
                "cpk_updated": cpk_updated,
                "yield_updated": yield_updated,
            }
        )

    if not updated_tests:
        raise ActionCancelled("No tests updated.")

    charts.refresh_tests(context, updated_tests, include_spec=True, include_proposed=True)
    _reapply_cpk_formulas(context)

    summary_parts = [f"Updated proposed limits for {len(updated_tests)} test(s)."]
    if metrics_updates:
        summary_parts.append(f"Recomputed metrics for {metrics_updates} test(s).")
    summary_text = " ".join(summary_parts)

    return {
        "summary": summary_text,
        "warnings": warnings,
        "audit": {
            "scope": resolved["scope"],
            "tests": _tests_to_strings(updated_tests),
            "parameters": {
                "target_cpk": target_cpk,
                "per_test": audit_details,
            },
        },
        "replay_params": resolved,
        "mark_dirty": mark_dirty,
    }


def calculate_proposed_limits_grr(context: PostProcessContext, io: PostProcessIO, params: Optional[dict]) -> dict:
    descriptors = _build_test_descriptors(context)
    if not descriptors:
        raise ActionCancelled("Summary sheet is empty.")

    params = dict(params or {})
    grr_available = params.get("grr_available")
    if grr_available not in {True, False}:
        grr_available = io.confirm("Is GRR data available for proposed limit calculation?", default=True)
    if not grr_available:
        raise ActionCancelled("GRR data is required to compute guardbanded proposed limits.")
    params["grr_available"] = True

    def _prompt_grr_path(previous: str | None) -> Path:
        attempt = _safe_str(previous)
        while True:
            if not attempt:
                attempt = _safe_str(io.prompt("Enter directory containing Total_GRR.xlsx:", default=""))
            if not attempt:
                raise ActionCancelled("GRR directory not provided.")
            candidate = Path(attempt).expanduser()
            workbook_path = candidate if candidate.is_file() else candidate / "Total_GRR.xlsx"
            if workbook_path.exists():
                return candidate
            io.warn(f"GRR workbook not found at {workbook_path}.")
            if not io.confirm("Try another location?", default=True):
                raise ActionCancelled("Unable to locate GRR workbook.")
            attempt = ""

    grr_root = _prompt_grr_path(params.get("grr_path"))
    params["grr_path"] = str(grr_root)

    try:
        grr_table = proposed_limits_grr.load_grr_table(grr_root)
    except (FileNotFoundError, ValueError) as exc:
        raise ActionCancelled(str(exc))

    _ensure_grr_reference_sheet(context.workbook(), grr_table)

    min_cpk = _coerce_positive_float(params.get("cpk_min"))
    while min_cpk is None:
        value = io.prompt_float("Enter minimum FT CPK:", default=None)
        if value is None:
            io.warn("Minimum FT CPK is required.")
            continue
        if value <= 0:
            io.warn("Minimum FT CPK must be positive.")
            continue
        min_cpk = value
    params["cpk_min"] = min_cpk

    max_cpk = _coerce_positive_float(params.get("cpk_max"))
    while max_cpk is None:
        value = io.prompt_float("Enter maximum FT CPK:", default=None)
        if value is None:
            io.warn("Maximum FT CPK is required.")
            continue
        if value <= 0:
            io.warn("Maximum FT CPK must be positive.")
            continue
        max_cpk = value
    while max_cpk < min_cpk:
        io.warn("Maximum FT CPK must be greater than or equal to the minimum.")
        value = io.prompt_float("Enter maximum FT CPK:", default=max_cpk)
        if value is None or value <= 0:
            io.warn("Maximum FT CPK must be positive.")
            continue
        max_cpk = value
    params["cpk_max"] = max_cpk

    resolved = _prompt_scope(io, params, allow_single=True, prompt_target=False)
    selected_tests = _resolve_tests(descriptors, context, io, resolved)

    template_ws = context.template_sheet()
    template_header_row, template_headers = sheet_utils.build_header_map(template_ws)
    ll_prop_col = _resolve_column(template_headers, ["LL_PROP", "LL Proposal", "LL Proposed"])
    ul_prop_col = _resolve_column(template_headers, ["UL_PROP", "UL Proposal", "UL Proposed"])
    cpk_prop_col = _resolve_column(template_headers, ["CPK_PROP", "CPK Proposed"])
    yld_prop_col = _resolve_column(template_headers, ["%YLD LOSS_PROP", "%YLD Loss Proposed"])
    if not all([ll_prop_col, ul_prop_col, cpk_prop_col, yld_prop_col]):
        missing = []
        if not ll_prop_col:
            missing.append("LL_PROP")
        if not ul_prop_col:
            missing.append("UL_PROP")
        if not cpk_prop_col:
            missing.append("CPK_PROP")
        if not yld_prop_col:
            missing.append("%YLD LOSS_PROP")
        raise ActionCancelled(
            "Template sheet missing proposed-limit columns: " + ", ".join(missing) + "."
        )

    spec_columns = _ensure_proposed_spec_columns(template_ws, template_header_row, template_headers)
    template_header_row = spec_columns["header_row"]
    spec_prop_lower_col = spec_columns["spec_lower"]
    spec_prop_upper_col = spec_columns["spec_upper"]
    _ = spec_columns["spec_cpk"]

    spec_lower_col = _resolve_column(template_headers, ["Spec Lower", "Spec Lower Limit"])
    spec_upper_col = _resolve_column(template_headers, ["Spec Upper", "Spec Upper Limit"])

    summary_df = context.summary_frame()
    summary_lookup = _summaries_by_key(summary_df)
    measurements_df = context.measurements_frame()
    metadata_proposals = _ensure_proposal_state(context.metadata)
    timestamp = datetime.now(timezone.utc).isoformat()

    warnings: List[str] = []
    updated_tests: List[TestDescriptor] = []
    skipped_tests: List[TestDescriptor] = []
    metrics_updates = 0
    audit_details: List[dict] = []
    unit_override_cache: Dict[Tuple[str, str], bool] = {}
    mark_dirty = False

    for descriptor in selected_tests:
        row = summary_lookup.get(descriptor.key())
        if row is None:
            _warn_if_missing(io, warnings, f"Summary entry not found for {descriptor.label()} - skipping.")
            continue

        mean = _safe_float(row.get("MEAN"))
        stdev = _safe_float(row.get("STDEV"))
        if mean is None or stdev is None or stdev <= 0:
            _warn_if_missing(io, warnings, f"Cannot compute limits for {descriptor.label()} - missing stats.")
            continue

        template_rows = sheet_utils.find_rows_by_test(
            template_ws,
            template_header_row,
            template_headers,
            test_name=descriptor.test_name,
            test_number=descriptor.test_number,
        )
        if not template_rows:
            _warn_if_missing(io, warnings, f"Template row not found for {descriptor.label()} - skipping.")
            continue

        grr_record = grr_table.find(descriptor.test_number, descriptor.test_name)
        if grr_record is None or grr_record.guardband_full is None or grr_record.guardband_full <= 0:
            _warn_if_missing(io, warnings, f"GRR data missing for {descriptor.label()} - skipping.")
            continue

        workbook_unit = proposed_limits_grr.normalize_unit(descriptor.unit)
        grr_unit = grr_record.unit_normalized
        if workbook_unit and grr_unit and workbook_unit != grr_unit:
            proceed = _resolve_unit_override(
                io,
                unit_override_cache,
                descriptor.label(),
                descriptor.unit,
                grr_record.unit_raw,
            )
            if not proceed:
                warnings.append(
                    f"Skipped {descriptor.label()} due to unmatched units: workbook '{descriptor.unit or 'n/a'}' "
                    f"vs GRR '{grr_record.unit_raw or 'n/a'}'."
                )
                continue

        spec_lower_value = grr_record.spec_lower
        spec_upper_value = grr_record.spec_upper
        if spec_lower_value is None or spec_upper_value is None:
            spec_lower_value = _read_first_from_rows(template_ws, template_rows, spec_lower_col)
            spec_upper_value = _read_first_from_rows(template_ws, template_rows, spec_upper_col)
        spec_lower_f = _safe_float(spec_lower_value)
        spec_upper_f = _safe_float(spec_upper_value)
        if spec_lower_f is None or spec_upper_f is None:
            _warn_if_missing(io, warnings, f"Spec limits missing or non-numeric for {descriptor.label()} - skipping.")
            continue

        skip_guardband, skip_info = _should_skip_grr(
            descriptor,
            template_ws,
            template_rows,
            template_headers,
            spec_lower=spec_lower_f,
            spec_upper=spec_upper_f,
            min_cpk=float(min_cpk),
            max_cpk=float(max_cpk),
            guardband_full=float(grr_record.guardband_full),
        )

        if skip_guardband:
            ll_ate = skip_info.get("ll_ate")
            ul_ate = skip_info.get("ul_ate")
            for row_idx in template_rows:
                if ll_ate is not None:
                    sheet_utils.set_cell(template_ws, row_idx, ll_prop_col, ll_ate)
                if ul_ate is not None:
                    sheet_utils.set_cell(template_ws, row_idx, ul_prop_col, ul_ate)
                sheet_utils.set_cell(template_ws, row_idx, spec_prop_lower_col, spec_lower_f)
                sheet_utils.set_cell(template_ws, row_idx, spec_prop_upper_col, spec_upper_f)
            mark_dirty = True

            if measurements_df.empty:
                _warn_if_missing(io, warnings, f"No measurements available to compute yield for {descriptor.label()}.")
            elif ll_ate is not None and ul_ate is not None:
                yield_value = _compute_yield_loss(
                    measurements_df,
                    descriptor.test_name,
                    descriptor.test_number,
                    ll_ate,
                    ul_ate,
                )
                if yield_value is not None and np.isfinite(yield_value):
                    for row_idx in template_rows:
                        sheet_utils.set_cell(template_ws, row_idx, yld_prop_col, yield_value)
                    metrics_updates += 1

            metadata_proposals[descriptor.key()] = {
                "ll": ll_ate,
                "ul": ul_ate,
                "spec_ll": spec_lower_f,
                "spec_ul": spec_upper_f,
                "guardband": "Existing ATE",
                "timestamp": timestamp,
                "skipped": True,
                "skip_reason": skip_info.get("reason"),
                "required_guardband": float(grr_record.guardband_full),
            }

            audit_entry = {
                "test": descriptor.key(),
                "skipped": True,
                "skip_reason": skip_info.get("reason"),
                "ft_lower": ll_ate,
                "ft_upper": ul_ate,
                "ft_cpk": descriptor.cpk,
                "spec_lower": spec_lower_f,
                "spec_upper": spec_upper_f,
                "existing_lower_guardband_percent": skip_info.get("lower_guardband_percent"),
                "existing_upper_guardband_percent": skip_info.get("upper_guardband_percent"),
                "existing_lower_guardband_width": skip_info.get("lower_guardband_width"),
                "existing_upper_guardband_width": skip_info.get("upper_guardband_width"),
                "required_guardband": skip_info.get("required_guardband"),
            }
            if grr_record.guardband_full is not None:
                audit_entry["grr_full"] = grr_record.guardband_full
            audit_details.append(audit_entry)
            skipped_tests.append(descriptor)
            updated_tests.append(descriptor)
            continue

        try:
            computation = proposed_limits_grr.compute_proposed_limits(
                mean=float(mean),
                stdev=float(stdev),
                spec_lower=spec_lower_f,
                spec_upper=spec_upper_f,
                guardband_full=float(grr_record.guardband_full),
                cpk_min=float(min_cpk),
                cpk_max=float(max_cpk),
            )
        except proposed_limits_grr.ComputationError as exc:
            _warn_if_missing(io, warnings, f"{descriptor.label()}: {exc}")
            continue

        for row_idx in template_rows:
            sheet_utils.set_cell(template_ws, row_idx, ll_prop_col, computation.ft_lower)
            sheet_utils.set_cell(template_ws, row_idx, ul_prop_col, computation.ft_upper)
            sheet_utils.set_cell(template_ws, row_idx, spec_prop_lower_col, computation.spec_lower)
            sheet_utils.set_cell(template_ws, row_idx, spec_prop_upper_col, computation.spec_upper)
        mark_dirty = True

        if measurements_df.empty:
            _warn_if_missing(io, warnings, f"No measurements available to compute yield for {descriptor.label()}.")
        else:
            yield_value = _compute_yield_loss(
                measurements_df,
                descriptor.test_name,
                descriptor.test_number,
                computation.ft_lower,
                computation.ft_upper,
            )
            if yield_value is not None and np.isfinite(yield_value):
                for row_idx in template_rows:
                    sheet_utils.set_cell(template_ws, row_idx, yld_prop_col, yield_value)
                metrics_updates += 1

        if computation.ft_cpk is not None:
            metrics_updates += 1

        metadata_proposals[descriptor.key()] = {
            "ll": computation.ft_lower,
            "ul": computation.ft_upper,
            "spec_ll": computation.spec_lower,
            "spec_ul": computation.spec_upper,
            "guardband": computation.guardband_label,
            "timestamp": timestamp,
            "skipped": False,
        }

        audit_entry = {
            "test": descriptor.key(),
            "guardband": computation.guardband_label,
            "guardband_value": computation.guardband_value,
            "ft_lower": computation.ft_lower,
            "ft_upper": computation.ft_upper,
            "ft_cpk": computation.ft_cpk,
            "spec_lower": computation.spec_lower,
            "spec_upper": computation.spec_upper,
            "spec_cpk": computation.spec_cpk,
            "spec_widened": computation.spec_widened,
            "notes": computation.notes,
        }
        if grr_record.guardband_full is not None:
            audit_entry["grr_full"] = grr_record.guardband_full
        audit_details.append(audit_entry)
        updated_tests.append(descriptor)

    if not updated_tests:
        raise ActionCancelled("No tests updated.")

    charts.refresh_tests(
        context,
        updated_tests,
        include_spec=True,
        include_proposed=False,
        include_proposed_spec=False,
        build_proposed_sheets=True,
    )
    _reapply_cpk_formulas(context)
    template_ws = context.template_sheet()
    template_header_row, template_headers = sheet_utils.build_header_map(template_ws)
    _apply_spec_difference_cf(template_ws, template_headers, template_header_row)

    summary_parts = [f"Calculated GRR-based proposed limits for {len(updated_tests)} test(s)."]
    if metrics_updates:
        summary_parts.append(f"Updated metrics for {metrics_updates} test(s).")
    if skipped_tests:
        summary_parts.append(f"Skipped {len(skipped_tests)} test(s) (guardband already satisfied).")
    summary_text = " ".join(summary_parts)

    return {
        "summary": summary_text,
        "warnings": warnings,
        "audit": {
            "scope": resolved["scope"],
            "tests": _tests_to_strings(updated_tests),
            "parameters": {
                "grr_path": params.get("grr_path"),
                "cpk_min": min_cpk,
                "cpk_max": max_cpk,
                "per_test": audit_details,
                "skipped": _tests_to_strings(skipped_tests),
            },
        },
        "replay_params": {
            "scope": resolved["scope"],
            "test_key": resolved.get("test_key"),
            "grr_available": True,
            "grr_path": params.get("grr_path"),
            "cpk_min": min_cpk,
            "cpk_max": max_cpk,
        },
        "mark_dirty": mark_dirty,
    }
def _compute_yield_loss(
    measurements: pd.DataFrame,
    test_name: str,
    test_number: str,
    lower: float | None,
    upper: float | None,
) -> float:
    if measurements.empty:
        return float("nan")
    filtered = measurements.copy()
    target_name = _safe_str(test_name)
    target_number = _safe_str(test_number)
    if target_name:
        for column in ("Test Name", "TEST NAME"):
            if column in filtered.columns:
                mask = filtered[column].map(_safe_str)
                filtered = filtered[mask == target_name]
                break
    if target_number:
        for column in ("Test Number", "TEST NUM"):
            if column in filtered.columns:
                mask = filtered[column].map(_safe_str)
                filtered = filtered[mask == target_number]
                break

    if filtered.empty or "Value" not in filtered.columns:
        return float("nan")

    values = pd.to_numeric(filtered["Value"], errors="coerce").dropna()
    if values.empty:
        return float("nan")

    # Filter out Inf/-Inf values (dropna() only removes NaN).
    # Rationale: ±Inf typically indicates instrumentation artefacts (saturation,
    # divide-by-zero, etc.) rather than true out-of-spec parts, so counting them
    # as failures would distort the yield metric.
    finite_values = values[np.isfinite(values)]
    if finite_values.empty:
        return float("nan")

    failures = 0
    if lower is not None:
        failures += int(np.sum(finite_values < lower))
    if upper is not None:
        failures += int(np.sum(finite_values > upper))

    if lower is None and upper is None:
        return float("nan")

    return failures / len(finite_values)

