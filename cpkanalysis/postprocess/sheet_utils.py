"""Utilities for working with workbook sheets."""

from __future__ import annotations

from typing import Dict, Iterable, List, Tuple

from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    "build_header_map",
    "find_rows_by_test",
    "set_cell",
    "get_cell",
    "normalize_header",
]


def normalize_header(text: str) -> str:
    stripped = " ".join(text.split()).lower()
    return stripped


def build_header_map(sheet: Worksheet, *, scan_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    """Return (header_row_index, header_map)."""
    best_headers: Dict[str, int] = {}
    header_row_idx = 1

    max_scan = min(scan_rows, sheet.max_row)
    for row_idx in range(1, max_scan + 1):
        row = sheet[row_idx]
        mapping: Dict[str, int] = {}
        for cell in row:
            if cell.value is None:
                continue
            header = normalize_header(str(cell.value))
            if not header:
                continue
            mapping[header] = cell.column
        if len(mapping) > len(best_headers):
            best_headers = mapping
            header_row_idx = row_idx
    return header_row_idx, best_headers


def find_rows_by_test(
    sheet: Worksheet,
    header_row: int,
    header_map: Dict[str, int],
    *,
    test_name: str,
    test_number: str,
) -> List[int]:
    """Return all data row indices matching the supplied test metadata."""
    test_name_key = normalize_header("test name")
    test_num_key = normalize_header("test num")
    test_number_alt = normalize_header("test number")

    name_col = header_map.get(test_name_key)
    if name_col is None:
        name_col = header_map.get(normalize_header("test_name"))
    number_col = header_map.get(test_num_key) or header_map.get(test_number_alt)

    matches: List[int] = []
    if name_col is None or number_col is None:
        return matches

    target_name = normalize_header(test_name)
    target_number = normalize_header(test_number)

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        name_value = sheet.cell(row=row_idx, column=name_col).value
        number_value = sheet.cell(row=row_idx, column=number_col).value
        if name_value is None:
            continue
        if normalize_header(str(name_value)) != target_name:
            continue
        if target_number:
            if number_value is None:
                continue
            if normalize_header(str(number_value)) != target_number:
                continue
        matches.append(row_idx)
    return matches


def set_cell(sheet: Worksheet, row: int, column: int, value) -> None:
    sheet.cell(row=row, column=column, value=value)


def get_cell(sheet: Worksheet, row: int, column: int):
    return sheet.cell(row=row, column=column).value
