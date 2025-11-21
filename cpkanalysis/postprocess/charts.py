"""Utilities for regenerating charts after post-processing updates."""

from __future__ import annotations

import math
import os
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

from .. import mpl_charts
from ..workbook_builder import (
    AXIS_META_SHEET,
    IMAGE_ANCHOR_COLUMN,
    ROW_STRIDE,
    _compute_axis_bounds,
    _ensure_row_height,
    _ensure_site_block_width,
    _format_site_label,
    _normalise_site_value,
    _place_image,
    _safe_sheet_name,
    _remove_axis_tracking_sheet,
    _write_axis_ranges,
)
from .context import PostProcessContext
from .sheet_utils import build_header_map, find_rows_by_test, get_cell, normalize_header

__all__ = ["refresh_tests"]

SPEC_COLOR = "#4B6CB7"
WHAT_IF_COLOR = "#FFA500"
PROPOSED_COLOR = "#008B8B"
PROPOSED_SPEC_COLOR = "#20B2AA"
CHART_PREFIXES = ("Histogram", "CDF", "TimeSeries")
PROPOSED_CHART_PREFIXES = ("HistogramProposed", "CDFProposed", "TimeSeriesProposed")


@dataclass
class LimitInfo:
    stdf_lower: float | None = None
    stdf_upper: float | None = None
    spec_lower: float | None = None
    spec_upper: float | None = None
    what_lower: float | None = None
    what_upper: float | None = None
    proposed_lower: float | None = None
    proposed_upper: float | None = None
    proposed_spec_lower: float | None = None
    proposed_spec_upper: float | None = None


def _ensure_chart_state(metadata: Dict[str, Any]) -> Dict[str, Any]:
    state = metadata.setdefault("post_processing_state", {})
    charts = state.setdefault("chart_positions", {})
    return charts

# Debug logging for chart placement (writes to repo_root/debug_logs)
_DEBUG_ENABLED = os.getenv("CPK_CHART_DEBUG", "").strip().lower() in {"1", "true", "yes", "on"}
_RUN_ID = datetime.now().strftime("%Y%m%d_%H%M%S")
if _DEBUG_ENABLED:
    try:
        _REPO_ROOT = Path(__file__).resolve().parents[2]
        _DEBUG_DIR = _REPO_ROOT / "debug_logs"
        _DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        _DEBUG_LOG = _DEBUG_DIR / f"charts_{_RUN_ID}.log"
    except Exception:
        _DEBUG_LOG = None
else:
    _DEBUG_LOG = None


def _dbg(msg: str) -> None:
    try:
        if _DEBUG_LOG is None:
            return
        with _DEBUG_LOG.open("a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


def _make_test_key(file_key: str, test_name: str, test_number: str) -> str:
    return f"{file_key}|{test_name}|{test_number}"


def _resolve_test_index(chart_state: Dict[str, Any], file_key: str, test_key: str) -> int:
    order_map: Dict[str, Dict[str, int]] = chart_state.setdefault("_order", {})
    file_order = order_map.setdefault(file_key, {})
    if test_key not in file_order:
        file_order[test_key] = len(file_order)
    return file_order[test_key]


def _record_prefix_index(chart_state: Dict[str, Any], prefix: str, file_key: str, test_key: str, index: int) -> None:
    prefix_map: Dict[str, Dict[str, int]] = chart_state.setdefault(prefix, {})
    file_map = prefix_map.setdefault(file_key, {})
    file_map[test_key] = index


def _anchor_row_for_index(index: int) -> int:
    return 2 + (index * ROW_STRIDE)


def _get_plot_sheet(workbook, file_key: str, prefix: str):
    sheet_name = _safe_sheet_name(f"{prefix}_{file_key}")
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    header_text = f"{prefix} plots for {file_key}"
    if sheet.cell(row=1, column=1).value != header_text:
        sheet.cell(row=1, column=1, value=header_text)
    label_col = get_column_letter(1)
    image_col = get_column_letter(IMAGE_ANCHOR_COLUMN)
    if sheet.column_dimensions[label_col].width is None or sheet.column_dimensions[label_col].width < 28:
        sheet.column_dimensions[label_col].width = 28
    if sheet.column_dimensions[image_col].width is None or sheet.column_dimensions[image_col].width < 60:
        sheet.column_dimensions[image_col].width = 60
    return sheet


def _has_existing_charts(workbook, file_key: str) -> bool:
    for prefix in CHART_PREFIXES + PROPOSED_CHART_PREFIXES:
        sheet_name = _safe_sheet_name(f"{prefix}_{file_key}")
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            images = getattr(sheet, "_images", None)
            if images:
                return True
    return False


def _requires_full_refresh(chart_state: Dict[str, Any], workbook, target_keys: Set[Tuple[str, str, str]]) -> bool:
    order_map = chart_state.get("_order") or {}
    axis_cache: Dict[Tuple[str, str, str], Dict[str, float | None]] | None = None
    for file_key, _, _ in target_keys:
        file_order = order_map.get(file_key) or {}
        if file_order:
            continue
        if _has_existing_charts(workbook, file_key):
            return True
        if axis_cache is None:
            axis_cache = _load_axis_ranges(workbook)
        if axis_cache:
            for axis_key in axis_cache:
                if axis_key[0] == file_key:
                    return True
    return False


def _ensure_site_state(chart_state: Dict[str, Any]) -> Dict[str, Dict[str, Dict[Any, int]]]:
    return chart_state.setdefault("sites", {})


def _resolve_site_column(frame: pd.DataFrame) -> Optional[str]:
    for candidate in ("Site", "site"):
        if candidate in frame.columns:
            return candidate
    return None


def _remove_image(sheet, anchor_cell: str) -> None:
    for idx, image in enumerate(list(getattr(sheet, "_images", []))):
        anchor = getattr(image, "anchor", None)
        ref = None
        if hasattr(anchor, "ref"):
            ref = anchor.ref
        elif isinstance(anchor, str):
            ref = anchor
        if ref == anchor_cell:
            sheet._images.pop(idx)
            break


def _compute_cpk(mean: float | None, stdev: float | None, lower: float | None, upper: float | None) -> float | None:
    if mean is None or stdev is None or stdev <= 0:
        return None
    candidates: List[float] = []
    if upper is not None:
        candidates.append((upper - mean) / (3.0 * stdev))
    if lower is not None:
        candidates.append((mean - lower) / (3.0 * stdev))
    if not candidates:
        return None
    result = min(candidates)
    return float(result) if math.isfinite(result) else None


def _dataset_stats(values: np.ndarray) -> tuple[float | None, float | None]:
    if values.size == 0:
        return None, None
    mean = float(np.mean(values))
    stdev = float(np.std(values, ddof=0))
    if not math.isfinite(mean):
        mean = None
    if not math.isfinite(stdev):
        stdev = None
    return mean, stdev


def _image_cell(column: int, row: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _load_axis_ranges(workbook) -> Dict[Tuple[str, str, str], Dict[str, float | None]]:
    if AXIS_META_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[AXIS_META_SHEET]
    ranges: Dict[Tuple[str, str, str], Dict[str, float | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue
        file_key = _safe_text(row[0])
        test_name = _safe_text(row[1])
        test_number = _safe_text(row[2])
        key = (file_key, test_name, test_number)
        ranges[key] = {
            "data_min": _safe_float(row[3]),
            "data_max": _safe_float(row[4]),
            "lower_limit": _safe_float(row[5]),
            "upper_limit": _safe_float(row[6]),
            "axis_min": _safe_float(row[7]),
            "axis_max": _safe_float(row[8]),
        }
    return ranges


def _rewrite_axis_ranges(workbook, axis_ranges: Dict[Tuple[str, str, str], Dict[str, float | None]]) -> None:
    _remove_axis_tracking_sheet(workbook)
    if not axis_ranges:
        return
    _write_axis_ranges(workbook, axis_ranges)


def _replace_chart_image(sheet, index: int, anchor_row: int, label: str, image_bytes: bytes) -> None:
    if index < len(sheet._images):
        sheet._images.pop(index)
    _place_image(sheet, image_bytes, anchor_row, label, position=index)
    _ensure_row_height(sheet, anchor_row)


def refresh_tests(
    context: PostProcessContext,
    tests: Sequence,
    *,
    include_spec: bool = False,
    include_proposed: bool = False,
    include_proposed_spec: bool = False,
    build_proposed_sheets: bool = False,
) -> None:
    """Regenerate plot sheets to reflect updated limits."""
    workbook = context.workbook()
    _dbg("refresh_tests: start")
    summary_df = context.summary_frame(refresh=True)
    measurements_df = context.measurements_frame(refresh=True)
    limits_df = context.limits_frame(refresh=True)
    template_ws = context.template_sheet()

    limit_lookup = _collect_limit_info(limits_df, template_ws)
    summary_lookup = {
        (
            _safe_text(row.get("File")),
            _safe_text(row.get("Test Name")),
            _safe_text(row.get("Test Number")),
        ): row
        for _, row in summary_df.iterrows()
    }

    chart_state = _ensure_chart_state(context.metadata)

    target_keys: Optional[Set[Tuple[str, str, str]]] = None
    if tests:
        target_keys = {
            (
                _safe_text(getattr(test, "file", "")),
                _safe_text(getattr(test, "test_name", "")),
                _safe_text(getattr(test, "test_number", "")),
            )
            for test in tests
        }

    if measurements_df.empty:
        existing_axis = _load_axis_ranges(workbook) if target_keys else {}
        _rewrite_axis_ranges(workbook, existing_axis)
        context.invalidate_frames("measurements")
        return

    if not target_keys:
        _refresh_all_tests(
            context,
            workbook,
            summary_lookup,
            measurements_df,
            limit_lookup,
            include_spec,
            include_proposed,
            include_proposed_spec,
            build_proposed_sheets,
            chart_state,
        )
    else:
        if _requires_full_refresh(chart_state, workbook, target_keys):
            _refresh_all_tests(
                context,
                workbook,
                summary_lookup,
                measurements_df,
                limit_lookup,
                include_spec,
                include_proposed,
                include_proposed_spec,
                build_proposed_sheets,
                chart_state,
            )
        else:
            _refresh_subset_tests(
                context,
                workbook,
                summary_lookup,
                measurements_df,
                limit_lookup,
                include_spec,
                include_proposed,
                include_proposed_spec,
                build_proposed_sheets,
                chart_state,
                target_keys,
            )

    context.invalidate_frames("measurements")


def _refresh_all_tests(
    context: PostProcessContext,
    workbook,
    summary_lookup: Dict[Tuple[str, str, str], pd.Series],
    measurements_df: pd.DataFrame,
    limit_lookup: Dict[Tuple[str, str], LimitInfo],
    include_spec: bool,
    include_proposed: bool,
    include_proposed_spec: bool,
    build_proposed_sheets: bool,
    chart_state: Dict[str, Any],
) -> None:
    _prune_plot_sheets(workbook)
    chart_state.clear()
    chart_state["_order"] = {}
    chart_state["sites"] = {}
    order_map: Dict[str, Dict[str, int]] = chart_state["_order"]
    site_state: Dict[str, Dict[str, Dict[Any, int]]] = chart_state["sites"]
    axis_ranges: Dict[Tuple[str, str, str], Dict[str, float | None]] = {}
    next_index: Dict[str, int] = {}

    grouped = measurements_df.groupby(["File", "Test Name", "Test Number"], sort=False, dropna=False)
    for (file_name, test_name, test_number), group in grouped:
        file_key = _safe_text(file_name)
        safe_test_name = _safe_text(test_name)
        safe_test_number = _safe_text(test_number)
        test_key = (file_key, safe_test_name, safe_test_number)
        test_key_str = _make_test_key(file_key, safe_test_name, safe_test_number)

        numeric_values = pd.to_numeric(group["Value"], errors="coerce")
        valid_mask = numeric_values.notna() & np.isfinite(numeric_values)
        if not bool(valid_mask.any()):
            continue
        filtered_group = group.loc[valid_mask].copy()

        serial_numbers = np.arange(1, len(filtered_group) + 1, dtype=float)
        filtered_group["_serial_number"] = serial_numbers
        sort_keys = ["_serial_number"]
        if "Measurement Index" in filtered_group.columns:
            sort_keys.append("Measurement Index")
        elif "measurement_index" in filtered_group.columns:
            sort_keys.append("measurement_index")
        filtered_group.sort_values(by=sort_keys, kind="mergesort", inplace=True)
        serial_numbers = filtered_group["_serial_number"].to_numpy()
        values = pd.to_numeric(filtered_group["Value"], errors="coerce").to_numpy()

        limit_info = limit_lookup.get((safe_test_name, safe_test_number), LimitInfo())
        markers = _build_markers(
            limit_info,
            include_spec=include_spec,
            include_proposed=include_proposed,
            include_proposed_spec=include_proposed_spec,
        )

        data_limits = [m.value for m in markers if m.value is not None]
        if include_spec:
            data_limits.extend(
                value
                for value in (
                    limit_info.spec_lower,
                    limit_info.spec_upper,
                    limit_info.what_lower,
                    limit_info.what_upper,
                )
                if value is not None
            )
        if include_proposed:
            data_limits.extend(
                value
                for value in (
                    limit_info.proposed_lower,
                    limit_info.proposed_upper,
                )
                if value is not None
            )
        if include_proposed_spec:
            data_limits.extend(
                value
                for value in (
                    limit_info.proposed_spec_lower,
                    limit_info.proposed_spec_upper,
                )
                if value is not None
            )
        lower_limit = limit_info.stdf_lower
        upper_limit = limit_info.stdf_upper

        axis_min, axis_max, data_min, data_max = _bounds_with_markers(values, lower_limit, upper_limit, data_limits)
        x_range = (axis_min, axis_max)
        y_range = (axis_min, axis_max)

        summary_row = summary_lookup.get(test_key)
        cpk = None
        unit_label = ""
        mean = None
        stdev = None
        if summary_row is not None:
            cpk = _safe_float(summary_row.get("CPK"))
            unit_label = _safe_text(summary_row.get("Unit"))
            mean = _safe_float(summary_row.get("MEAN"))
            stdev = _safe_float(summary_row.get("STDEV"))
        proposed_cpk = _compute_cpk(mean, stdev, limit_info.proposed_lower, limit_info.proposed_upper)

        test_label = safe_test_name if not safe_test_number else f"{safe_test_name} (Test {safe_test_number})"
        index = next_index.get(file_key, 0)
        next_index[file_key] = index + 1

        order_map.setdefault(file_key, {})[test_key_str] = index
        anchor_row = _anchor_row_for_index(index)
        _dbg(f"full:{file_key}|{safe_test_name}|{safe_test_number} idx={index} row={anchor_row}")

        hist_sheet = None
        if context.analysis_inputs.generate_histogram:
            hist_sheet = _get_plot_sheet(workbook, file_key, "Histogram")
            image_bytes = mpl_charts.render_histogram(
                values,
                lower_limit=lower_limit,
                upper_limit=upper_limit,
                x_range=x_range,
                test_label=test_label,
                cpk=cpk,
                proposed_cpk=proposed_cpk,
                unit_label=unit_label,
                extra_markers=markers,
                title_font_size=10,
                cpk_font_size=8,
            )
            _dbg(f"place Histogram full idx={index} row={anchor_row}")
            _replace_chart_image(hist_sheet, index, anchor_row, test_label, image_bytes)
            _record_prefix_index(chart_state, "Histogram", file_key, test_key_str, index)

        cdf_sheet = None
        if context.analysis_inputs.generate_cdf:
            cdf_sheet = _get_plot_sheet(workbook, file_key, "CDF")
            image_bytes = mpl_charts.render_cdf(
                values,
                lower_limit=lower_limit,
                upper_limit=upper_limit,
                x_range=x_range,
                test_label=test_label,
                cpk=cpk,
                proposed_cpk=proposed_cpk,
                unit_label=unit_label,
                extra_markers=markers,
                title_font_size=10,
                cpk_font_size=8,
            )
            _dbg(f"place CDF full idx={index} row={anchor_row}")
            _replace_chart_image(cdf_sheet, index, anchor_row, test_label, image_bytes)
            _record_prefix_index(chart_state, "CDF", file_key, test_key_str, index)

        ts_sheet = None
        if context.analysis_inputs.generate_time_series:
            ts_sheet = _get_plot_sheet(workbook, file_key, "TimeSeries")
            image_bytes = mpl_charts.render_time_series(
                x=serial_numbers,
                y=values,
                lower_limit=lower_limit,
                upper_limit=upper_limit,
                y_range=y_range,
                test_label=test_label,
                cpk=cpk,
                proposed_cpk=proposed_cpk,
                unit_label=unit_label,
                extra_markers=_horizontalised_markers(markers),
                title_font_size=10,
                cpk_font_size=8,
            )
            _dbg(f"place TimeSeries full idx={index} row={anchor_row}")
            _replace_chart_image(ts_sheet, index, anchor_row, test_label, image_bytes)
            _record_prefix_index(chart_state, "TimeSeries", file_key, test_key_str, index)

        if build_proposed_sheets:
            _render_proposed_overlay(
                context,
                workbook,
                file_key,
                test_key_str,
                test_label,
                index,
                anchor_row,
                serial_numbers,
                values,
                limit_info,
                axis_min,
                axis_max,
                lower_limit,
                upper_limit,
                cpk,
                proposed_cpk,
                unit_label,
                chart_state,
            )

        axis_ranges[test_key] = {
            "data_min": _safe_float(data_min),
            "data_max": _safe_float(data_max),
            "lower_limit": _safe_float(lower_limit),
            "upper_limit": _safe_float(upper_limit),
            "axis_min": _safe_float(axis_min),
            "axis_max": _safe_float(axis_max),
        }

        _refresh_site_charts(
            site_state,
            workbook,
            file_key,
            test_key_str,
            test_label,
            anchor_row,
            filtered_group,
            markers,
            axis_min,
            axis_max,
            limit_info,
            unit_label,
            context.analysis_inputs.generate_histogram,
            context.analysis_inputs.generate_cdf,
            context.analysis_inputs.generate_time_series,
            hist_sheet,
            cdf_sheet,
            ts_sheet,
        )

    _rewrite_axis_ranges(workbook, axis_ranges)


def _refresh_subset_tests(
    context: PostProcessContext,
    workbook,
    summary_lookup: Dict[Tuple[str, str, str], pd.Series],
    measurements_df: pd.DataFrame,
    limit_lookup: Dict[Tuple[str, str], LimitInfo],
    include_spec: bool,
    include_proposed: bool,
    include_proposed_spec: bool,
    build_proposed_sheets: bool,
    chart_state: Dict[str, Any],
    target_keys: Set[Tuple[str, str, str]],
) -> None:
    existing_axis = _load_axis_ranges(workbook)
    axis_ranges = dict(existing_axis)
    order_map: Dict[str, Dict[str, int]] = chart_state.setdefault("_order", {})
    site_state = _ensure_site_state(chart_state)
    updated = False

    grouped = measurements_df.groupby(["File", "Test Name", "Test Number"], sort=False, dropna=False)
    for (file_name, test_name, test_number), group in grouped:
        file_key = _safe_text(file_name)
        safe_test_name = _safe_text(test_name)
        safe_test_number = _safe_text(test_number)
        test_key = (file_key, safe_test_name, safe_test_number)
        if test_key not in target_keys:
            continue

        numeric_values = pd.to_numeric(group["Value"], errors="coerce")
        valid_mask = numeric_values.notna() & np.isfinite(numeric_values)
        if not bool(valid_mask.any()):
            continue
        filtered_group = group.loc[valid_mask].copy()

        serial_numbers = np.arange(1, len(filtered_group) + 1, dtype=float)
        filtered_group["_serial_number"] = serial_numbers
        sort_keys = ["_serial_number"]
        if "Measurement Index" in filtered_group.columns:
            sort_keys.append("Measurement Index")
        elif "measurement_index" in filtered_group.columns:
            sort_keys.append("measurement_index")
        filtered_group.sort_values(by=sort_keys, kind="mergesort", inplace=True)
        serial_numbers = filtered_group["_serial_number"].to_numpy()
        values = pd.to_numeric(filtered_group["Value"], errors="coerce").to_numpy()

        limit_info = limit_lookup.get((safe_test_name, safe_test_number), LimitInfo())
        markers = _build_markers(
            limit_info,
            include_spec=include_spec,
            include_proposed=include_proposed,
            include_proposed_spec=include_proposed_spec,
        )

        data_limits = [m.value for m in markers if m.value is not None]
        if include_spec:
            data_limits.extend(
                value
                for value in (
                    limit_info.spec_lower,
                    limit_info.spec_upper,
                    limit_info.what_lower,
                    limit_info.what_upper,
                )
                if value is not None
            )
        if include_proposed:
            data_limits.extend(
                value
                for value in (
                    limit_info.proposed_lower,
                    limit_info.proposed_upper,
                )
                if value is not None
            )
        lower_limit = limit_info.stdf_lower
        upper_limit = limit_info.stdf_upper

        axis_min, axis_max, data_min, data_max = _bounds_with_markers(values, lower_limit, upper_limit, data_limits)
        x_range = (axis_min, axis_max)
        y_range = (axis_min, axis_max)

        summary_row = summary_lookup.get(test_key)
        cpk = None
        unit_label = ""
        mean = None
        stdev = None
        if summary_row is not None:
            cpk = _safe_float(summary_row.get("CPK"))
            unit_label = _safe_text(summary_row.get("Unit"))
            mean = _safe_float(summary_row.get("MEAN"))
            stdev = _safe_float(summary_row.get("STDEV"))
        proposed_cpk = _compute_cpk(mean, stdev, limit_info.proposed_lower, limit_info.proposed_upper)

        test_label = safe_test_name if not safe_test_number else f"{safe_test_name} (Test {safe_test_number})"
        test_key_str = _make_test_key(file_key, safe_test_name, safe_test_number)
        index = _resolve_test_index(chart_state, file_key, test_key_str)
        order_map.setdefault(file_key, {})[test_key_str] = index
        anchor_row = _anchor_row_for_index(index)
        _dbg(f"subset:{file_key}|{safe_test_name}|{safe_test_number} idx={index} row={anchor_row}")

        hist_sheet = None
        if context.analysis_inputs.generate_histogram:
            hist_sheet = _get_plot_sheet(workbook, file_key, "Histogram")
            image_bytes = mpl_charts.render_histogram(
                values,
                lower_limit=lower_limit,
                upper_limit=upper_limit,
                x_range=x_range,
                test_label=test_label,
                cpk=cpk,
                proposed_cpk=proposed_cpk,
                unit_label=unit_label,
                extra_markers=markers,
                title_font_size=10,
                cpk_font_size=8,
            )
            _dbg(f"place Histogram subset idx={index} row={anchor_row}")
            _replace_chart_image(hist_sheet, index, anchor_row, test_label, image_bytes)
            _record_prefix_index(chart_state, "Histogram", file_key, test_key_str, index)

        cdf_sheet = None
        if context.analysis_inputs.generate_cdf:
            cdf_sheet = _get_plot_sheet(workbook, file_key, "CDF")
            image_bytes = mpl_charts.render_cdf(
                values,
                lower_limit=lower_limit,
                upper_limit=upper_limit,
                x_range=x_range,
                test_label=test_label,
                cpk=cpk,
                proposed_cpk=proposed_cpk,
                unit_label=unit_label,
                extra_markers=markers,
                title_font_size=10,
                cpk_font_size=8,
            )
            _dbg(f"place CDF subset idx={index} row={anchor_row}")
            _replace_chart_image(cdf_sheet, index, anchor_row, test_label, image_bytes)
            _record_prefix_index(chart_state, "CDF", file_key, test_key_str, index)

        ts_sheet = None
        if context.analysis_inputs.generate_time_series:
            ts_sheet = _get_plot_sheet(workbook, file_key, "TimeSeries")
            image_bytes = mpl_charts.render_time_series(
                x=serial_numbers,
                y=values,
                lower_limit=lower_limit,
                upper_limit=upper_limit,
                y_range=y_range,
                test_label=test_label,
                cpk=cpk,
                proposed_cpk=proposed_cpk,
                unit_label=unit_label,
                extra_markers=_horizontalised_markers(markers),
                title_font_size=10,
                cpk_font_size=8,
            )
            _dbg(f"place TimeSeries subset idx={index} row={anchor_row}")
            _replace_chart_image(ts_sheet, index, anchor_row, test_label, image_bytes)
            _record_prefix_index(chart_state, "TimeSeries", file_key, test_key_str, index)

        if build_proposed_sheets:
            _render_proposed_overlay(
                context,
                workbook,
                file_key,
                test_key_str,
                test_label,
                index,
                anchor_row,
                serial_numbers,
                values,
                limit_info,
                axis_min,
                axis_max,
                lower_limit,
                upper_limit,
                cpk,
                proposed_cpk,

                unit_label,
                chart_state,
            )

        axis_ranges[test_key] = {
            "data_min": _safe_float(data_min),
            "data_max": _safe_float(data_max),
            "lower_limit": _safe_float(lower_limit),
            "upper_limit": _safe_float(upper_limit),
            "axis_min": _safe_float(axis_min),
            "axis_max": _safe_float(axis_max),
        }
        updated = True

        _refresh_site_charts(
            site_state,
            workbook,
            file_key,
            test_key_str,
            test_label,
            anchor_row,
            filtered_group,
            markers,
            axis_min,
            axis_max,
            limit_info,
            unit_label,
            context.analysis_inputs.generate_histogram,
            context.analysis_inputs.generate_cdf,
            context.analysis_inputs.generate_time_series,
            hist_sheet,
            cdf_sheet,
            ts_sheet,
        )

    if updated:
        _rewrite_axis_ranges(workbook, axis_ranges)


def _render_proposed_overlay(
    context: PostProcessContext,
    workbook,
    file_key: str,
    test_key: str,
    test_label: str,
    index: int,
    anchor_row: int,
    serial_numbers: np.ndarray,
    values: np.ndarray,
    limit_info: LimitInfo,
    axis_min: float,
    axis_max: float,
    lower_limit: float | None,
    upper_limit: float | None,
    cpk: float | None,
    proposed_cpk: float | None,
    unit_label: str,
    chart_state: Dict[str, Any],
) -> None:
    markers = _build_markers(
        limit_info,
        include_spec=True,
        include_proposed=True,
        include_proposed_spec=True,
    )
    x_range = (axis_min, axis_max)
    y_range = (axis_min, axis_max)

    if context.analysis_inputs.generate_histogram:
        hist_sheet = _get_plot_sheet(workbook, file_key, "HistogramProposed")
        image_bytes = mpl_charts.render_histogram(
            values,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            x_range=x_range,
            test_label=test_label,
            cpk=cpk,
            proposed_cpk=proposed_cpk,
            unit_label=unit_label,
            extra_markers=markers,
            title_font_size=10,
            cpk_font_size=8,
        )
        _replace_chart_image(hist_sheet, index, anchor_row, test_label, image_bytes)
        _record_prefix_index(chart_state, "HistogramProposed", file_key, test_key, index)

    if context.analysis_inputs.generate_cdf:
        cdf_sheet = _get_plot_sheet(workbook, file_key, "CDFProposed")
        image_bytes = mpl_charts.render_cdf(
            values,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            x_range=x_range,
            test_label=test_label,
            cpk=cpk,
            proposed_cpk=proposed_cpk,
            unit_label=unit_label,
            extra_markers=markers,
            title_font_size=10,
            cpk_font_size=8,
        )
        _replace_chart_image(cdf_sheet, index, anchor_row, test_label, image_bytes)
        _record_prefix_index(chart_state, "CDFProposed", file_key, test_key, index)

    if context.analysis_inputs.generate_time_series:
        ts_sheet = _get_plot_sheet(workbook, file_key, "TimeSeriesProposed")
        image_bytes = mpl_charts.render_time_series(
            x=serial_numbers,
            y=values,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            y_range=y_range,
            test_label=test_label,
            cpk=cpk,
            proposed_cpk=proposed_cpk,
            unit_label=unit_label,
            extra_markers=_horizontalised_markers(markers),
            title_font_size=10,
            cpk_font_size=8,
        )
        _replace_chart_image(ts_sheet, index, anchor_row, test_label, image_bytes)
        _record_prefix_index(chart_state, "TimeSeriesProposed", file_key, test_key, index)


def _refresh_site_charts(
    site_state: Dict[str, Dict[str, Dict[Any, int]]],
    workbook,
    file_key: str,
    test_key: str,
    test_label: str,
    anchor_row: int,
    filtered_group: pd.DataFrame,
    markers: Sequence[mpl_charts.ChartMarker],
    axis_min: float,
    axis_max: float,
    limit_info: LimitInfo,
    unit_label: str,
    include_histogram: bool,
    include_cdf: bool,
    include_time_series: bool,
    hist_sheet,
    cdf_sheet,
    ts_sheet,
) -> None:
    site_column = _resolve_site_column(filtered_group)
    if site_column is None:
        return

    file_sites = site_state.setdefault(file_key, {})
    test_sites = file_sites.setdefault(test_key, {})
    processed_sites: Set[Any] = set()

    site_groups = filtered_group.groupby(site_column, dropna=False, sort=False)
    for site_value_raw, site_group in site_groups:
        site_value_norm = _normalise_site_value(site_value_raw)
        processed_sites.add(site_value_norm)
        block_index = test_sites.get(site_value_norm)
        if block_index is None:
            block_index = len(test_sites) + 1
            test_sites[site_value_norm] = block_index

        sort_keys_site = ["_serial_number"]
        if "Measurement Index" in site_group.columns:
            sort_keys_site.append("Measurement Index")
        elif "measurement_index" in site_group.columns:
            sort_keys_site.append("measurement_index")
        site_group_sorted = site_group.sort_values(by=sort_keys_site, kind="mergesort")
        site_values = pd.to_numeric(site_group_sorted["Value"], errors="coerce").to_numpy()
        site_serials = site_group_sorted["_serial_number"].to_numpy()
        site_mean, site_stdev = _dataset_stats(site_values)
        site_cpk = _compute_cpk(site_mean, site_stdev, limit_info.stdf_lower, limit_info.stdf_upper)
        site_proposed_cpk = _compute_cpk(site_mean, site_stdev, limit_info.proposed_lower, limit_info.proposed_upper)
        display_label = f"{test_label} – Site {_format_site_label(site_value_raw)}"
        x_range = (axis_min, axis_max)
        y_range = (axis_min, axis_max)

        if include_histogram and hist_sheet is not None:
            label_col, image_col = _ensure_site_block_width(hist_sheet, block_index)
            target_cell = _image_cell(image_col, anchor_row)
            _remove_image(hist_sheet, target_cell)
            image_bytes = mpl_charts.render_histogram(
                site_values,
                lower_limit=limit_info.stdf_lower,
                upper_limit=limit_info.stdf_upper,
                x_range=x_range,
                test_label=display_label,
                cpk=site_cpk,
                proposed_cpk=site_proposed_cpk,
                unit_label=unit_label,
                extra_markers=markers,
                title_font_size=10,
                cpk_font_size=8,
            )
            cell_ref = _place_image(
                hist_sheet,
                image_bytes,
                anchor_row,
                label=display_label,
                label_column=label_col,
                image_column=image_col,
                anchor_cell=target_cell,
            )
            _dbg(f"per-site Histogram site={site_value_norm} block={block_index} row={anchor_row} cell={cell_ref}")
            _ensure_row_height(hist_sheet, anchor_row)

        if include_cdf and cdf_sheet is not None:
            label_col, image_col = _ensure_site_block_width(cdf_sheet, block_index)
            target_cell = _image_cell(image_col, anchor_row)
            _remove_image(cdf_sheet, target_cell)
            image_bytes = mpl_charts.render_cdf(
                site_values,
                lower_limit=limit_info.stdf_lower,
                upper_limit=limit_info.stdf_upper,
                x_range=x_range,
                test_label=display_label,
                cpk=site_cpk,
                proposed_cpk=site_proposed_cpk,
                unit_label=unit_label,
                extra_markers=markers,
                title_font_size=10,
                cpk_font_size=8,
            )
            cell_ref = _place_image(
                cdf_sheet,
                image_bytes,
                anchor_row,
                label=display_label,
                label_column=label_col,
                image_column=image_col,
                anchor_cell=target_cell,
            )
            _dbg(f"per-site CDF site={site_value_norm} block={block_index} row={anchor_row} cell={cell_ref}")
            _ensure_row_height(cdf_sheet, anchor_row)

        if include_time_series and ts_sheet is not None:
            label_col, image_col = _ensure_site_block_width(ts_sheet, block_index)
            target_cell = _image_cell(image_col, anchor_row)
            _remove_image(ts_sheet, target_cell)
            image_bytes = mpl_charts.render_time_series(
                x=site_serials,
                y=site_values,
                lower_limit=limit_info.stdf_lower,
                upper_limit=limit_info.stdf_upper,
                y_range=y_range,
                test_label=display_label,
                cpk=site_cpk,
                proposed_cpk=site_proposed_cpk,
                unit_label=unit_label,
                extra_markers=_horizontalised_markers(markers),
                title_font_size=10,
                cpk_font_size=8,
            )
            cell_ref = _place_image(
                ts_sheet,
                image_bytes,
                anchor_row,
                label=display_label,
                label_column=label_col,
                image_column=image_col,
                anchor_cell=target_cell,
            )
            _dbg(f"per-site TimeSeries site={site_value_norm} block={block_index} row={anchor_row} cell={cell_ref}")
            _ensure_row_height(ts_sheet, anchor_row)

    stale_sites = set(test_sites.keys()) - processed_sites
    for stale_site in stale_sites:
        block_index = test_sites.pop(stale_site)
        if include_histogram and hist_sheet is not None:
            label_col, image_col = _ensure_site_block_width(hist_sheet, block_index)
            target_cell = _image_cell(image_col, anchor_row)
            _remove_image(hist_sheet, target_cell)
            hist_sheet.cell(row=anchor_row, column=label_col, value="")
        if include_cdf and cdf_sheet is not None:
            label_col, image_col = _ensure_site_block_width(cdf_sheet, block_index)
            target_cell = _image_cell(image_col, anchor_row)
            _remove_image(cdf_sheet, target_cell)
            cdf_sheet.cell(row=anchor_row, column=label_col, value="")
        if include_time_series and ts_sheet is not None:
            label_col, image_col = _ensure_site_block_width(ts_sheet, block_index)
            target_cell = _image_cell(image_col, anchor_row)
            _remove_image(ts_sheet, target_cell)
            ts_sheet.cell(row=anchor_row, column=label_col, value="")


def _collect_limit_info(limits_df: pd.DataFrame, template_ws) -> Dict[Tuple[str, str], LimitInfo]:
    lookup: Dict[Tuple[str, str], LimitInfo] = {}
    if not limits_df.empty:
        for _, row in limits_df.iterrows():
            key = (
                _safe_text(_row_value(row, ["Test name", "test_name", "Test Name"])),
                _safe_text(_row_value(row, ["Test number", "test_number", "Test Number"])),
            )
            info = lookup.setdefault(key, LimitInfo())
            info.stdf_lower = _safe_float(_row_value(row, ["STDF Lower Limit", "stdf_lower"]))
            info.stdf_upper = _safe_float(_row_value(row, ["STDF Upper Limit", "stdf_upper"]))
            info.spec_lower = _safe_float(_row_value(row, ["Spec Lower Limit", "spec_lower"]))
            info.spec_upper = _safe_float(_row_value(row, ["Spec Upper Limit", "spec_upper"]))
            info.what_lower = _safe_float(_row_value(row, ["User What-If Lower Limit", "what_if_lower"]))
            info.what_upper = _safe_float(_row_value(row, ["User What-If Upper Limit", "what_if_upper"]))

    template_header_row, template_headers = build_header_map(template_ws)
    ll_ate_col = _resolve_template_column(template_headers, ["LL_ATE", "LL ATE", "Lower ATE"])
    ul_ate_col = _resolve_template_column(template_headers, ["UL_ATE", "UL ATE", "Upper ATE"])
    spec_lower_col = _resolve_template_column(template_headers, ["Spec Lower", "Spec Lower Limit"])
    spec_upper_col = _resolve_template_column(template_headers, ["Spec Upper", "Spec Upper Limit"])
    what_lower_col = _resolve_template_column(template_headers, ["What-if Lower", "What If Lower", "What-If Lower"])
    what_upper_col = _resolve_template_column(template_headers, ["What-if Upper", "What If Upper", "What-If Upper"])
    prop_lower_col = _resolve_template_column(template_headers, ["LL_PROP", "LL Proposed"])
    prop_upper_col = _resolve_template_column(template_headers, ["UL_PROP", "UL Proposed"])
    test_name_col = _resolve_template_column(template_headers, ["TEST NAME", "Test Name"])
    test_number_col = _resolve_template_column(template_headers, ["TEST NUM", "Test Number"])

    for row_idx in range(template_header_row + 1, template_ws.max_row + 1):
        test_name = _safe_text(get_cell(template_ws, row_idx, test_name_col)) if test_name_col else ""
        test_number = _safe_text(get_cell(template_ws, row_idx, test_number_col)) if test_number_col else ""
        if not test_name and not test_number:
            continue
        key = (test_name, test_number)
        info = lookup.setdefault(key, LimitInfo())
        if ll_ate_col:
            value = _safe_float(get_cell(template_ws, row_idx, ll_ate_col))
            if value is not None:
                info.stdf_lower = value
        if ul_ate_col:
            value = _safe_float(get_cell(template_ws, row_idx, ul_ate_col))
            if value is not None:
                info.stdf_upper = value
        if spec_lower_col:
            value = _safe_float(get_cell(template_ws, row_idx, spec_lower_col))
            if value is not None:
                info.spec_lower = value
        if spec_upper_col:
            value = _safe_float(get_cell(template_ws, row_idx, spec_upper_col))
            if value is not None:
                info.spec_upper = value
        if what_lower_col:
            value = _safe_float(get_cell(template_ws, row_idx, what_lower_col))
            if value is not None:
                info.what_lower = value
        if what_upper_col:
            value = _safe_float(get_cell(template_ws, row_idx, what_upper_col))
            if value is not None:
                info.what_upper = value
        if prop_lower_col:
            value = _safe_float(get_cell(template_ws, row_idx, prop_lower_col))
            if value is not None:
                info.proposed_lower = value
        if prop_upper_col:
            value = _safe_float(get_cell(template_ws, row_idx, prop_upper_col))
            if value is not None:
                info.proposed_upper = value

    return lookup


def _resolve_template_column(header_map: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        normalized = normalize_header(alias)
        if normalized in header_map:
            return header_map[normalized]
    return None


def _prune_plot_sheets(workbook) -> None:
    for sheet_name in list(workbook.sheetnames):
        if sheet_name.startswith(("Histogram", "CDF", "TimeSeries")):
            del workbook[sheet_name]
    _remove_axis_tracking_sheet(workbook)


def _build_markers(
    info: LimitInfo,
    *,
    include_spec: bool,
    include_proposed: bool,
    include_proposed_spec: bool,
) -> List[mpl_charts.ChartMarker]:
    markers: List[mpl_charts.ChartMarker] = []
    if include_spec:
        if info.spec_lower is not None:
            markers.append(mpl_charts.ChartMarker("Spec Lower", info.spec_lower, "vertical", SPEC_COLOR))
        if info.spec_upper is not None:
            markers.append(mpl_charts.ChartMarker("Spec Upper", info.spec_upper, "vertical", SPEC_COLOR))
        if info.what_lower is not None:
            markers.append(mpl_charts.ChartMarker("What-If Lower", info.what_lower, "vertical", WHAT_IF_COLOR))
        if info.what_upper is not None:
            markers.append(mpl_charts.ChartMarker("What-If Upper", info.what_upper, "vertical", WHAT_IF_COLOR))
    if include_proposed:
        if info.proposed_lower is not None:
            markers.append(mpl_charts.ChartMarker("Proposed Lower", info.proposed_lower, "vertical", PROPOSED_COLOR, linestyle="-"))
        if info.proposed_upper is not None:
            markers.append(mpl_charts.ChartMarker("Proposed Upper", info.proposed_upper, "vertical", PROPOSED_COLOR, linestyle="-"))
    if include_proposed_spec:
        if info.proposed_spec_lower is not None:
            markers.append(
                mpl_charts.ChartMarker(
                    "Proposed Spec Lower",
                    info.proposed_spec_lower,
                    "vertical",
                    PROPOSED_SPEC_COLOR,
                    linestyle="--",
                )
            )
        if info.proposed_spec_upper is not None:
            markers.append(
                mpl_charts.ChartMarker(
                    "Proposed Spec Upper",
                    info.proposed_spec_upper,
                    "vertical",
                    PROPOSED_SPEC_COLOR,
                    linestyle="--",
                )
            )
    return markers


def _horizontalised_markers(markers: Sequence[mpl_charts.ChartMarker]) -> List[mpl_charts.ChartMarker]:
    converted = []
    for marker in markers:
        if marker.orientation == "horizontal":
            converted.append(marker)
        else:
            converted.append(mpl_charts.ChartMarker(marker.label, marker.value, "horizontal", marker.color, marker.linestyle, marker.linewidth))
    return converted


def _bounds_with_markers(values: np.ndarray, lower, upper, markers: Sequence[float]) -> Tuple[float, float, float, float]:
    axis_min, axis_max, data_min, data_max = _compute_axis_bounds(values, lower, upper, desired_ticks=10)
    candidates = [m for m in markers if m is not None and np.isfinite(m)]
    if candidates:
        min_marker = float(min(candidates))
        max_marker = float(max(candidates))
        axis_min = float(min(axis_min, min_marker))
        axis_max = float(max(axis_max, max_marker))
        span = axis_max - axis_min
        if not np.isfinite(span) or span <= 0:
            span = max(abs(axis_min), abs(axis_max), 1.0)
        padding = max(span * 0.02, 1e-6)
        axis_min = float(min(axis_min, min_marker - padding))
        axis_max = float(max(axis_max, max_marker + padding))
    return axis_min, axis_max, data_min, data_max


def _safe_float(value) -> float | None:
    if value in (None, "", "nan"):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if np.isnan(result):
        return None
    return result


def _safe_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        return "" if text.lower() == "nan" else text
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _row_value(row: pd.Series, candidates: Sequence[str]):
    for candidate in candidates:
        if candidate in row.index:
            value = row.get(candidate)
            if pd.notna(value):
                return value
    return None
