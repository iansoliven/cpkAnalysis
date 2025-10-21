from __future__ import annotations

import math
import numbers
import os
import re
import time
import multiprocessing
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .mpl_charts import (
    render_cdf,
    render_histogram,
    render_pareto_chart,
    render_time_series,
    render_yield_chart,
)
from .stats import SUMMARY_COLUMNS, YIELD_SUMMARY_COLUMNS, PARETO_COLUMNS, PARETO_COLUMNS_SITE

_BASE_FALLBACK_DECIMALS = 4
FALLBACK_DECIMALS = _BASE_FALLBACK_DECIMALS
_STDF_FORMAT_RE = re.compile(r"%(?:[-+ #0']*)?(?:\d+)?(?:\.(\d+))?([A-Za-z])")


def set_fallback_decimals(value: Optional[int]) -> None:
    """Configure the default number of decimal places used when STDF hints are absent."""
    global FALLBACK_DECIMALS
    if value is None:
        FALLBACK_DECIMALS = _BASE_FALLBACK_DECIMALS
        return
    try:
        numeric = int(value)
    except (TypeError, ValueError):
        FALLBACK_DECIMALS = _BASE_FALLBACK_DECIMALS
        return
    FALLBACK_DECIMALS = max(0, min(9, numeric))


def _clean_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).strip()
    if not text or text.upper() in {"NAN", "<NA>"}:
        return None
    return text


def _clean_optional_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _excel_number_format(
    stdf_format: Any,
    scale: Any,
    *,
    default_decimals: Optional[int] = None,
) -> str:
    if default_decimals is None:
        default_decimals = FALLBACK_DECIMALS
    fmt_text = _clean_optional_text(stdf_format)
    if fmt_text:
        match = _STDF_FORMAT_RE.search(fmt_text)
        if match:
            precision_text, spec = match.groups()
            spec = (spec or "").lower()
            if spec in {"f"}:
                decimals = int(precision_text) if precision_text is not None else 0
                decimals = max(0, min(decimals, 9))
                if decimals == 0:
                    return "0"
                return "0." + "0" * decimals
            if spec in {"e"}:
                decimals = int(precision_text) if precision_text is not None else 6
                decimals = max(0, min(decimals, 9))
                if decimals == 0:
                    return "0E+00"
                return "0." + "0" * decimals + "E+00"
            if spec in {"g"}:
                decimals = int(precision_text) if precision_text is not None else default_decimals
                decimals = max(0, min(decimals, 6))
                if decimals == 0:
                    return "General"
                return "0." + "#" * decimals
            if spec in {"d", "i"}:
                return "0"
    scale_value = _clean_optional_int(scale)
    if scale_value is not None:
        decimals = min(6, abs(scale_value))
        if decimals == 0:
            return "0"
        return "0." + "#" * decimals
    if default_decimals <= 0:
        return "0"
    return "0." + "#" * default_decimals


def _row_number_format(
    row: Any,
    format_fields: Iterable[Optional[str]],
    scale_fields: Iterable[Optional[str]],
    *,
    default_decimals: Optional[int] = None,
) -> str:
    if default_decimals is None:
        default_decimals = FALLBACK_DECIMALS
    selected_format: Optional[str] = None
    for field in format_fields:
        if not field:
            continue
        candidate = row.get(field) if hasattr(row, "get") else None
        cleaned = _clean_optional_text(candidate)
        if cleaned:
            selected_format = cleaned
            break
    selected_scale: Optional[int] = None
    for field in scale_fields:
        if not field:
            continue
        candidate = row.get(field) if hasattr(row, "get") else None
        cleaned_scale = _clean_optional_int(candidate)
        if cleaned_scale is not None:
            selected_scale = cleaned_scale
            break
    return _excel_number_format(selected_format, selected_scale, default_decimals=default_decimals)


def _set_number_format(cell, number_format: str) -> None:
    value = cell.value
    if isinstance(value, bool) or value is None:
        return
    if isinstance(value, numbers.Real):
        numeric = float(value)
    else:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return
    if math.isnan(numeric):
        return
    cell.number_format = number_format


def _fixed_decimal_format(decimals: int) -> str:
    if decimals <= 0:
        return "0"
    return "0." + "0" * decimals

MEAS_COLUMNS = [
    ("file", "File"),
    ("device_id", "DeviceID"),
    ("test_name", "Test Name"),
    ("test_number", "Test Number"),
    ("value", "Value"),
    ("units", "Units"),
    ("timestamp", "Timestamp/Index"),
]

MEAS_VALUE_COLUMN_INDEX = next(
    idx for idx, (source, _) in enumerate(MEAS_COLUMNS, start=1) if source == "value"
)

TEST_LIMIT_COLUMNS = [
    ("test_name", "Test name"),
    ("test_number", "Test number"),
    ("stdf_upper", "STDF Upper Limit"),
    ("stdf_lower", "STDF Lower Limit"),
    ("unit", "Unit"),
    ("spec_upper", "Spec Upper Limit"),
    ("spec_lower", "Spec Lower Limit"),
    ("what_if_upper", "User What-If Upper Limit"),
    ("what_if_lower", "User What-If Lower Limit"),
]

TEST_LIMIT_COLUMN_INDEX = {
    source: idx for idx, (source, _) in enumerate(TEST_LIMIT_COLUMNS, start=1)
}

def _format_limit_cells(ws, excel_row: int, row: Any) -> None:
    lower_columns = ("stdf_lower", "spec_lower", "what_if_lower")
    lower_number_format = _row_number_format(
        row,
        format_fields=("stdf_lower_format", "stdf_result_format", "stdf_upper_format"),
        scale_fields=("stdf_lower_scale", "stdf_result_scale", "stdf_upper_scale"),
    )
    for column_name in lower_columns:
        column_idx = TEST_LIMIT_COLUMN_INDEX.get(column_name)
        if column_idx is None:
            continue
        cell = ws.cell(row=excel_row, column=column_idx)
        _set_number_format(cell, lower_number_format)

    upper_columns = ("stdf_upper", "spec_upper", "what_if_upper")
    upper_number_format = _row_number_format(
        row,
        format_fields=("stdf_upper_format", "stdf_result_format", "stdf_lower_format"),
        scale_fields=("stdf_upper_scale", "stdf_result_scale", "stdf_lower_scale"),
    )
    for column_name in upper_columns:
        column_idx = TEST_LIMIT_COLUMN_INDEX.get(column_name)
        if column_idx is None:
            continue
        cell = ws.cell(row=excel_row, column=column_idx)
        _set_number_format(cell, upper_number_format)

CPK_COLUMNS = [
    "File",
    "Site",
    "TEST NAME",
    "TEST NUM",
    "UNITS",
    "LL_ATE",
    "UL_ATE",
    "COUNT",
    "MEAN",
    "MEDIAN",
    "STDEV",
    "IQR",
    "CPL",
    "CPU",
    "CPK",
    "%YLD LOSS",
    "LL_2CPK",
    "UL_2CPK",
    "CPK_2.0",
    "%YLD LOSS_2.0",
    "LL_3IQR",
    "UL_3IQR",
    "CPK_3IQR",
    "%YLD LOSS_3IQR",
    "PLOTS",
    "Proposal",
    "Lot Qual",
]

DEFAULT_ROW_HEIGHT_PX = 18  # Approximate Excel default
IMAGE_WIDTH = 640
IMAGE_HEIGHT = 360
IMAGE_ANCHOR_COLUMN = 2
ROW_STRIDE = max(int(IMAGE_HEIGHT / DEFAULT_ROW_HEIGHT_PX) + 4, 32)
COL_STRIDE = 18
AXIS_META_SHEET = "_PlotAxisRanges"
ROW_HEIGHT_POINTS = IMAGE_HEIGHT * 0.75


def _normalise_site_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return int(value)
    return value


def _format_site_label(value: Any) -> str:
    if value is None:
        return "Unknown"
    if isinstance(value, float) and math.isnan(value):
        return "Unknown"
    return str(value)


def _site_block_columns(block_index: int) -> tuple[int, int]:
    label_column = 1 + block_index * COL_STRIDE
    image_column = IMAGE_ANCHOR_COLUMN + block_index * COL_STRIDE
    return label_column, image_column


def _ensure_site_block_width(sheet, block_index: int) -> tuple[int, int]:
    label_column, image_column = _site_block_columns(block_index)
    label_letter = get_column_letter(label_column)
    image_letter = get_column_letter(image_column)
    if sheet.column_dimensions[label_letter].width is None or sheet.column_dimensions[label_letter].width < 28:
        sheet.column_dimensions[label_letter].width = 28
    if sheet.column_dimensions[image_letter].width is None or sheet.column_dimensions[image_letter].width < 60:
        sheet.column_dimensions[image_letter].width = 60
    return label_column, image_column


def build_workbook(
    *,
    summary: pd.DataFrame,
    measurements: pd.DataFrame,
    test_limits: pd.DataFrame,
    limit_sources: dict[tuple[str, str, str], dict[str, str]],
    outlier_summary: dict[str, Any],
    per_file_stats: list[dict[str, Any]],
    output_path: Path,
    template_path: Optional[Path],
    include_histogram: bool,
    include_cdf: bool,
    include_time_series: bool,
    include_yield_pareto: bool = False,
    yield_summary: Optional[pd.DataFrame] = None,
    pareto_summary: Optional[pd.DataFrame] = None,
    site_summary: Optional[pd.DataFrame] = None,
    site_limit_sources: Optional[dict[tuple[Any, ...], dict[str, str]]] = None,
    site_yield_summary: Optional[pd.DataFrame] = None,
    site_pareto_summary: Optional[pd.DataFrame] = None,
    site_enabled: bool = False,
    fallback_decimals: Optional[int] = None,
    temp_dir: Path,
    timing_collector: Optional[dict[str, float]] = None,
    histogram_rug: bool = False,
    max_render_processes: Optional[int] = None,
    defer_save: bool = False,
    workbook_obj: Optional[Workbook] = None,
) -> Workbook:
    previous_decimals = FALLBACK_DECIMALS
    set_fallback_decimals(fallback_decimals)
    workbook = workbook_obj if workbook_obj is not None else _load_base_workbook(template_path)
    try:
        _write_summary_sheet(workbook, summary)
        _write_measurements(workbook, measurements)
        _write_test_limits(workbook, test_limits)
        plot_links: dict[tuple[str, str, str], str] = {}
        site_plot_links: dict[tuple[str, Any, str, str], str] = {}
        if include_histogram or include_cdf or include_time_series:
            charts_start = time.perf_counter()
            plot_links, site_plot_links = _create_plot_sheets(
                workbook,
                measurements,
                test_limits,
                summary,
                include_histogram=include_histogram,
                include_cdf=include_cdf,
                include_time_series=include_time_series,
                timings=timing_collector,
                include_rug=histogram_rug,
                max_render_processes=max_render_processes,
                site_summary=site_summary if site_enabled else None,
                site_enabled=site_enabled,
            )
            if timing_collector is not None:
                timing_collector["charts.total"] = timing_collector.get("charts.total", 0.0) + (
                    time.perf_counter() - charts_start
                )
        _populate_cpk_report(
            workbook,
            summary,
            test_limits,
            plot_links,
            site_summary=site_summary if site_enabled else None,
            site_limit_sources=site_limit_sources if site_enabled else None,
            site_plot_links=site_plot_links if site_enabled else None,
        )

        if include_yield_pareto:
            yield_df = yield_summary if yield_summary is not None else pd.DataFrame(columns=YIELD_SUMMARY_COLUMNS)
            pareto_df = pareto_summary if pareto_summary is not None else pd.DataFrame(columns=PARETO_COLUMNS)
            _write_yield_pareto_sheet(
                workbook,
                yield_df,
                pareto_df,
                site_yield_summary=site_yield_summary if site_enabled else None,
                site_pareto_summary=site_pareto_summary if site_enabled else None,
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if not defer_save:
            save_start = time.perf_counter()
            workbook.save(output_path)
            if timing_collector is not None:
                timing_collector["workbook.save"] = timing_collector.get("workbook.save", 0.0) + (
                    time.perf_counter() - save_start
                )
    finally:
        set_fallback_decimals(previous_decimals)
    return workbook


def _load_base_workbook(template_path: Optional[Path]) -> Workbook:
    if template_path and template_path.exists():
        return load_workbook(template_path)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    return workbook


def _write_summary_sheet(workbook: Workbook, summary: pd.DataFrame) -> None:
    if "Summary" in workbook.sheetnames:
        del workbook["Summary"]
    ws = workbook.create_sheet("Summary", 0)
    ws.append(SUMMARY_COLUMNS)

    for _, row in summary.iterrows():
        ws.append([row.get(column) for column in SUMMARY_COLUMNS])

    last_row = ws.max_row
    if last_row > 1:
        column_index = {name: idx for idx, name in enumerate(SUMMARY_COLUMNS, start=1)}
        percent_columns = {"%YLD LOSS", "%YLD LOSS_2.0", "%YLD LOSS_3IQR"}
        integer_columns = {"COUNT"}
        text_columns = {"File", "Test Name", "Test Number", "Unit"}
        general_format = _fixed_decimal_format(FALLBACK_DECIMALS)
        for row_idx in range(2, last_row + 1):
            for name, idx in column_index.items():
                if name in text_columns:
                    continue
                cell = ws.cell(row=row_idx, column=idx)
                if name in percent_columns:
                    _set_number_format(cell, "0.00%")
                elif name in integer_columns:
                    _set_number_format(cell, "0")
                else:
                    _set_number_format(cell, general_format)

    table = Table(displayName="SummaryTable", ref=f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"

    widths = [18, 42, 18, 12, 10, 12, 12, 12, 12, 10, 10, 10, 14, 14, 14, 12, 16, 14, 14, 12, 18]
    for idx, width in enumerate(widths[: len(SUMMARY_COLUMNS)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_measurements(workbook: Workbook, measurements: pd.DataFrame) -> None:
    for sheet in list(workbook.sheetnames):
        if sheet.startswith("Measurements"):
            del workbook[sheet]

    max_rows = 1_048_576
    chunk_size = max_rows - 1
    sources = [source for source, _ in MEAS_COLUMNS]
    headers = [target for _, target in MEAS_COLUMNS]
    total_rows = len(measurements)

    if total_rows == 0:
        ws = workbook.create_sheet("Measurements", index=1)
        ws.append(headers)
        ws.freeze_panes = "A2"
        for idx, width in enumerate([18, 20, 36, 18, 18, 16, 18][: len(headers)], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        return

    for index, start in enumerate(range(0, total_rows, chunk_size)):
        end = min(start + chunk_size, total_rows)
        chunk = measurements.iloc[start:end]
        suffix = "" if index == 0 else f"_{index + 1}"
        sheet_name = f"Measurements{suffix}"
        sheet_index = 1 + index
        ws = workbook.create_sheet(sheet_name, index=sheet_index)
        ws.append(headers)
        data_view = chunk.loc[:, sources]
        for row_values in data_view.itertuples(index=False, name=None):
            ws.append(row_values)
        last_row = ws.max_row
        table_ref = f"A1:{get_column_letter(len(headers))}{last_row}"
        table_name = f"MeasurementsTable{index + 1}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
        ws.freeze_panes = "A2"
        widths = [18, 20, 36, 18, 18, 16, 18]
        for idx, width in enumerate(widths[: len(headers)], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width


def _write_test_limits(workbook: Workbook, test_limits: pd.DataFrame) -> None:
    if "Test List and Limits" in workbook.sheetnames:
        del workbook["Test List and Limits"]
    ws = workbook.create_sheet("Test List and Limits", index=len(workbook.sheetnames))

    data_frame = test_limits.copy()
    original_frame = data_frame.copy()
    if data_frame.empty:
        for _, header in TEST_LIMIT_COLUMNS:
            ws.append([header])
        return

    ordered_columns = [source for source, _ in TEST_LIMIT_COLUMNS if source in data_frame.columns]
    data_frame = data_frame.loc[:, ordered_columns]
    headers = [target for _, target in TEST_LIMIT_COLUMNS]
    ws.append(headers)
    for offset, (_, row) in enumerate(data_frame.iterrows()):
        ws.append([row.get(source) for source, _ in TEST_LIMIT_COLUMNS])
        excel_row = ws.max_row
        original_row = original_frame.iloc[offset]
        _format_limit_cells(ws, excel_row, original_row)

    last_row = ws.max_row
    table = Table(displayName="LimitsTable", ref=f"A1:{get_column_letter(len(headers))}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    widths = [32, 18, 18, 18, 12, 18, 18, 24, 24]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _create_plot_sheets(
    workbook: Workbook,
    measurements: pd.DataFrame,
    test_limits: pd.DataFrame,
    summary: pd.DataFrame,
    *,
    include_histogram: bool,
    include_cdf: bool,
    include_time_series: bool,
    timings: Optional[dict[str, float]] = None,
    include_rug: bool = False,
    max_render_processes: Optional[int] = None,
    site_summary: Optional[pd.DataFrame] = None,
    site_enabled: bool = False,
) -> tuple[dict[tuple[str, str, str], str], dict[tuple[str, Any, str, str], str]]:
    limit_map = _limit_lookup(test_limits)
    summary_map = {
        (str(row.get("File")), str(row.get("Test Name")), str(row.get("Test Number"))): _maybe_float(row.get("CPK"))
        for _, row in summary.iterrows()
    }
    site_summary_map: dict[tuple[str, Any, str, str], pd.Series] = {}
    if site_summary is not None:
        for _, row in site_summary.iterrows():
            key = (
                str(row.get("File")),
                _normalise_site_value(row.get("Site")),
                str(row.get("Test Name")),
                str(row.get("Test Number")),
            )
            site_summary_map[key] = row

    plot_links: dict[tuple[str, str, str], str] = {}
    site_plot_links: dict[tuple[str, Any, str, str], str] = {}
    axis_ranges: dict[tuple[str, str, str], dict[str, float | None]] = {}
    prepared_measurements = _prepare_measurements_for_plots(measurements)
    grouped = prepared_measurements.groupby(["file", "test_name", "test_number"], sort=False)

    hist_sheets: dict[str, Any] = {}
    cdf_sheets: dict[str, Any] = {}
    time_sheets: dict[str, Any] = {}
    _remove_axis_tracking_sheet(workbook)

    plot_tasks: list[dict[str, Any]] = []

    overall_start = time.perf_counter()
    render_elapsed = 0.0

    rug_active = include_rug and include_histogram

    for (file_name, test_name, test_number), group in grouped:
        limit_info = limit_map.get((test_name, test_number), {})
        finite_mask = group["_value_finite"].to_numpy(copy=False)
        if not bool(np.any(finite_mask)):
            continue
        filtered_group = group.loc[finite_mask]
        values = filtered_group["_value_numeric"].to_numpy(copy=False)
        serial_numbers = filtered_group["_serial_number"].to_numpy(copy=False)
        lower_limit = limit_info.get("active_lower")
        upper_limit = limit_info.get("active_upper")
        unit_label = _sanitize_label(str(limit_info.get("unit") or ""))
        if not unit_label and "units" in filtered_group.columns:
            units_series = filtered_group["units"].dropna().astype(str).str.strip()
            if not units_series.empty:
                unit_label = _sanitize_label(units_series.iloc[0])

        axis_min, axis_max, data_min, data_max = _compute_axis_bounds(
            values,
            lower_limit,
            upper_limit,
            desired_ticks=10,
        )
        x_range = (axis_min, axis_max)
        axis_ranges[(file_name, test_name, test_number)] = {
            "data_min": data_min,
            "data_max": data_max,
            "lower_limit": lower_limit,
            "upper_limit": upper_limit,
            "axis_min": axis_min,
            "axis_max": axis_max,
            "unit": unit_label,
        }

        test_label = test_name if not test_number else f"{test_name} (Test {test_number})"
        test_label = _sanitize_label(test_label)
        cpk_value = summary_map.get((str(file_name), str(test_name), str(test_number)))

        plot_tasks.append(
            {
                "key": (file_name, test_name, test_number),
                "file_name": file_name,
                "test_name": test_name,
                "test_number": test_number,
                "values": values,
                "serial_numbers": serial_numbers,
                "lower_limit": lower_limit,
                "upper_limit": upper_limit,
                "x_range": x_range,
                "axis_min": axis_min,
                "axis_max": axis_max,
                "test_label": test_label,
                "cpk_value": cpk_value,
                "unit_label": unit_label,
                "include_rug": rug_active,
                "block_index": 0,
                "site_value": None,
                "site_label": "",
            }
        )

        if site_enabled and "site" in filtered_group.columns:
            site_groups = filtered_group.groupby("site", dropna=False, sort=False)
            for idx, (site_value_raw, site_group) in enumerate(site_groups, start=1):
                site_values = site_group["_value_numeric"].to_numpy(copy=False)
                if site_values.size == 0:
                    continue
                site_serials = site_group["_serial_number"].to_numpy(copy=False)
                site_data_min = float(np.min(site_values)) if site_values.size else None
                site_data_max = float(np.max(site_values)) if site_values.size else None
                site_value_norm = _normalise_site_value(site_value_raw)
                site_key = (str(file_name), site_value_norm, str(test_name), str(test_number))
                site_row = site_summary_map.get(site_key)
                site_cpk = None
                if site_row is not None:
                    site_cpk = _maybe_float(site_row.get("CPK"))
                axis_entry = axis_ranges[(file_name, test_name, test_number)]
                site_ranges = axis_entry.setdefault("site_ranges", {})
                site_ranges[site_value_norm] = {
                    "data_min": site_data_min,
                    "data_max": site_data_max,
                }
                plot_tasks.append(
                    {
                        "key": ("site", file_name, site_value_norm, test_name, test_number),
                        "file_name": file_name,
                        "test_name": test_name,
                        "test_number": test_number,
                        "values": site_values,
                        "serial_numbers": site_serials,
                        "lower_limit": lower_limit,
                        "upper_limit": upper_limit,
                        "x_range": x_range,
                        "axis_min": axis_min,
                        "axis_max": axis_max,
                        "test_label": test_label,
                        "cpk_value": site_cpk,
                        "unit_label": unit_label,
                        "include_rug": rug_active,
                        "block_index": idx,
                        "site_value": site_value_norm,
                        "site_label": _format_site_label(site_value_raw),
                        "site_data_min": site_data_min,
                        "site_data_max": site_data_max,
                    }
                )

    rendered_images: dict[tuple[Any, ...], dict[str, Optional[bytes]]] = {}
    if plot_tasks:
        render_start = time.perf_counter()
        available_cpus = os.cpu_count() or 1
        if max_render_processes is not None:
            max_workers = max(1, min(int(max_render_processes), available_cpus))
        elif available_cpus <= 2:
            max_workers = 1
        else:
            max_workers = max(1, min(available_cpus - 2, 4))
        if max_workers > 1:
            with multiprocessing.get_context("spawn").Pool(processes=max_workers) as pool:
                args_iter = (
                    (
                        task,
                        include_histogram,
                        include_cdf,
                        include_time_series,
                        task.get("include_rug", False),
                    )
                    for task in plot_tasks
                )
                for result, key in pool.imap_unordered(_render_plot_images_process, args_iter):
                    rendered_images[key] = result
        else:
            for task in plot_tasks:
                rendered_images[task["key"]] = _render_plot_images(
                    task,
                    include_histogram,
                    include_cdf,
                    include_time_series,
                    task.get("include_rug", False),
                )
        render_elapsed = time.perf_counter() - render_start

    plot_positions: dict[tuple[str, str, str], dict[str, tuple[Any, int]]] = {}

    for task in plot_tasks:
        key = task["key"]
        file_name = task["file_name"]
        test_name = task["test_name"]
        test_number = task["test_number"]
        test_label = task["test_label"]
        cpk_value = task["cpk_value"]
        unit_label = task["unit_label"]
        axis_min = task["axis_min"]
        axis_max = task["axis_max"]
        lower_limit = task["lower_limit"]
        upper_limit = task["upper_limit"]
        block_index = task.get("block_index", 0)
        site_value = task.get("site_value")
        site_label = task.get("site_label") or ""
        image_bundle = rendered_images.get(key, {})

        base_identifier = (file_name, test_name, test_number)

        if block_index == 0:
            if include_histogram:
                anchor = _ensure_plot_anchor(hist_sheets, workbook, file_name, "Histogram")
                row, sheet = anchor["row"], anchor["sheet"]
                _ensure_site_block_width(sheet, 0)
                image_bytes = image_bundle.get("histogram")
                if image_bytes is not None:
                    cell = _place_image(sheet, image_bytes, row, label=test_label)
                    sheet_ref = sheet.title.replace("'", "''")
                    plot_links[(file_name, test_name, test_number)] = f"#'{sheet_ref}'!{cell}"
                    _ensure_row_height(sheet, row)
                hist_sheets[file_name]["row"] += ROW_STRIDE
                plot_positions.setdefault(base_identifier, {})["histogram"] = (sheet, row)

            if include_cdf:
                anchor_cdf = _ensure_plot_anchor(cdf_sheets, workbook, file_name, "CDF")
                _ensure_site_block_width(anchor_cdf["sheet"], 0)
                image_bytes = image_bundle.get("cdf")
                if image_bytes is not None:
                    _place_image(anchor_cdf["sheet"], image_bytes, anchor_cdf["row"], label=test_label)
                    _ensure_row_height(anchor_cdf["sheet"], anchor_cdf["row"])
                cdf_sheets[file_name]["row"] += ROW_STRIDE
                plot_positions.setdefault(base_identifier, {})["cdf"] = (anchor_cdf["sheet"], anchor_cdf["row"])

            if include_time_series:
                anchor_ts = _ensure_plot_anchor(time_sheets, workbook, file_name, "TimeSeries")
                _ensure_site_block_width(anchor_ts["sheet"], 0)
                image_bytes = image_bundle.get("time_series")
                if image_bytes is not None:
                    _place_image(
                        anchor_ts["sheet"],
                        image_bytes,
                        anchor_ts["row"],
                        label=test_label,
                    )
                    _ensure_row_height(anchor_ts["sheet"], anchor_ts["row"])
                time_sheets[file_name]["row"] += ROW_STRIDE
                plot_positions.setdefault(base_identifier, {})["time_series"] = (anchor_ts["sheet"], anchor_ts["row"])
            continue

        positions = plot_positions.get(base_identifier)
        if not positions:
            continue

        display_label = f"{test_label} – Site {site_label}"
        site_key = (file_name, site_value, test_name, test_number)

        if include_histogram and "histogram" in positions:
            sheet, row = positions["histogram"]
            label_col, image_col = _ensure_site_block_width(sheet, block_index)
            image_bytes = image_bundle.get("histogram")
            if image_bytes is not None:
                cell = _place_image(
                    sheet,
                    image_bytes,
                    row,
                    label=display_label,
                    label_column=label_col,
                    image_column=image_col,
                )
                sheet_ref = sheet.title.replace("'", "''")
                site_plot_links[(file_name, site_value, test_name, test_number)] = f"#'{sheet_ref}'!{cell}"
                _ensure_row_height(sheet, row)

        if include_cdf and "cdf" in positions:
            sheet, row = positions["cdf"]
            label_col, image_col = _ensure_site_block_width(sheet, block_index)
            image_bytes = image_bundle.get("cdf")
            if image_bytes is not None:
                _place_image(
                    sheet,
                    image_bytes,
                    row,
                    label=display_label,
                    label_column=label_col,
                    image_column=image_col,
                )
                _ensure_row_height(sheet, row)

        if include_time_series and "time_series" in positions:
            sheet, row = positions["time_series"]
            label_col, image_col = _ensure_site_block_width(sheet, block_index)
            image_bytes = image_bundle.get("time_series")
            if image_bytes is not None:
                _place_image(
                    sheet,
                    image_bytes,
                    row,
                    label=display_label,
                    label_column=label_col,
                    image_column=image_col,
                )
                _ensure_row_height(sheet, row)

    _write_axis_ranges(workbook, axis_ranges)
    total_elapsed = time.perf_counter() - overall_start
    if timings is not None:
        if render_elapsed > 0:
            timings["charts.render"] = timings.get("charts.render", 0.0) + render_elapsed
        embed_elapsed = total_elapsed - render_elapsed
        if embed_elapsed > 0:
            timings["charts.embed"] = timings.get("charts.embed", 0.0) + embed_elapsed
    return plot_links, site_plot_links


def _prepare_measurements_for_plots(measurements: pd.DataFrame) -> pd.DataFrame:
    prepared = measurements.copy()
    prepared["_value_numeric"] = pd.to_numeric(prepared["value"], errors="coerce")
    prepared["_value_finite"] = np.isfinite(prepared["_value_numeric"])

    if "device_sequence" in prepared.columns:
        sequence_numeric = pd.to_numeric(prepared["device_sequence"], errors="coerce")
    else:
        sequence_numeric = pd.Series(np.nan, index=prepared.index, dtype=float)

    fallback = prepared.groupby(["file", "test_name", "test_number"]).cumcount() + 1
    fallback = fallback.astype(float)
    prepared["_serial_number"] = np.where(np.isfinite(sequence_numeric), sequence_numeric, fallback.to_numpy())

    if "measurement_index" in prepared.columns:
        measurement_numeric = (
            pd.to_numeric(prepared["measurement_index"], errors="coerce").fillna(0.0).astype(float)
        )
    else:
        measurement_numeric = pd.Series(0.0, index=prepared.index, dtype=float)
    prepared["_measurement_index_numeric"] = measurement_numeric

    prepared.sort_values(
        ["file", "test_name", "test_number", "_serial_number", "_measurement_index_numeric"],
        kind="mergesort",
        inplace=True,
    )
    return prepared


def _render_plot_images_process(args: tuple[dict[str, Any], bool, bool, bool, bool]) -> tuple[dict[str, Optional[bytes]], tuple[str, str, str]]:
    task, include_histogram, include_cdf, include_time_series, include_rug = args
    key = task["key"]
    result = _render_plot_images(
        task,
        include_histogram,
        include_cdf,
        include_time_series,
        include_rug,
    )
    return result, key


def _render_plot_images(
    task: dict[str, Any],
    include_histogram: bool,
    include_cdf: bool,
    include_time_series: bool,
    include_rug: bool,
) -> dict[str, Optional[bytes]]:
    values = task["values"]
    serial_numbers = task["serial_numbers"]
    lower_limit = task["lower_limit"]
    upper_limit = task["upper_limit"]
    x_range = task["x_range"]
    axis_min = task["axis_min"]
    axis_max = task["axis_max"]
    test_label = task["test_label"]
    cpk_value = task["cpk_value"]
    unit_label = task["unit_label"]

    results: dict[str, Optional[bytes]] = {"histogram": None, "cdf": None, "time_series": None}

    if include_histogram:
        rug_enabled = include_rug and task.get("include_rug", False)
        results["histogram"] = render_histogram(
            values,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            x_range=x_range,
            test_label=test_label,
            cpk=cpk_value,
            unit_label=unit_label,
            rug=rug_enabled,
        )

    if include_cdf:
        results["cdf"] = render_cdf(
            values,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            x_range=x_range,
            test_label=test_label,
            cpk=cpk_value,
            unit_label=unit_label,
        )

    if include_time_series:
        results["time_series"] = render_time_series(
            x=serial_numbers,
            y=values,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            y_range=(axis_min, axis_max),
            test_label=test_label,
            cpk=cpk_value,
            unit_label=unit_label,
            x_label="Device Sequence",
        )

    return results


def _write_yield_pareto_sheet(
    workbook: Workbook,
    yield_summary: pd.DataFrame,
    pareto_summary: pd.DataFrame,
    *,
    site_yield_summary: Optional[pd.DataFrame] = None,
    site_pareto_summary: Optional[pd.DataFrame] = None,
) -> None:
    sheet_name = "Yield and Pareto"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, index=len(workbook.sheetnames))

    if yield_summary is None or yield_summary.empty:
        ws.cell(row=1, column=1, value="No yield data available.")
        return

    row_cursor = _write_yield_summary_table(ws, yield_summary)
    for index, (_, yield_row) in enumerate(yield_summary.iterrows(), start=1):
        section_start = row_cursor
        row_cursor = _write_yield_section(ws, yield_row, pareto_summary, index, row_cursor)

        if site_yield_summary is not None and not site_yield_summary.empty:
            file_name = str(yield_row.get("file", ""))
            site_subset = site_yield_summary[site_yield_summary["file"] == file_name]
            if not site_subset.empty:
                for site_idx, (_, site_row) in enumerate(site_subset.iterrows(), start=1):
                    site_value = site_row.get("site")
                    site_label = _format_site_label(site_value)
                    if site_pareto_summary is not None and not site_pareto_summary.empty:
                        site_pareto = site_pareto_summary[
                            (site_pareto_summary["file"] == file_name)
                            & (site_pareto_summary["site"] == site_value)
                        ]
                    else:
                        site_pareto = pd.DataFrame(columns=PARETO_COLUMNS_SITE)
                    _write_site_yield_section(
                        ws,
                        site_row,
                        site_pareto,
                        index,
                        section_start,
                        column_offset=site_idx * COL_STRIDE,
                        site_label=site_label,
                        site_index=site_idx,
                    )


def _write_yield_summary_table(ws, yield_summary: pd.DataFrame) -> int:
    headers = ["File", "Devices Total", "Devices Pass", "Devices Fail", "Yield %"]
    start_row = 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)

    current_row = start_row
    for _, row in yield_summary.iterrows():
        current_row += 1
        ws.cell(row=current_row, column=1, value=str(row.get("file", "")))
        ws.cell(row=current_row, column=2, value=int(row.get("devices_total", 0)))
        ws.cell(row=current_row, column=3, value=int(row.get("devices_pass", 0)))
        ws.cell(row=current_row, column=4, value=int(row.get("devices_fail", 0)))
        yield_percent = row.get("yield_percent")
        percent_value = None
        try:
            value = float(yield_percent)
            if math.isfinite(value):
                percent_value = value
        except (TypeError, ValueError):
            percent_value = None
        ws.cell(row=current_row, column=5, value=percent_value)

    table_ref = f"A{start_row}:{get_column_letter(len(headers))}{current_row}"
    table = Table(displayName="YieldSummaryTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    for r in range(start_row + 1, current_row + 1):
        _set_number_format(ws.cell(row=r, column=2), "0")
        _set_number_format(ws.cell(row=r, column=3), "0")
        _set_number_format(ws.cell(row=r, column=4), "0")
        _set_number_format(ws.cell(row=r, column=5), "0.00%")

    def _ensure_width(column: str, minimum: float) -> None:
        dim = ws.column_dimensions[column]
        existing = dim.width or 0
        if existing < minimum:
            dim.width = minimum

    _ensure_width("A", 32)
    _ensure_width("B", 14)
    _ensure_width("C", 14)
    _ensure_width("D", 14)
    _ensure_width("E", 14)

    return current_row + 2


def _write_yield_section(
    ws,
    yield_row: pd.Series,
    pareto_summary: pd.DataFrame,
    section_index: int,
    start_row: int,
) -> int:
    row_cursor = start_row
    file_name = str(yield_row.get("file", ""))
    header_cell = ws.cell(row=row_cursor, column=1, value=f"File: {file_name}")
    header_cell.font = Font(bold=True)
    row_cursor += 1

    yield_header_row = row_cursor
    yield_headers = ["Outcome", "Units", "Percent"]
    for col, header in enumerate(yield_headers, start=1):
        cell = ws.cell(row=row_cursor, column=col, value=header)
        cell.font = Font(bold=True)

    pass_units = int(yield_row.get("devices_pass", 0))
    fail_units = int(yield_row.get("devices_fail", 0))
    yield_percent = yield_row.get("yield_percent")
    yield_ratio = None
    chart_yield = None
    try:
        value = float(yield_percent)
        if math.isfinite(value):
            yield_ratio = value
            chart_yield = value * 100.0
    except (TypeError, ValueError):
        yield_ratio = None
        chart_yield = None

    data_rows = [
        ("Pass", pass_units, yield_ratio),
        ("Fail", fail_units, None if yield_ratio is None else max(0.0, 1.0 - yield_ratio)),
    ]

    for outcome, units, percent in data_rows:
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value=outcome)
        ws.cell(row=row_cursor, column=2, value=int(units))
        ws.cell(row=row_cursor, column=3, value=percent)

    yield_end_row = row_cursor
    table_ref = f"A{yield_header_row}:{get_column_letter(3)}{yield_end_row}"
    table = Table(displayName=f"YieldTable{section_index}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    for r in range(yield_header_row + 1, yield_end_row + 1):
        _set_number_format(ws.cell(row=r, column=2), "0")
        _set_number_format(ws.cell(row=r, column=3), "0.00%")

    def _ensure_width(column: str, minimum: float) -> None:
        dim = ws.column_dimensions[column]
        existing = dim.width or 0
        if existing < minimum:
            dim.width = minimum

    _ensure_width("A", 32)
    _ensure_width("B", 14)
    _ensure_width("C", 12)

    yield_title = f"{_sanitize_label(file_name)} Yield"
    chart_bytes = render_yield_chart(pass_units, fail_units, yield_percent=chart_yield, title=yield_title)
    yield_chart_col = 6
    _ensure_width(get_column_letter(yield_chart_col), 45)
    _place_image_at(ws, chart_bytes, yield_header_row, yield_chart_col)
    _ensure_row_height(ws, yield_header_row)

    next_row = max(yield_end_row + 2, yield_header_row + ROW_STRIDE)

    pareto_section_start = next_row
    pareto_header_cell = ws.cell(row=pareto_section_start, column=1, value="Pareto")
    pareto_header_cell.font = Font(bold=True)
    row_cursor = pareto_section_start + 1
    pareto_headers = [
        "Test Name",
        "Test Number",
        "Devices Fail",
        "Fail Rate",
        "Cumulative",
        "Lower Limit",
        "Upper Limit",
    ]
    for col, header in enumerate(pareto_headers, start=1):
        cell = ws.cell(row=row_cursor, column=col, value=header)
        cell.font = Font(bold=True)

    if pareto_summary is None or pareto_summary.empty or "file" not in pareto_summary.columns:
        subset = pareto_summary.iloc[0:0] if pareto_summary is not None else pd.DataFrame(columns=PARETO_COLUMNS)
    else:
        subset = pareto_summary[pareto_summary["file"] == file_name]

    chart_labels: list[str] = []
    chart_counts: list[float] = []
    chart_cumulative: list[float] = []

    if subset.empty:
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="No failing tests")
        ws.cell(row=row_cursor, column=2, value="")
        ws.cell(row=row_cursor, column=3, value=0)
        ws.cell(row=row_cursor, column=4, value=None)
        ws.cell(row=row_cursor, column=5, value=None)
        ws.cell(row=row_cursor, column=6, value=None)
        ws.cell(row=row_cursor, column=7, value=None)
    else:
        for _, pareto_row in subset.iterrows():
            row_cursor += 1
            test_name = str(pareto_row.get("test_name", ""))
            test_number = str(pareto_row.get("test_number", ""))
            ws.cell(row=row_cursor, column=1, value=test_name)
            ws.cell(row=row_cursor, column=2, value=test_number)
            fail_units = int(pareto_row.get("devices_fail", 0))
            ws.cell(row=row_cursor, column=3, value=fail_units)

            fail_value = None
            try:
                candidate = float(pareto_row.get("fail_rate_percent"))
                if math.isfinite(candidate):
                    fail_value = candidate
            except (TypeError, ValueError):
                fail_value = None

            cumulative_value = None
            try:
                candidate = float(pareto_row.get("cumulative_percent"))
                if math.isfinite(candidate):
                    cumulative_value = candidate
            except (TypeError, ValueError):
                cumulative_value = None

            ws.cell(row=row_cursor, column=4, value=fail_value if fail_value is not None else None)
            ws.cell(row=row_cursor, column=5, value=cumulative_value if cumulative_value is not None else None)
            ws.cell(row=row_cursor, column=6, value=pareto_row.get("lower_limit"))
            ws.cell(row=row_cursor, column=7, value=pareto_row.get("upper_limit"))

            label = f"{test_name} ({test_number})".strip()
            chart_labels.append(_sanitize_label(label) or "(unnamed)")
            chart_counts.append(float(fail_units))
            chart_cumulative.append(
                (cumulative_value * 100.0) if cumulative_value is not None else 0.0
            )

    pareto_end_row = row_cursor
    table_ref = f"A{pareto_section_start + 1}:{get_column_letter(len(pareto_headers))}{pareto_end_row}"
    table = Table(displayName=f"ParetoTable{section_index}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    for r in range(pareto_section_start + 2, pareto_end_row + 1):
        _set_number_format(ws.cell(row=r, column=3), "0")
        _set_number_format(ws.cell(row=r, column=4), "0.00%")
        _set_number_format(ws.cell(row=r, column=5), "0.00%")

    for column, minimum in zip("ABCDEFG", [32, 18, 14, 12, 12, 14, 14]):
        dim = ws.column_dimensions[column]
        existing = dim.width or 0
        if existing < minimum:
            dim.width = minimum

    pareto_chart_title = f"{_sanitize_label(file_name)} Pareto"
    chart_bytes = render_pareto_chart(chart_labels, chart_counts, chart_cumulative, title=pareto_chart_title)
    _place_image_at(ws, chart_bytes, pareto_section_start + 1, 6)
    _ensure_row_height(ws, pareto_section_start + 1)

    return max(pareto_end_row + 2, pareto_section_start + ROW_STRIDE)


def _write_site_yield_section(
    ws,
    yield_row: pd.Series,
    pareto_summary: pd.DataFrame,
    section_index: int,
    start_row: int,
    *,
    column_offset: int,
    site_label: str,
    site_index: int,
) -> None:
    base_col = column_offset + 1
    file_name = str(yield_row.get("file", ""))
    header_cell = ws.cell(row=start_row, column=base_col, value=f"File: {file_name} (Site {site_label})")
    header_cell.font = Font(bold=True)

    row_cursor = start_row + 1
    yield_headers = ["Outcome", "Units", "Percent"]
    for col_offset, header in enumerate(yield_headers):
        cell = ws.cell(row=row_cursor, column=base_col + col_offset, value=header)
        cell.font = Font(bold=True)

    pass_units = int(yield_row.get("devices_pass", 0))
    fail_units = int(yield_row.get("devices_fail", 0))
    yield_percent = yield_row.get("yield_percent")
    yield_ratio = None
    chart_yield = None
    try:
        value = float(yield_percent)
        if math.isfinite(value):
            yield_ratio = value
            chart_yield = value * 100.0
    except (TypeError, ValueError):
        yield_ratio = None
        chart_yield = None

    data_rows = [
        ("Pass", pass_units, yield_ratio),
        ("Fail", fail_units, None if yield_ratio is None else max(0.0, 1.0 - yield_ratio)),
    ]

    for outcome, units, percent in data_rows:
        row_cursor += 1
        ws.cell(row=row_cursor, column=base_col, value=outcome)
        ws.cell(row=row_cursor, column=base_col + 1, value=int(units))
        ws.cell(row=row_cursor, column=base_col + 2, value=percent)

    yield_end_row = row_cursor
    start_letter = get_column_letter(base_col)
    end_letter = get_column_letter(base_col + len(yield_headers) - 1)
    table_ref = f"{start_letter}{start_row + 1}:{end_letter}{yield_end_row}"
    table = Table(displayName=f"YieldTable{section_index}_Site{site_index}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    for r in range(start_row + 1, yield_end_row + 1):
        _set_number_format(ws.cell(row=r, column=base_col + 1), "0")
        _set_number_format(ws.cell(row=r, column=base_col + 2), "0.00%")

    for col_offset, minimum in zip(range(3), [32, 14, 12]):
        letter = get_column_letter(base_col + col_offset)
        dim = ws.column_dimensions[letter]
        existing = dim.width or 0
        if existing < minimum:
            dim.width = minimum

    yield_chart_col = column_offset + 6
    chart_bytes = render_yield_chart(pass_units, fail_units, yield_percent=chart_yield, title=f"{_sanitize_label(file_name)} Site {site_label} Yield")
    dim = ws.column_dimensions[get_column_letter(yield_chart_col)]
    if (dim.width or 0) < 45:
        dim.width = 45
    _place_image_at(ws, chart_bytes, start_row, yield_chart_col)
    _ensure_row_height(ws, start_row)

    next_row = max(yield_end_row + 2, start_row + ROW_STRIDE)

    pareto_section_start = next_row
    ws.cell(row=pareto_section_start, column=base_col, value="Pareto").font = Font(bold=True)
    row_cursor = pareto_section_start + 1
    pareto_headers = [
        "Test Name",
        "Test Number",
        "Devices Fail",
        "Fail Rate",
        "Cumulative",
        "Lower Limit",
        "Upper Limit",
    ]
    for col_offset, header in enumerate(pareto_headers):
        cell = ws.cell(row=row_cursor, column=base_col + col_offset, value=header)
        cell.font = Font(bold=True)

    site_value = yield_row.get("site")
    if pareto_summary is None or pareto_summary.empty:
        subset = pd.DataFrame(columns=PARETO_COLUMNS_SITE)
    else:
        subset = pareto_summary[
            (pareto_summary.get("file") == file_name)
            & (pareto_summary.get("site") == site_value)
        ]

    chart_labels: list[str] = []
    chart_counts: list[float] = []
    chart_cumulative: list[float] = []

    if subset.empty:
        row_cursor += 1
        for col_offset, value in enumerate(["No failing tests", "", 0, None, None, None, None]):
            ws.cell(row=row_cursor, column=base_col + col_offset, value=value)
    else:
        for _, pareto_row in subset.iterrows():
            row_cursor += 1
            test_name = str(pareto_row.get("test_name", ""))
            test_number = str(pareto_row.get("test_number", ""))
            ws.cell(row=row_cursor, column=base_col, value=test_name)
            ws.cell(row=row_cursor, column=base_col + 1, value=test_number)
            fail_units = int(pareto_row.get("devices_fail", 0))
            ws.cell(row=row_cursor, column=base_col + 2, value=fail_units)

            fail_value = None
            try:
                candidate = float(pareto_row.get("fail_rate_percent"))
                if math.isfinite(candidate):
                    fail_value = candidate
            except (TypeError, ValueError):
                fail_value = None

            cumulative_value = None
            try:
                candidate = float(pareto_row.get("cumulative_percent"))
                if math.isfinite(candidate):
                    cumulative_value = candidate
            except (TypeError, ValueError):
                cumulative_value = None

            ws.cell(row=row_cursor, column=base_col + 3, value=fail_value if fail_value is not None else None)
            ws.cell(row=row_cursor, column=base_col + 4, value=cumulative_value if cumulative_value is not None else None)
            ws.cell(row=row_cursor, column=base_col + 5, value=pareto_row.get("lower_limit"))
            ws.cell(row=row_cursor, column=base_col + 6, value=pareto_row.get("upper_limit"))

            label = f"{test_name} ({test_number})".strip()
            chart_labels.append(_sanitize_label(label) or "(unnamed)")
            chart_counts.append(float(fail_units))
            chart_cumulative.append((cumulative_value * 100.0) if cumulative_value is not None else 0.0)

    pareto_end_row = row_cursor
    start_letter = get_column_letter(base_col)
    end_letter = get_column_letter(base_col + len(pareto_headers) - 1)
    table_ref = f"{start_letter}{pareto_section_start + 1}:{end_letter}{pareto_end_row}"
    table = Table(displayName=f"ParetoTable{section_index}_Site{site_index}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    for r in range(pareto_section_start + 2, pareto_end_row + 1):
        _set_number_format(ws.cell(row=r, column=base_col + 2), "0")
        _set_number_format(ws.cell(row=r, column=base_col + 3), "0.00%")
        _set_number_format(ws.cell(row=r, column=base_col + 4), "0.00%")

    for col_offset, minimum in zip(range(7), [32, 18, 14, 12, 12, 14, 14]):
        letter = get_column_letter(base_col + col_offset)
        dim = ws.column_dimensions[letter]
        existing = dim.width or 0
        if existing < minimum:
            dim.width = minimum

    pareto_chart_title = f"{_sanitize_label(file_name)} Site {site_label} Pareto"
    chart_bytes = render_pareto_chart(chart_labels, chart_counts, chart_cumulative, title=pareto_chart_title)
    _place_image_at(ws, chart_bytes, pareto_section_start + 1, column_offset + 6)
    _ensure_row_height(ws, pareto_section_start + 1)

def _populate_cpk_report(
    workbook: Workbook,
    summary: pd.DataFrame,
    test_limits: pd.DataFrame,
    plot_links: dict[tuple[str, str, str], str],
    *,
    site_summary: Optional[pd.DataFrame] = None,
    site_limit_sources: Optional[dict[tuple[Any, ...], dict[str, str]]] = None,
    site_plot_links: Optional[dict[tuple[str, Any, str, str], str]] = None,
) -> None:
    sheet_name = "CPK Report"
    if sheet_name not in workbook.sheetnames:
        ws = workbook.create_sheet(sheet_name)
        ws.append(CPK_COLUMNS)
    ws = workbook[sheet_name]

    header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    column_map: dict[str, int] = {}
    for idx, column in enumerate(header, start=1):
        if column:
            column_map[column] = idx

    for column in CPK_COLUMNS:
        if column not in column_map:
            header.append(column)
            column_map[column] = len(header)
            ws.cell(row=1, column=column_map[column], value=column)

    # Create a lookup for test limits data (test_name, test_number) -> limits_row
    limits_lookup = {}
    for _, limits_row in test_limits.iterrows():
        key = (limits_row.get("test_name", ""), str(limits_row.get("test_number", "")))
        limits_lookup[key] = limits_row

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    general_format = _fixed_decimal_format(FALLBACK_DECIMALS)
    site_plot_links = site_plot_links or {}

    for _, row in summary.iterrows():
        excel_row_index = ws.max_row + 1
        ws.append([None] * len(header))

        limits_key = (row.get("Test Name", ""), str(row.get("Test Number", "")))
        limits_data = limits_lookup.get(limits_key, {})

        for column in CPK_COLUMNS:
            cell_index = column_map[column]
            cell = ws.cell(row=excel_row_index, column=cell_index)
            if column == "PLOTS":
                link_key = (row["File"], row["Test Name"], row["Test Number"])
                hyperlink = plot_links.get(link_key)
                if hyperlink:
                    cell.value = "Histogram"
                    cell.hyperlink = hyperlink
                    cell.style = "Hyperlink"
                else:
                    cell.value = ""
            elif column == "Site":
                cell.value = ""
            elif column == "Proposal" or column == "Lot Qual":
                cell.value = ""
            elif column == "UNITS":
                # Map "Unit" from summary to "UNITS" for template compatibility
                cell.value = row.get("Unit", "")
            elif column == "TEST NAME":
                # Map "Test Name" from summary to "TEST NAME" for template compatibility
                cell.value = row.get("Test Name", "")
            elif column == "TEST NUM":
                # Map "Test Number" from summary to "TEST NUM" for template compatibility
                cell.value = row.get("Test Number", "")
            elif column == "LL_ATE":
                # Map "stdf_lower" from test_limits to "LL_ATE"
                cell.value = limits_data.get("stdf_lower", "")
            elif column == "UL_ATE":
                # Map "stdf_upper" from test_limits to "UL_ATE"
                cell.value = limits_data.get("stdf_upper", "")
            else:
                cell.value = row.get(column, "")

        ll_index = column_map.get("LL_ATE")
        if ll_index is not None:
            ll_cell = ws.cell(row=excel_row_index, column=ll_index)
            ll_format = _row_number_format(
                limits_data,
                format_fields=("stdf_lower_format", "stdf_result_format", "stdf_upper_format"),
                scale_fields=("stdf_lower_scale", "stdf_result_scale", "stdf_upper_scale"),
            )
            _set_number_format(ll_cell, ll_format)
        ul_index = column_map.get("UL_ATE")
        if ul_index is not None:
            ul_cell = ws.cell(row=excel_row_index, column=ul_index)
            ul_format = _row_number_format(
                limits_data,
                format_fields=("stdf_upper_format", "stdf_result_format", "stdf_lower_format"),
                scale_fields=("stdf_upper_scale", "stdf_result_scale", "stdf_lower_scale"),
            )
            _set_number_format(ul_cell, ul_format)

        text_columns = {"File", "Site", "TEST NAME", "TEST NUM", "UNITS", "PLOTS", "Proposal", "Lot Qual"}
        percent_columns = {"%YLD LOSS", "%YLD LOSS_2.0", "%YLD LOSS_3IQR"}
        integer_columns = {"COUNT"}
        limit_columns = {"LL_ATE", "UL_ATE"}
        for column, cell_index in column_map.items():
            if column in text_columns or column in limit_columns:
                continue
            cell = ws.cell(row=excel_row_index, column=cell_index)
            if column in percent_columns:
                _set_number_format(cell, "0.00%")
            elif column in integer_columns:
                _set_number_format(cell, "0")
            else:
                _set_number_format(cell, general_format)

    if site_summary is not None and not site_summary.empty:
        for _, row in site_summary.iterrows():
            excel_row_index = ws.max_row + 1
            ws.append([None] * len(header))

            file_name = str(row.get("File", ""))
            test_name = row.get("Test Name", "")
            test_number = str(row.get("Test Number", ""))
            site_value = _normalise_site_value(row.get("Site"))
            display_site = _format_site_label(row.get("Site"))

            limits_key = (test_name, test_number)
            limits_data = limits_lookup.get(limits_key, {})

            for column in CPK_COLUMNS:
                cell_index = column_map[column]
                cell = ws.cell(row=excel_row_index, column=cell_index)
                if column == "PLOTS":
                    link_key = (file_name, site_value, test_name, test_number)
                    hyperlink = site_plot_links.get(link_key)
                    if hyperlink:
                        cell.value = "Histogram"
                        cell.hyperlink = hyperlink
                        cell.style = "Hyperlink"
                    else:
                        cell.value = ""
                elif column == "Site":
                    cell.value = display_site
                elif column == "Proposal" or column == "Lot Qual":
                    cell.value = ""
                elif column == "UNITS":
                    cell.value = row.get("Unit", "")
                elif column == "TEST NAME":
                    cell.value = test_name
                elif column == "TEST NUM":
                    cell.value = test_number
                elif column == "LL_ATE":
                    cell.value = limits_data.get("stdf_lower", "")
                elif column == "UL_ATE":
                    cell.value = limits_data.get("stdf_upper", "")
                else:
                    cell.value = row.get(column, "")

            ll_index = column_map.get("LL_ATE")
            if ll_index is not None:
                ll_cell = ws.cell(row=excel_row_index, column=ll_index)
                ll_format = _row_number_format(
                    limits_data,
                    format_fields=("stdf_lower_format", "stdf_result_format", "stdf_upper_format"),
                    scale_fields=("stdf_lower_scale", "stdf_result_scale", "stdf_upper_scale"),
                )
                _set_number_format(ll_cell, ll_format)
            ul_index = column_map.get("UL_ATE")
            if ul_index is not None:
                ul_cell = ws.cell(row=excel_row_index, column=ul_index)
                ul_format = _row_number_format(
                    limits_data,
                    format_fields=("stdf_upper_format", "stdf_result_format", "stdf_lower_format"),
                    scale_fields=("stdf_upper_scale", "stdf_result_scale", "stdf_lower_scale"),
                )
                _set_number_format(ul_cell, ul_format)

            text_columns = {"File", "Site", "TEST NAME", "TEST NUM", "UNITS", "PLOTS", "Proposal", "Lot Qual"}
            percent_columns = {"%YLD LOSS", "%YLD LOSS_2.0", "%YLD LOSS_3IQR"}
            integer_columns = {"COUNT"}
            limit_columns = {"LL_ATE", "UL_ATE"}
            for column, cell_index in column_map.items():
                if column in text_columns or column in limit_columns:
                    continue
                cell = ws.cell(row=excel_row_index, column=cell_index)
                if column in percent_columns:
                    _set_number_format(cell, "0.00%")
                elif column in integer_columns:
                    _set_number_format(cell, "0")
                else:
                    _set_number_format(cell, general_format)

    for idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18


def _ensure_plot_anchor(cache: dict[str, Any], workbook: Workbook, file_name: str, prefix: str) -> dict[str, Any]:
    entry = cache.get(file_name)
    if entry is not None:
        return entry
    sheet_name = _safe_sheet_name(f"{prefix}_{file_name}")
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    cache[file_name] = {"sheet": sheet, "row": 2}
    sheet.cell(row=1, column=1, value=f"{prefix} plots for {file_name}")
    label_col = get_column_letter(1)
    image_col = get_column_letter(IMAGE_ANCHOR_COLUMN)
    if sheet.column_dimensions[label_col].width is None or sheet.column_dimensions[label_col].width < 28:
        sheet.column_dimensions[label_col].width = 28
    if sheet.column_dimensions[image_col].width is None or sheet.column_dimensions[image_col].width < 60:
        sheet.column_dimensions[image_col].width = 60
    return cache[file_name]


def _place_image(
    sheet,
    image_bytes: bytes,
    anchor_row: int,
    label: str,
    *,
    label_column: int = 1,
    image_column: int = IMAGE_ANCHOR_COLUMN,
    position: int | None = None,
    anchor_cell: str | None = None,
) -> str:
    sheet.cell(row=anchor_row, column=label_column, value=label)
    image_stream = BytesIO(image_bytes)
    img = XLImage(image_stream)
    img.width = IMAGE_WIDTH
    img.height = IMAGE_HEIGHT
    target_cell = anchor_cell or f"{get_column_letter(image_column)}{anchor_row}"
    img.anchor = target_cell
    if position is None:
        sheet.add_image(img, target_cell)
    else:
        if position < 0:
            position = 0
        if position >= len(sheet._images):
            sheet._images.append(img)
        else:
            sheet._images.insert(position, img)
    return target_cell


def _place_image_at(sheet, image_bytes: bytes, anchor_row: int, anchor_column: int) -> str:
    image_stream = BytesIO(image_bytes)
    img = XLImage(image_stream)
    img.width = IMAGE_WIDTH
    img.height = IMAGE_HEIGHT
    target_cell = f"{get_column_letter(anchor_column)}{anchor_row}"
    sheet.add_image(img, target_cell)
    return target_cell


def _safe_sheet_name(name: str) -> str:
    invalid = set("\\/:*?[]")
    sanitized = "".join("_" if ch in invalid else ch for ch in name)
    if len(sanitized) > 31:
        sanitized = sanitized[:28] + "..."
    return sanitized or "Sheet"


def _ensure_row_height(sheet, anchor_row: int) -> None:
    row_dim = sheet.row_dimensions[anchor_row]
    if row_dim.height is None or row_dim.height < ROW_HEIGHT_POINTS:
        row_dim.height = ROW_HEIGHT_POINTS


def _remove_axis_tracking_sheet(workbook: Workbook) -> None:
    if AXIS_META_SHEET in workbook.sheetnames:
        del workbook[AXIS_META_SHEET]


def _write_axis_ranges(
    workbook: Workbook,
    axis_ranges: dict[tuple[str, str, str], dict[str, float | None]],
) -> None:
    if not axis_ranges:
        return
    ws = workbook.create_sheet(AXIS_META_SHEET)
    ws.sheet_state = "hidden"
    ws.append([
        "File",
        "Test Name",
        "Test Number",
        "Data Min",
        "Data Max",
        "Lower Limit",
        "Upper Limit",
        "Axis Min",
        "Axis Max",
    ])
    for (file_name, test_name, test_number), values in sorted(axis_ranges.items()):
        ws.append([
            file_name,
            test_name,
            test_number,
            _maybe_float(values.get("data_min")),
            _maybe_float(values.get("data_max")),
            _maybe_float(values.get("lower_limit")),
            _maybe_float(values.get("upper_limit")),
            _maybe_float(values.get("axis_min")),
            _maybe_float(values.get("axis_max")),
        ])


def _sanitize_label(label: str) -> str:
    if not label:
        return ""
    cleaned = []
    for ch in label:
        if ch == "\x96":
            cleaned.append("-")
            continue
        code = ord(ch)
        if code < 32 or ch in {"\t", "\r", "\n"}:
            cleaned.append(" ")
        else:
            cleaned.append(ch)
    return "".join(cleaned).strip()


def _limit_lookup(test_limits: pd.DataFrame) -> dict[tuple[str, str], dict[str, Any]]:
    mapping: dict[tuple[str, str], dict[str, Any]] = {}
    for _, row in test_limits.iterrows():
        key = (str(row.get("test_name", "")), str(row.get("test_number", "")))
        lsl = _coerce_float(row.get("what_if_lower"))
        if lsl is None:
            lsl = _coerce_float(row.get("spec_lower"))
        if lsl is None:
            lsl = _coerce_float(row.get("stdf_lower"))
        usl = _coerce_float(row.get("what_if_upper"))
        if usl is None:
            usl = _coerce_float(row.get("spec_upper"))
        if usl is None:
            usl = _coerce_float(row.get("stdf_upper"))
        mapping[key] = {
            "active_lower": lsl,
            "active_upper": usl,
        }
    return mapping


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _maybe_float(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    return None


def _compute_axis_bounds(
    values: np.ndarray,
    lower_limit: Optional[float],
    upper_limit: Optional[float],
    desired_ticks: int = 10,
) -> tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
    finite = values[np.isfinite(values)]
    data_min = float(finite.min()) if finite.size else None
    data_max = float(finite.max()) if finite.size else None

    min_candidates = [val for val in (data_min, lower_limit) if val is not None and math.isfinite(val)]
    max_candidates = [val for val in (data_max, upper_limit) if val is not None and math.isfinite(val)]

    axis_min = min(min_candidates) if min_candidates else data_min
    axis_max = max(max_candidates) if max_candidates else data_max

    if axis_min is None and axis_max is None:
        axis_min, axis_max = -1.0, 1.0
    elif axis_min is None:
        axis_min = axis_max - 1.0  # type: ignore[operator]
    elif axis_max is None:
        axis_max = axis_min + 1.0

    axis_min = float(axis_min)
    axis_max = float(axis_max)

    if axis_min == axis_max:
        delta = abs(axis_min) * 0.05 or 1.0
        axis_min -= delta
        axis_max += delta

    span = axis_max - axis_min
    tick_spacing = _nice_tick(span / max(desired_ticks, 1)) if span > 0 else 1.0
    if tick_spacing <= 0:
        tick_spacing = max(span * 0.1, 1.0)

    target_min_candidates = [val for val in (data_min, lower_limit) if val is not None and math.isfinite(val)]
    if target_min_candidates:
        target_min = min(target_min_candidates)
        axis_min = min(axis_min, target_min - tick_spacing)

    target_max_candidates = [val for val in (data_max, upper_limit) if val is not None and math.isfinite(val)]
    if target_max_candidates:
        target_max = max(target_max_candidates)
        axis_max = max(axis_max, target_max + tick_spacing)

    if axis_min >= axis_max:
        axis_min -= tick_spacing
        axis_max += tick_spacing

    return axis_min, axis_max, data_min, data_max


def _nice_tick(value: float) -> float:
    if value <= 0 or not math.isfinite(value):
        return 1.0
    exponent = math.floor(math.log10(value))
    fraction = value / (10 ** exponent)
    if fraction < 1.5:
        nice_fraction = 1.0
    elif fraction < 3.0:
        nice_fraction = 2.0
    elif fraction < 7.0:
        nice_fraction = 5.0
    else:
        nice_fraction = 10.0
    return nice_fraction * (10 ** exponent)

