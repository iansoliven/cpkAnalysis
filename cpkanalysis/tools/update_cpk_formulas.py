from __future__ import annotations

import argparse
from pathlib import Path
from typing import Union

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..postprocess import sheet_utils

HELPER_SHEET = "cpk_helpers"
GRR_REFERENCE_SHEET = "_GRR_reference"
SPEC_LEGEND_TEXT = "Light yellow highlight indicates Proposed Spec limit differs from Total_GRR spec."
GUARD_LEGEND_TEXT = (
    "Red guardband percent indicates guardband width is below 50% of required GRR guardband."
)
SPEC_LEGEND_COLOR = "FFFFF2CC"
GUARD_LEGEND_COLOR = "FFFFC7CE"
LEGEND_FONT = Font(bold=True)
SPEC_LEGEND_FILL = PatternFill(start_color=SPEC_LEGEND_COLOR, end_color=SPEC_LEGEND_COLOR, fill_type="solid")
GUARD_LEGEND_FILL = PatternFill(start_color=GUARD_LEGEND_COLOR, end_color=GUARD_LEGEND_COLOR, fill_type="solid")


def _set(cell, formula: str) -> None:
    cell.value = formula


def _clear_conditional_formatting(ws: Worksheet, range_str: str) -> None:
    """Remove existing conditional formatting rule for the given range if present."""
    cf = ws.conditional_formatting
    try:
        cf.remove(range_str)  # type: ignore[attr-defined]
        return
    except (AttributeError, KeyError):
        pass
    rules = getattr(cf, "_cf_rules", None)
    if rules is None:
        return
    if isinstance(rules, dict):
        keys_to_delete = [key for key in list(rules.keys()) if str(key) == range_str]
        for key in keys_to_delete:
            del rules[key]
    else:
        filtered = [rule for rule in rules if str(getattr(rule, "sqref", "")) != range_str]
        if len(filtered) != len(rules):
            setattr(cf, "_cf_rules", filtered)


def _quote_sheet(name: str) -> str:
    escaped = name.replace("'", "''")
    return f"'{escaped}'!"


def _require_column(header_map: dict[str, int], *aliases: str) -> int:
    for header in aliases:
        key = sheet_utils.normalize_header(header)
        column = header_map.get(key)
        if column is not None:
            return column
    alias_list = ", ".join(aliases)
    raise ValueError(f"Template sheet missing required column (aliases: {alias_list}).")


def _ensure_column(ws: Worksheet, header_row: int, header_map: dict[str, int], *aliases: str) -> int:
    for header in aliases:
        key = sheet_utils.normalize_header(header)
        column = header_map.get(key)
        if column is not None:
            return column
    header = aliases[0]
    column = ws.max_column + 1
    ws.cell(row=header_row, column=column, value=header)
    header_map[sheet_utils.normalize_header(header)] = column
    return column


def _cell(column_letter: str, row: int) -> str:
    return f"${column_letter}{row}"


def _sheet_cell(sheet_ref: str, column_letter: str, row: int) -> str:
    return f"{sheet_ref}${column_letter}{row}"


def apply_formulas(workbook: Workbook, template_sheet: Union[str, Worksheet]) -> None:
    if isinstance(template_sheet, Worksheet):
        ws = template_sheet
    else:
        if template_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {template_sheet!r} not found in workbook.")
        ws = workbook[template_sheet]

    sheet_ref = _quote_sheet(ws.title)
    header_row, header_map = sheet_utils.build_header_map(ws)
    data_start_row = header_row + 1

    mean_col = _require_column(header_map, "MEAN")
    stdev_col = _require_column(header_map, "STDEV")
    median_col = _require_column(header_map, "MEDIAN")
    iqr_col = _require_column(header_map, "IQR")
    ll_spec_col = _require_column(header_map, "LL_SPEC", "Spec Lower", "Spec Lower Limit", "Spec LL")
    ul_spec_col = _require_column(header_map, "UL_SPEC", "Spec Upper", "Spec Upper Limit", "Spec UL")
    ll_ate_col = _require_column(header_map, "LL_ATE", "LL ATE", "Lower ATE", "LL")
    ul_ate_col = _require_column(header_map, "UL_ATE", "UL ATE", "Upper ATE", "UL")
    cpl_col = _require_column(header_map, "CPL")
    cpu_col = _require_column(header_map, "CPU")
    cpk_col = _require_column(header_map, "CPK")
    ll_2cpk_col = _require_column(header_map, "LL_2CPK", "LL 2CPK")
    ul_2cpk_col = _require_column(header_map, "UL_2CPK", "UL 2CPK")
    cpk_2_col = _require_column(header_map, "CPK_2.0", "CPK 2.0")
    ll_3iqr_col = _require_column(header_map, "LL_3IQR", "LL 3IQR")
    ul_3iqr_col = _require_column(header_map, "UL_3IQR", "UL 3IQR")
    cpk_3iqr_col = _require_column(header_map, "CPK_3IQR", "CPK 3IQR")
    ll_prop_col = _require_column(header_map, "LL_PROP", "Proposed LL", "LL Proposed")
    ul_prop_col = _require_column(header_map, "UL_PROP", "Proposed UL", "UL Proposed")
    cpk_prop_col = _require_column(header_map, "CPK_PROP", "CPK Proposed")
    test_name_col = _require_column(header_map, "Test Name", "TEST NAME", "Test_Name")
    test_number_col = _require_column(header_map, "Test Number", "TEST NUM", "Test_Num")

    cpk_spec_col = _ensure_column(ws, header_row, header_map, "CPK_SPEC")
    spec_prop_lower_col = _ensure_column(ws, header_row, header_map, "Proposed Spec Lower")
    spec_prop_upper_col = _ensure_column(ws, header_row, header_map, "Proposed Spec Upper")
    spec_prop_cpk_col = _ensure_column(ws, header_row, header_map, "CPK Proposed Spec")
    lower_guardband_pct_col = _ensure_column(ws, header_row, header_map, "Lower Guardband Percent")
    upper_guardband_pct_col = _ensure_column(ws, header_row, header_map, "Upper Guardband Percent")
    lower_guardband_col = _ensure_column(ws, header_row, header_map, "Lower Guardband")
    upper_guardband_col = _ensure_column(ws, header_row, header_map, "Upper Guardband")

    letters = {
        "mean": get_column_letter(mean_col),
        "stdev": get_column_letter(stdev_col),
        "median": get_column_letter(median_col),
        "iqr": get_column_letter(iqr_col),
        "ll_spec": get_column_letter(ll_spec_col),
        "ul_spec": get_column_letter(ul_spec_col),
        "ll_ate": get_column_letter(ll_ate_col),
        "ul_ate": get_column_letter(ul_ate_col),
        "cpl": get_column_letter(cpl_col),
        "cpu": get_column_letter(cpu_col),
        "cpk": get_column_letter(cpk_col),
        "ll_2cpk": get_column_letter(ll_2cpk_col),
        "ul_2cpk": get_column_letter(ul_2cpk_col),
        "cpk_2": get_column_letter(cpk_2_col),
        "ll_3iqr": get_column_letter(ll_3iqr_col),
        "ul_3iqr": get_column_letter(ul_3iqr_col),
        "cpk_3iqr": get_column_letter(cpk_3iqr_col),
        "ll_prop": get_column_letter(ll_prop_col),
        "ul_prop": get_column_letter(ul_prop_col),
        "cpk_prop": get_column_letter(cpk_prop_col),
        "test_name": get_column_letter(test_name_col),
        "test_number": get_column_letter(test_number_col),
        "cpk_spec": get_column_letter(cpk_spec_col),
        "spec_prop_lower": get_column_letter(spec_prop_lower_col),
        "spec_prop_upper": get_column_letter(spec_prop_upper_col),
        "spec_prop_cpk": get_column_letter(spec_prop_cpk_col),
        "lower_guardband_pct": get_column_letter(lower_guardband_pct_col),
        "upper_guardband_pct": get_column_letter(upper_guardband_pct_col),
        "lower_guardband": get_column_letter(lower_guardband_col),
        "upper_guardband": get_column_letter(upper_guardband_col),
    }

    if HELPER_SHEET in workbook.sheetnames:
        helper_index = workbook.sheetnames.index(HELPER_SHEET)
        del workbook[HELPER_SHEET]
        helper_insert_index = helper_index
    else:
        helper_insert_index = workbook.sheetnames.index(ws.title) + 1
    helper_ws = workbook.create_sheet(HELPER_SHEET, helper_insert_index)
    helper_ws.sheet_state = "hidden"

    helper_ws.cell(row=header_row, column=1, value="CPL_SPEC")
    helper_ws.cell(row=header_row, column=2, value="CPU_SPEC")
    helper_ws.cell(row=header_row, column=3, value="CPK_SPEC")

    def _write_legend(row: int, text: str, fill: PatternFill) -> None:
        if not spec_prop_lower_col:
            return
        cell = ws.cell(row=row, column=spec_prop_lower_col)
        if cell.value != text:
            cell.value = text
        cell.font = LEGEND_FONT
        cell.fill = fill

    _write_legend(1, GUARD_LEGEND_TEXT, GUARD_LEGEND_FILL)
    _write_legend(2, SPEC_LEGEND_TEXT, SPEC_LEGEND_FILL)

    guardband_lookup: dict[str, str] | None = None
    if GRR_REFERENCE_SHEET in workbook.sheetnames:
        ref_ws = workbook[GRR_REFERENCE_SHEET]
        ref_header_row, ref_headers = sheet_utils.build_header_map(ref_ws)
        gb_col = ref_headers.get(sheet_utils.normalize_header("Guardband Requirement"))
        name_col = ref_headers.get(sheet_utils.normalize_header("Test Name"))
        number_col = ref_headers.get(sheet_utils.normalize_header("Test Number"))
        if gb_col and name_col and number_col:
            ref_start_row = ref_header_row + 1
            ref_end_row = max(ref_ws.max_row, ref_start_row)
            if ref_start_row <= ref_end_row:
                gb_letter = get_column_letter(gb_col)
                name_letter = get_column_letter(name_col)
                number_letter = get_column_letter(number_col)
                guardband_lookup = {
                    "guardband_range": f"'{GRR_REFERENCE_SHEET}'!${gb_letter}${ref_start_row}:${gb_letter}${ref_end_row}",
                    "name_range": f"'{GRR_REFERENCE_SHEET}'!${name_letter}${ref_start_row}:${name_letter}${ref_end_row}",
                    "number_range": f"'{GRR_REFERENCE_SHEET}'!${number_letter}${ref_start_row}:${number_letter}${ref_end_row}",
                }

    max_row = ws.max_row
    for row in range(data_start_row, max_row + 1):
        mean = _cell(letters["mean"], row)
        stdev = _cell(letters["stdev"], row)
        median = _cell(letters["median"], row)
        iqr = _cell(letters["iqr"], row)
        ll_spec = _cell(letters["ll_spec"], row)
        ul_spec = _cell(letters["ul_spec"], row)
        ll_ate = _cell(letters["ll_ate"], row)
        ul_ate = _cell(letters["ul_ate"], row)
        cpl_ref = _cell(letters["cpl"], row)
        cpu_ref = _cell(letters["cpu"], row)
        ll_2 = _cell(letters["ll_2cpk"], row)
        ul_2 = _cell(letters["ul_2cpk"], row)
        ll_3 = _cell(letters["ll_3iqr"], row)
        ul_3 = _cell(letters["ul_3iqr"], row)
        ll_prop = _cell(letters["ll_prop"], row)
        ul_prop = _cell(letters["ul_prop"], row)
        test_name = _sheet_cell(sheet_ref, letters["test_name"], row)
        test_number = _sheet_cell(sheet_ref, letters["test_number"], row)

        _set(
            ws.cell(row=row, column=cpl_col),
            f"=IF(OR({stdev}<=0,AND({ll_ate}=\"\",{ll_spec}=\"\")),\"\",({mean}-IF({ll_ate}<>\"\",{ll_ate},{ll_spec}))/(3*{stdev}))",
        )
        _set(
            ws.cell(row=row, column=cpu_col),
            f"=IF(OR({stdev}<=0,AND({ul_ate}=\"\",{ul_spec}=\"\")),\"\",(IF({ul_ate}<>\"\",{ul_ate},{ul_spec})-{mean})/(3*{stdev}))",
        )
        _set(
            ws.cell(row=row, column=cpk_col),
            f"=IF(OR({cpl_ref}=\"\",{cpu_ref}=\"\"),\"\",MIN({cpl_ref},{cpu_ref}))",
        )
        _set(
            ws.cell(row=row, column=ll_2cpk_col),
            f"=IF({stdev}>0,{mean}-6*{stdev},\"\")",
        )
        _set(
            ws.cell(row=row, column=ul_2cpk_col),
            f"=IF({stdev}>0,{mean}+6*{stdev},\"\")",
        )
        _set(
            ws.cell(row=row, column=cpk_2_col),
            f"=IF(AND({stdev}>0,{ll_2}<>\"\",{ul_2}<>\"\"),MIN(({mean}-{ll_2})/(3*{stdev}),({ul_2}-{mean})/(3*{stdev})),\"\")",
        )
        _set(
            ws.cell(row=row, column=ll_3iqr_col),
            f"=IF(AND({median}<>\"\",{iqr}<>\"\"),{median}-3*{iqr},\"\")",
        )
        _set(
            ws.cell(row=row, column=ul_3iqr_col),
            f"=IF(AND({median}<>\"\",{iqr}<>\"\"),{median}+3*{iqr},\"\")",
        )
        _set(
            ws.cell(row=row, column=cpk_3iqr_col),
            f"=IF(AND({stdev}>0,{ll_3}<>\"\",{ul_3}<>\"\"),MIN(({mean}-{ll_3})/(3*{stdev}),({ul_3}-{mean})/(3*{stdev})),\"\")",
        )
        _set(
            ws.cell(row=row, column=cpk_prop_col),
            f"=IF(AND({stdev}>0,{ll_prop}<>\"\",{ul_prop}<>\"\"),MIN(({mean}-{ll_prop})/(3*{stdev}),({ul_prop}-{mean})/(3*{stdev})),\"\")",
        )
        spec_prop_lower = _cell(letters["spec_prop_lower"], row)
        spec_prop_upper = _cell(letters["spec_prop_upper"], row)
        _set(
            ws.cell(row=row, column=spec_prop_cpk_col),
            f"=IF(AND({stdev}>0,{spec_prop_lower}<>\"\",{spec_prop_upper}<>\"\"),"
            f"MIN(({mean}-{spec_prop_lower})/(3*{stdev}),({spec_prop_upper}-{mean})/(3*{stdev})),\"\")",
        )
        _set(
            ws.cell(row=row, column=lower_guardband_col),
            f"=IF(OR({ll_prop}=\"\",{spec_prop_lower}=\"\"),\"\",ABS({ll_prop}-{spec_prop_lower}))",
        )
        _set(
            ws.cell(row=row, column=upper_guardband_col),
            f"=IF(OR({ul_prop}=\"\",{spec_prop_upper}=\"\"),\"\",ABS({spec_prop_upper}-{ul_prop}))",
        )
        spec_width = f"({spec_prop_upper}-{spec_prop_lower})"
        if guardband_lookup:
            guardband_expr = (
                f"IFERROR(SUMIFS({guardband_lookup['guardband_range']},"
                f"{guardband_lookup['name_range']},{test_name},"
                f"{guardband_lookup['number_range']},{test_number}),\"\")"
            )
            lower_pct_formula = (
                f"=IF(OR({spec_prop_lower}=\"\",{ll_prop}=\"\",{guardband_expr}=\"\",{guardband_expr}=0),\"\","
                f"ABS({ll_prop}-{spec_prop_lower})/{guardband_expr})"
            )
            upper_pct_formula = (
                f"=IF(OR({spec_prop_upper}=\"\",{ul_prop}=\"\",{guardband_expr}=\"\",{guardband_expr}=0),\"\","
                f"ABS({spec_prop_upper}-{ul_prop})/{guardband_expr})"
            )
        else:
            lower_pct_formula = (
                f"=IF(OR({spec_width}=0,{spec_prop_lower}=\"\",{spec_prop_upper}=\"\",{ll_prop}=\"\"),\"\","
                f"ABS({ll_prop}-{spec_prop_lower})/{spec_width})"
            )
            upper_pct_formula = (
                f"=IF(OR({spec_width}=0,{spec_prop_lower}=\"\",{spec_prop_upper}=\"\",{ul_prop}=\"\"),\"\","
                f"ABS({spec_prop_upper}-{ul_prop})/{spec_width})"
            )

        _set(ws.cell(row=row, column=lower_guardband_pct_col), lower_pct_formula)
        _set(ws.cell(row=row, column=upper_guardband_pct_col), upper_pct_formula)

        sheet_mean = _sheet_cell(sheet_ref, letters["mean"], row)
        sheet_stdev = _sheet_cell(sheet_ref, letters["stdev"], row)
        sheet_ll_spec = _sheet_cell(sheet_ref, letters["ll_spec"], row)
        sheet_ul_spec = _sheet_cell(sheet_ref, letters["ul_spec"], row)

        _set(
            helper_ws.cell(row=row, column=1),
            f"=IF(AND({sheet_stdev}>0,{sheet_ll_spec}<>\"\"),({sheet_mean}-{sheet_ll_spec})/(3*{sheet_stdev}),\"\")",
        )
        _set(
            helper_ws.cell(row=row, column=2),
            f"=IF(AND({sheet_stdev}>0,{sheet_ul_spec}<>\"\"),({sheet_ul_spec}-{sheet_mean})/(3*{sheet_stdev}),\"\")",
        )
        _set(
            helper_ws.cell(row=row, column=3),
            f"=IF(OR($A{row}=\"\",$B{row}=\"\"),\"\",MIN($A{row},$B{row}))",
        )

        _set(
            ws.cell(row=row, column=cpk_spec_col),
            f"=IF('{HELPER_SHEET}'!$C{row}=\"\",\"\",'{HELPER_SHEET}'!$C{row})",
        )

    if data_start_row > max_row:
        return

    decimal_columns = {
        ll_spec_col,
        ul_spec_col,
        ll_ate_col,
        ul_ate_col,
        cpl_col,
        cpu_col,
        cpk_col,
        ll_2cpk_col,
        ul_2cpk_col,
        cpk_2_col,
        ll_3iqr_col,
        ul_3iqr_col,
        cpk_3iqr_col,
        ll_prop_col,
        ul_prop_col,
        cpk_prop_col,
        cpk_spec_col,
        spec_prop_lower_col,
        spec_prop_upper_col,
        spec_prop_cpk_col,
        lower_guardband_col,
        upper_guardband_col,
    }
    for col in decimal_columns:
        if not col:
            continue
        for row in range(data_start_row, max_row + 1):
            ws.cell(row=row, column=col).number_format = "0.000"

    percent_columns = {lower_guardband_pct_col, upper_guardband_pct_col}
    for col in percent_columns:
        if not col:
            continue
        for row in range(data_start_row, max_row + 1):
            ws.cell(row=row, column=col).number_format = "0.0%"

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    for key in ("lower_guardband_pct", "upper_guardband_pct"):
        col_letter = letters.get(key)
        if not col_letter:
            continue
        range_str = f"{col_letter}{data_start_row}:{col_letter}{max_row}"
        _clear_conditional_formatting(ws, range_str)
        ws.conditional_formatting.add(range_str, CellIsRule(operator="lessThan", formula=["0.5"], fill=red_fill))


def update_workbook(path: Path, template_sheet: str) -> None:
    wb = openpyxl.load_workbook(path)
    try:
        apply_formulas(wb, template_sheet)
        wb.save(path)
    finally:
        wb.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inject CPK formulas into the template workbook.")
    parser.add_argument(
        "workbook",
        type=Path,
        help="Path to the template workbook.",
    )
    parser.add_argument(
        "--sheet",
        required=True,
        help="Template sheet name as resolved by the pipeline.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    update_workbook(args.workbook.resolve(), template_sheet=args.sheet)


if __name__ == "__main__":
    main()

