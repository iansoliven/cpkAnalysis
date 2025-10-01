#!/usr/bin/env python3
"""Generate histogram charts grouped by Test Name and Lot.

The script scans every worksheet whose name starts with "Measurements" inside the
combined workbook produced by mergeXlsx.py. For each Event/Test Name/Lot trio it
builds histograms of the Measurement column, overlays them by INT/Interval subgroups,
and annotates the mean shift between the smallest and largest INT buckets. INT series
are ordered so "INT Prescreen" appears first, numeric INT labels follow from smallest to
largest, and any remaining text labels (including "INT Missing") are appended afterward.
Separate sheets named `Histogram_<EVENT>` are inserted immediately after the Measurements
sheets, and a ShiftSummary sheet lists the per-lot limits (taken from the largest INT
dataset) together with the computed mean shifts. Chart worksheet names are sanitised
to satisfy Excel's 31-character limit (falling back to `Unknown Event` when no event
label is present). Charts are arranged with one Test Name per row and one Lot per
column, each sized approximately 10" x 5" (width x height). Legends are rendered just
outside the chart area so the plot surface stays clear of labels.
"""

from __future__ import annotations

import argparse
import math
import re
from collections import OrderedDict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")  # Always render off-screen
import matplotlib.pyplot as plt  # type: ignore
import numpy as np  # type: ignore
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

MEASUREMENTS_PREFIX = "Measurements"
INCH_TO_PIXELS = 96  # Excel assumes ~96 DPI for embedded images
CHART_WIDTH_IN = 10
CHART_HEIGHT_IN = 5
CHART_WIDTH_PX = int(CHART_WIDTH_IN * INCH_TO_PIXELS)
CHART_HEIGHT_PX = int(CHART_HEIGHT_IN * INCH_TO_PIXELS)
HISTOGRAM_SHEET_PREFIX = "Histogram_"
SHIFT_SUMMARY_SHEET_NAME = "ShiftSummary"
MAX_SHEET_NAME_LENGTH = 31
INVALID_SHEET_CHARS = re.compile(r"[\/:*?\[\]]")

SHIFT_SUMMARY_HEADER = [
    "Event",
    "Test Name",
    "Test Number",
    "Unit",
    "Lot",
    "Min INT",
    "Max INT",
    "Min INT Mean",
    "Max INT Mean",
    "Mean Shift",
    "Low Limit (Max INT)",
    "High Limit (Max INT)",
]
ROW_STRIDE = 27  # heuristic rows to leave between chart anchors
COL_STRIDE = 15  # heuristic columns to leave between chart anchors
TEST_LABEL_COLUMN = 1
FIRST_CHART_COLUMN = 2
HEADER_ROW = 3
DEFAULT_MAX_LOTS = 0  # 0 means include every lot discovered


@dataclass
class LotData:
    lot: str
    measurements: List[float]
    low_limits: List[float]
    high_limits: List[float]
    intervals: "OrderedDict[str, List[float]]"
    interval_numeric_values: Dict[str, float]
    interval_low_limits: Dict[str, List[float]]
    interval_high_limits: Dict[str, List[float]]

    def representative_limits(self) -> Tuple[Optional[float], Optional[float]]:
        return _first_finite(self.low_limits), _first_finite(self.high_limits)


@dataclass
class TestData:
    name: str
    number: Optional[str]
    unit: Optional[str]
    lots: "OrderedDict[str, LotData]"

    @property
    def title(self) -> str:
        if self.number and self.number.strip():
            return f"{self.name} (Test {self.number})"
        return self.name


@dataclass
class EventData:
    name: str
    tests: "OrderedDict[Tuple[str, Optional[str]], TestData]"


@dataclass
class IntervalSummary:
    min_label: Optional[str]
    max_label: Optional[str]
    min_mean: Optional[float]
    max_mean: Optional[float]
    mean_shift: Optional[float]
    low_limit_max: Optional[float]
    high_limit_max: Optional[float]


def _first_finite(values: Iterable[float]) -> Optional[float]:
    for value in values:
        if math.isfinite(value):
            return value
    return None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add histogram charts for each Test Name/Lot pairing in the combined workbook."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=Path("combined_workbook.xlsx"),
        help="Path to the workbook that already contains Measurements sheets (default: combined_workbook.xlsx).",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Optional path to save the augmented workbook. Defaults to overwriting the input.",
    )
    parser.add_argument(
        "--max-lots",
        type=int,
        default=DEFAULT_MAX_LOTS,
        help="Maximum number of Lot columns to render (0 means all lots).",
    )
    return parser.parse_args()


def load_measurement_tests(workbook_path: Path) -> "OrderedDict[str, EventData]":
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        measurement_sheet_names = [name for name in wb.sheetnames if name.startswith(MEASUREMENTS_PREFIX)]
        if not measurement_sheet_names:
            raise ValueError("No Measurements sheets were found in the workbook.")

        events: "OrderedDict[str, EventData]" = OrderedDict()
        for sheet_name in measurement_sheet_names:
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            header_map = {str(value).strip(): idx for idx, value in enumerate(header) if value is not None}
            required_columns = ["Test Name", "Measurement", "Lot"]
            for column in required_columns:
                if column not in header_map:
                    raise ValueError(f"Column '{column}' was not found in sheet '{sheet_name}'.")

            idx_test_name = header_map["Test Name"]
            idx_measurement = header_map["Measurement"]
            idx_lot = header_map["Lot"]
            idx_test_number = header_map.get("Test Number")
            idx_unit = header_map.get("Test Unit")
            idx_low_limit = header_map.get("Low Limit")
            idx_high_limit = header_map.get("High Limit")
            idx_event = header_map.get("Event")
            idx_interval = None
            for candidate in ("INT", "Int", "int", "Interval", "interval"):
                idx_interval = header_map.get(candidate)
                if idx_interval is not None:
                    break
            if idx_interval is None:
                for header_key, column_idx in header_map.items():
                    if header_key and str(header_key).strip().lower() in {"int", "interval"}:
                        idx_interval = column_idx
                        break

            for row in rows:
                test_name = _safe_string(row, idx_test_name)
                if not test_name:
                    continue
                measurement_raw = row[idx_measurement] if idx_measurement < len(row) else None
                measurement = _coerce_float(measurement_raw)
                if measurement is None:
                    continue

                lot_label = _safe_string(row, idx_lot) or "Unknown Lot"
                test_number = _safe_string(row, idx_test_number)
                unit_text = _safe_string(row, idx_unit)
                event_label = _safe_string(row, idx_event) if idx_event is not None else None
                event_key = event_label if event_label else "Unknown Event"
                interval_label = _safe_string(row, idx_interval) if idx_interval is not None else None
                interval_numeric = None
                if idx_interval is not None and idx_interval < len(row):
                    interval_numeric = _coerce_float(row[idx_interval])
                if interval_numeric is None and interval_label:
                    interval_numeric = _coerce_float(interval_label)

                event_entry = events.get(event_key)
                if event_entry is None:
                    event_entry = EventData(name=event_key, tests=OrderedDict())
                    events[event_key] = event_entry

                key = (test_name, test_number)
                entry = event_entry.tests.get(key)
                if entry is None:
                    entry = TestData(
                        name=test_name,
                        number=test_number,
                        unit=unit_text,
                        lots=OrderedDict(),
                    )
                    event_entry.tests[key] = entry
                elif not entry.unit and unit_text:
                    entry.unit = unit_text

                lot_entry = entry.lots.get(lot_label)
                if lot_entry is None:
                    lot_entry = LotData(
                        lot=lot_label,
                        measurements=[],
                        low_limits=[],
                        high_limits=[],
                        intervals=OrderedDict(),
                        interval_numeric_values={},
                        interval_low_limits={},
                        interval_high_limits={},
                    )
                    entry.lots[lot_label] = lot_entry

                lot_entry.measurements.append(measurement)
                interval_key = interval_label if interval_label else ("INT Missing" if idx_interval is not None else "All Data")
                interval_group = lot_entry.intervals.setdefault(interval_key, [])
                interval_group.append(measurement)
                if interval_numeric is not None and interval_key not in lot_entry.interval_numeric_values:
                    lot_entry.interval_numeric_values[interval_key] = interval_numeric
                if idx_low_limit is not None and idx_low_limit < len(row):
                    low_value = _coerce_float(row[idx_low_limit])
                    if low_value is not None:
                        lot_entry.low_limits.append(low_value)
                        lot_entry.interval_low_limits.setdefault(interval_key, []).append(low_value)
                if idx_high_limit is not None and idx_high_limit < len(row):
                    high_value = _coerce_float(row[idx_high_limit])
                    if high_value is not None:
                        lot_entry.high_limits.append(high_value)
                        lot_entry.interval_high_limits.setdefault(interval_key, []).append(high_value)
        if not events:
            raise ValueError("No measurement rows were found across the Measurements sheets.")
        return events
    finally:
        wb.close()



def _coerce_float(value: Optional[object]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        value_float = float(value)
        return value_float if math.isfinite(value_float) else None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def _safe_string(row: Tuple[object, ...], index: Optional[int]) -> Optional[str]:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    return str(value).strip()


def determine_axis_limits(values: List[float], declared_low: Optional[float], declared_high: Optional[float]) -> Tuple[float, float]:
    if not values:
        raise ValueError("determine_axis_limits called with empty values")
    min_value = min(values)
    max_value = max(values)
    low = declared_low if declared_low is not None else min_value
    high = declared_high if declared_high is not None else max_value
    low = min(low, min_value)
    high = max(high, max_value)
    if math.isclose(low, high):
        span = max(abs(low) * 0.1, 1.0)
        low -= span
        high += span
    return low, high


def choose_bin_count(values: List[float], low: float, high: float) -> int:
    if not values:
        return 10
    span = high - low
    if span <= 0:
        span = max(abs(low) * 0.1, 1.0)
        high = low + span
    n = len(values)
    bins = max(10, min(60, int(round(math.sqrt(n) * 1.5))))
    return bins


def _sanitize_sheet_title(base_name: str, existing_names: Iterable[str]) -> str:
    sanitized = INVALID_SHEET_CHARS.sub("_", base_name).strip()
    if not sanitized:
        sanitized = "Sheet"
    sanitized = sanitized[:MAX_SHEET_NAME_LENGTH]
    candidate = sanitized
    index = 1
    existing_set = set(existing_names)
    while candidate in existing_set:
        suffix = f"_{index}"
        trim_length = MAX_SHEET_NAME_LENGTH - len(suffix)
        candidate = f"{sanitized[:trim_length]}{suffix}"
        index += 1
    return candidate


def _histogram_sheet_name(event_name: str, existing_names: Iterable[str]) -> str:
    label = event_name or "Unknown Event"
    base_title = f"{HISTOGRAM_SHEET_PREFIX}{label}"
    return _sanitize_sheet_title(base_title, existing_names)

def _format_interval_label(label: str) -> str:
    label_str = str(label)
    if label_str in {"All Data", "INT Missing"}:
        return label_str
    if label_str.upper().startswith("INT "):
        return label_str
    return f"INT {label_str}"


def _prepare_interval_groups(lot_data: LotData) -> Tuple[List[Tuple[str, List[float]]], List[Tuple[float, str, List[float]]]]:
    interval_items = [(label, values) for label, values in lot_data.intervals.items() if values]
    if not interval_items:
        interval_items = [("All Data", lot_data.measurements)]

    numeric_groups: List[Tuple[float, str, List[float]]] = []
    prescreen_group: Optional[Tuple[str, List[float]]] = None
    fallback_groups: List[Tuple[str, List[float]]] = []
    for label, values in interval_items:
        numeric_value = lot_data.interval_numeric_values.get(label)
        if numeric_value is None:
            try:
                numeric_value = float(label)
            except (TypeError, ValueError):
                numeric_value = None
        if numeric_value is None:
            label_str = str(label).strip()
            normalized = label_str.lower()
            if normalized.startswith("int "):
                normalized = normalized[4:].strip()
            if normalized == "prescreen" and prescreen_group is None:
                prescreen_group = (label, values)
            else:
                fallback_groups.append((label, values))
        else:
            numeric_groups.append((numeric_value, label, values))
    numeric_groups.sort(key=lambda item: item[0])
    ordered_groups: List[Tuple[str, List[float]]] = []
    if prescreen_group is not None:
        ordered_groups.append(prescreen_group)
    ordered_groups.extend((label, values) for _, label, values in numeric_groups)
    ordered_groups.extend(fallback_groups)
    if not ordered_groups:
        ordered_groups = [("All Data", lot_data.measurements)]
    return ordered_groups, numeric_groups

def _summarize_interval_shift(
    lot_data: LotData, numeric_groups: Optional[List[Tuple[float, str, List[float]]]] = None
) -> IntervalSummary:
    if numeric_groups is None:
        _, numeric_groups = _prepare_interval_groups(lot_data)

    low_declared, high_declared = lot_data.representative_limits()
    if not numeric_groups:
        return IntervalSummary(
            min_label=None,
            max_label=None,
            min_mean=None,
            max_mean=None,
            mean_shift=None,
            low_limit_max=low_declared,
            high_limit_max=high_declared,
        )

    _, min_label, min_values = numeric_groups[0]
    _, max_label, max_values = numeric_groups[-1]
    min_mean = float(np.mean(min_values)) if min_values else None
    max_mean = float(np.mean(max_values)) if max_values else None
    mean_shift = None
    if min_mean is not None and max_mean is not None:
        mean_shift = max_mean - min_mean

    low_list = lot_data.interval_low_limits.get(max_label)
    high_list = lot_data.interval_high_limits.get(max_label)
    low_limit_max = _first_finite(low_list) if low_list else _first_finite(lot_data.low_limits)
    high_limit_max = _first_finite(high_list) if high_list else _first_finite(lot_data.high_limits)

    return IntervalSummary(
        min_label=_format_interval_label(min_label),
        max_label=_format_interval_label(max_label),
        min_mean=min_mean,
        max_mean=max_mean,
        mean_shift=mean_shift,
        low_limit_max=low_limit_max if low_limit_max is not None else low_declared,
        high_limit_max=high_limit_max if high_limit_max is not None else high_declared,
    )





def build_chart_image(test: TestData, lot_data: LotData) -> Optional[BytesIO]:
    if not lot_data.measurements:
        return None
    low_declared, high_declared = lot_data.representative_limits()
    axis_low_raw, axis_high_raw = determine_axis_limits(lot_data.measurements, low_declared, high_declared)
    bin_count = choose_bin_count(lot_data.measurements, axis_low_raw, axis_high_raw)
    bin_count = max(1, bin_count)

    span = axis_high_raw - axis_low_raw
    bin_width = span / bin_count if bin_count else 0.0
    padding = max(bin_width / 2.0, abs(span) * 0.05, 1e-6)
    if not math.isfinite(padding) or padding <= 0:
        padding = 0.5
    axis_low = axis_low_raw - padding
    axis_high = axis_high_raw + padding
    if not math.isfinite(axis_low):
        axis_low = axis_low_raw if math.isfinite(axis_low_raw) else -0.5
    if not math.isfinite(axis_high):
        axis_high = axis_high_raw if math.isfinite(axis_high_raw) else 0.5
    if axis_low >= axis_high:
        axis_low -= 0.5
        axis_high += 0.5

    bin_edges = np.linspace(axis_low, axis_high, bin_count + 1)

    fig, ax = plt.subplots(figsize=(CHART_WIDTH_IN, CHART_HEIGHT_IN))

    ordered_groups, numeric_groups = _prepare_interval_groups(lot_data)
    group_arrays = [values for _, values in ordered_groups]
    cmap = plt.get_cmap("tab20")
    colors = [cmap(index % cmap.N) for index in range(len(group_arrays))]
    group_labels = [_format_interval_label(label) for label, _ in ordered_groups]
    ax.hist(group_arrays, bins=bin_edges, color=colors, edgecolor="#1F497D", alpha=0.7, label=group_labels)

    axis_label = "Measurement"
    if test.unit:
        axis_label += f" ({test.unit})"
    ax.set_xlabel(axis_label)
    ax.set_ylabel("Count")
    ax.set_xlim(axis_low, axis_high)
    ax.set_title(f"{lot_data.lot} - {test.title}")

    interval_summary = _summarize_interval_shift(lot_data, numeric_groups)
    if interval_summary.mean_shift is not None:
        shift_text = f"Mean shift: {interval_summary.mean_shift:.3g}"
        if test.unit:
            shift_text += f" {test.unit}"
        ax.text(0.98, 0.98, shift_text, transform=ax.transAxes, ha="right", va="top", fontsize=12, fontweight="bold", bbox={"boxstyle": "round,pad=0.3", "facecolor": "white", "alpha": 0.7})

    if low_declared is not None:
        ax.axvline(low_declared, color="#C0504D", linestyle="--", linewidth=2, label="Low Limit")
    if high_declared is not None:
        ax.axvline(high_declared, color="#9BBB59", linestyle="--", linewidth=2, label="High Limit")

    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(handles, labels, loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0)

    ax.grid(axis="y", linestyle=":", linewidth=0.6, alpha=0.7)
    fig.tight_layout(rect=[0, 0, 0.78, 1])

    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=150)
    plt.close(fig)
    buffer.seek(0)
    return buffer


def add_charts_to_workbook(
    workbook_path: Path,
    output_path: Path,
    events: "OrderedDict[str, EventData]",
    max_lots: int,
) -> int:
    wb = load_workbook(workbook_path)
    for sheet_name in list(wb.sheetnames):
        if sheet_name.startswith(HISTOGRAM_SHEET_PREFIX) or sheet_name == SHIFT_SUMMARY_SHEET_NAME:
            del wb[sheet_name]

    measurement_indices = [idx for idx, name in enumerate(wb.sheetnames) if name.startswith(MEASUREMENTS_PREFIX)]
    insert_index = max(measurement_indices) + 1 if measurement_indices else len(wb.sheetnames)

    summary_rows: List[List[object]] = []

    for event_name, event_data in events.items():
        sheet_name = _histogram_sheet_name(event_name, wb.sheetnames)
        charts_ws = wb.create_sheet(sheet_name, index=insert_index)
        insert_index = wb.sheetnames.index(sheet_name) + 1
        charts_ws["A1"] = (
            f"Histogram charts generated by {Path(__file__).name} for event '{event_name}' ({sheet_name})"
        )
        charts_ws.freeze_panes = "B4"

        sorted_tests = [event_data.tests[key] for key in event_data.tests]

        lot_order: "OrderedDict[str, None]" = OrderedDict()
        for test in sorted_tests:
            for lot_name in test.lots:
                lot_order.setdefault(lot_name, None)
        lot_names = list(lot_order.keys())
        if max_lots > 0:
            lot_names = lot_names[:max_lots]

        for col_idx, lot_name in enumerate(lot_names):
            header_column = FIRST_CHART_COLUMN + col_idx * COL_STRIDE
            charts_ws.cell(row=HEADER_ROW, column=header_column, value=lot_name)

        for test_idx, test in enumerate(sorted_tests):
            anchor_row = HEADER_ROW + 1 + test_idx * ROW_STRIDE
            charts_ws.cell(row=anchor_row, column=TEST_LABEL_COLUMN, value=test.title)

            for col_idx, lot_name in enumerate(lot_names):
                lot_data = test.lots.get(lot_name)
                if lot_data is None or not lot_data.measurements:
                    continue
                interval_summary = _summarize_interval_shift(lot_data)
                summary_rows.append([
                    event_name,
                    test.name,
                    test.number or "",
                    test.unit or "",
                    lot_name,
                    interval_summary.min_label or "",
                    interval_summary.max_label or "",
                    interval_summary.min_mean,
                    interval_summary.max_mean,
                    interval_summary.mean_shift,
                    interval_summary.low_limit_max,
                    interval_summary.high_limit_max,
                ])
                image_stream = build_chart_image(test, lot_data)
                if image_stream is None:
                    continue
                img = XLImage(image_stream)
                img.width = CHART_WIDTH_PX
                img.height = CHART_HEIGHT_PX

                anchor_col = FIRST_CHART_COLUMN + col_idx * COL_STRIDE
                anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"
                charts_ws.add_image(img, anchor_cell)

    summary_index = insert_index
    summary_ws = wb.create_sheet(SHIFT_SUMMARY_SHEET_NAME, index=summary_index)
    summary_ws.append(SHIFT_SUMMARY_HEADER)
    for row in summary_rows:
        summary_ws.append(row)
    summary_ws.freeze_panes = "A2"

    wb.save(output_path)
    wb.close()
    return len(summary_rows)



def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    events = load_measurement_tests(workbook_path)
    output_path = (args.output or workbook_path).expanduser().resolve()
    summary_count = add_charts_to_workbook(workbook_path, output_path, events, max_lots=max(args.max_lots, 0))
    total_histograms = sum(len(test.lots) for event in events.values() for test in event.tests.values())
    total_tests = sum(len(event.tests) for event in events.values())
    print(
        f"Added histogram sheets for {len(events)} event(s) with {total_histograms} histogram(s) across "
        f"{total_tests} test(s) and recorded {summary_count} ShiftSummary row(s) in {output_path}"
    )



if __name__ == "__main__":
    main()

