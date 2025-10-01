from __future__ import annotations

import statistics
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..models import MeasurementRow


@dataclass
class ShiftRow:
    event: str
    test_name: str
    test_number: str
    unit: str
    lot: str
    min_interval: str
    max_interval: str
    min_interval_mean: float | None
    max_interval_mean: float | None
    mean_shift: float | None
    stddev_change: float | None
    min_interval_median: float | None
    max_interval_median: float | None
    low_limit_max: float | None
    high_limit_max: float | None


def run(
    measurements: Sequence[MeasurementRow],
    workbook_path: Path,
) -> list[ShiftRow]:
    shift_rows = _build_shift_rows(measurements)
    if not shift_rows:
        return []

    workbook_path = workbook_path.expanduser().resolve()
    workbook = load_workbook(workbook_path)
    if 'ShiftSummary' in workbook.sheetnames:
        del workbook['ShiftSummary']
    sheet = workbook.create_sheet('ShiftSummary')

    headers = [
        'Event',
        'Test Name',
        'Test Number',
        'Unit',
        'Lot',
        'Min Interval',
        'Max Interval',
        'Min Interval Mean',
        'Max Interval Mean',
        'Mean Shift',
        'Std Dev Change',
        'Low Limit (Max Int)',
        'High Limit (Max Int)',
    ]
    sheet.append(headers)
    for row in shift_rows:
        sheet.append([
            row.event,
            row.test_name,
            row.test_number,
            row.unit,
            row.lot,
            row.min_interval,
            row.max_interval,
            row.min_interval_mean,
            row.max_interval_mean,
            row.mean_shift,
            row.stddev_change,
            row.low_limit_max,
            row.high_limit_max,
        ])

    last_row = len(shift_rows) + 1
    table = Table(displayName='ShiftSummaryTable', ref=f'A1:M{last_row}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = 'A2'

    widths = (18, 42, 18, 12, 14, 18, 18, 20, 20, 16, 16, 18, 18)
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    workbook.save(workbook_path)
    return shift_rows


def _build_shift_rows(measurements: Sequence[MeasurementRow]) -> list[ShiftRow]:
    if not measurements:
        return []

    group_map: Dict[Tuple[str, str, str, str, str], Dict[str, List[float]]] = defaultdict(lambda: defaultdict(list))
    limit_map: Dict[Tuple[str, str, str, str, str], Dict[str, Tuple[float | None, float | None]]] = defaultdict(dict)

    for row in measurements:
        value = _coerce_numeric(row.measurement)
        if value is None:
            continue
        key = (row.event or '', row.test_name or '', row.test_number or '', row.test_unit or '', row.lot or '')
        interval = row.interval or ''
        group_map[key][interval].append(value)
        limits = limit_map[key].get(interval, (None, None))
        low, high = limits
        new_low = low if low is not None else row.low_limit
        new_high = high if high is not None else row.high_limit
        limit_map[key][interval] = (new_low, new_high)

    shift_rows: list[ShiftRow] = []
    for key, interval_map in group_map.items():
        if not interval_map:
            continue
        ordered_intervals = _order_intervals(interval_map.keys())
        if not ordered_intervals:
            continue
        min_interval = ordered_intervals[0]
        max_interval = ordered_intervals[-1]
        min_values = interval_map.get(min_interval, [])
        max_values = interval_map.get(max_interval, [])
        if not min_values or not max_values:
            continue

        min_mean = statistics.fmean(min_values)
        max_mean = statistics.fmean(max_values)
        mean_shift = max_mean - min_mean
        min_std = statistics.pstdev(min_values) if len(min_values) > 1 else 0.0
        max_std = statistics.pstdev(max_values) if len(max_values) > 1 else 0.0
        std_change = max_std - min_std
        min_median = statistics.median(min_values)
        max_median = statistics.median(max_values)
        low_limit, high_limit = limit_map[key].get(max_interval, (None, None))

        shift_rows.append(
            ShiftRow(
                event=key[0],
                test_name=key[1],
                test_number=key[2],
                unit=key[3],
                lot=key[4],
                min_interval=min_interval,
                max_interval=max_interval,
                min_interval_mean=min_mean,
                max_interval_mean=max_mean,
                mean_shift=mean_shift,
                stddev_change=std_change,
                min_interval_median=min_median,
                max_interval_median=max_median,
                low_limit_max=low_limit,
                high_limit_max=high_limit,
            )
        )

    shift_rows.sort(
        key=lambda row: (
            row.event.lower(),
            row.test_name.lower(),
            row.test_number,
            row.lot.lower(),
            _interval_sort_key(row.min_interval),
            _interval_sort_key(row.max_interval),
        )
    )
    return shift_rows


def _coerce_numeric(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def _order_intervals(intervals: Iterable[str]) -> List[str]:
    unique = {interval for interval in intervals if interval is not None}
    return sorted(unique, key=_interval_sort_key)


def _interval_sort_key(label: str) -> Tuple[int, float, str]:
    if label is None:
        return (3, 0.0, '')
    text = label.strip()
    if not text:
        return (3, 0.0, '')
    lowered = text.lower()
    if lowered == 'prescreen':
        return (0, 0.0, lowered)
    try:
        numeric = float(text)
        return (1, numeric, lowered)
    except ValueError:
        return (2, 0.0, lowered)

