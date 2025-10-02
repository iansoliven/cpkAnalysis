"""Convert source measurement files into a consolidated DataWorkBook."""

from __future__ import annotations

import math
import sys
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ..models import MeasurementRow, SourceFile, SummaryRow

ISTDF_SRC = Path(__file__).resolve().parents[2] / "Submodules" / "istdf" / "src"
if ISTDF_SRC.exists() and ISTDF_SRC.is_dir():
    if str(ISTDF_SRC) not in sys.path:
        sys.path.insert(0, str(ISTDF_SRC))
try:  # pragma: no cover - dependency injection
    from istdf import STDFReader  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    STDFReader = None  # type: ignore[assignment]

_PART_FAIL_MASK = 0x08


@dataclass
class _ExcelMeasurement:
    lot: str
    event: str
    interval: str
    source_sheet: str
    sn: str
    status: str
    test_number: str
    test_name: str
    test_unit: str
    low_limit: Optional[float]
    high_limit: Optional[float]
    measurement: Any
    row_index: int


@dataclass
class _TestMetadata:
    test_name: str = ''
    unit: str = ''
    low_limit: Optional[float] = None
    high_limit: Optional[float] = None
    scale: Optional[int] = None

def run(
    sources: Sequence[SourceFile],
    output_path: Path,
    values_only: bool = False,
) -> tuple[list[SummaryRow], list[MeasurementRow]]:
    if not sources:
        raise ValueError("No sources were provided to ConvertToData stage.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_rows: list[SummaryRow] = []
    measurement_rows: list[MeasurementRow] = []

    for source in sources:
        if source.file_type == "xlsx":
            summary, measurements = _process_excel_source(source, values_only)
        elif source.file_type == "stdf":
            summary, measurements = _process_stdf_source(source)
        else:
            raise ValueError(f"Unsupported file type: {source.file_type}")
        if summary is not None:
            summary_rows.append(summary)
        measurement_rows.extend(measurements)

    _create_summary_sheet(workbook, summary_rows)
    _create_measurements_sheet(workbook, measurement_rows)

    resolved_output = output_path.expanduser().resolve()
    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(resolved_output)
    return summary_rows, measurement_rows


def _process_excel_source(
    source: SourceFile,
    values_only: bool,
) -> tuple[SummaryRow, list[MeasurementRow]]:
    try:
        workbook = load_workbook(source.path, data_only=values_only, read_only=False)
    except Exception as exc:  # pragma: no cover - surface open errors
        raise RuntimeError(f"Failed to open Excel workbook: {source.path}") from exc

    total_pass = 0
    total_fail = 0
    excel_measurements: list[_ExcelMeasurement] = []

    try:
        for sheet in workbook.worksheets:
            pass_count, fail_count = analyze_unit_results(sheet)
            total_pass += pass_count
            total_fail += fail_count
            excel_measurements.extend(
                extract_measurements(
                    sheet,
                    lot=source.lot,
                    event=source.event,
                    interval=source.interval,
                    source_sheet=sheet.title,
                )
            )
    finally:
        workbook.close()

    collapsed = collapse_measurement_records(excel_measurements)
    measurements = [
        MeasurementRow(
            lot=record.lot,
            event=record.event,
            interval=record.interval,
            source=source.path.name,
            status=record.status,
            test_number=record.test_number,
            test_name=record.test_name,
            test_unit=record.test_unit,
            low_limit=record.low_limit,
            high_limit=record.high_limit,
            measurement=record.measurement,
            serial_number=record.sn,
        )
        for record in collapsed
    ]

    summary = SummaryRow(
        lot=source.lot,
        event=source.event,
        interval=source.interval,
        file_name=source.path.name,
        source_path=source.path,
        pass_count=total_pass,
        fail_count=total_fail,
    )
    return summary, measurements


def _process_stdf_source(source: SourceFile) -> tuple[SummaryRow, list[MeasurementRow]]:
    if STDFReader is None:
        raise RuntimeError(
            "STDF support is unavailable. Ensure Submodules/istdf is initialised and importable."
        )

    attempts: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    fallback_index = 0
    current_measurements: list[dict[str, Any]] = []
    current_serial: Optional[str] = None
    current_site: Optional[int] = None
    test_metadata: Dict[str, _TestMetadata] = {}

    def store_attempt(serial: Optional[str], status: str) -> None:
        nonlocal fallback_index
        if serial is None or not str(serial).strip():
            serial_key = f"UNLABELED_{fallback_index}"
            fallback_index += 1
        else:
            serial_key = str(serial)
        attempt = {
            "status": status,
            "measurements": list(current_measurements),
        }
        if serial_key in attempts:
            del attempts[serial_key]
        attempts[serial_key] = attempt
        current_measurements.clear()

    with STDFReader(source.path, ignore_unknown=True) as reader:
        for record in reader:
            name = record.name
            data = record.to_dict()
            if name == "PIR":
                current_measurements.clear()
                current_serial = data.get("PART_ID")
                current_site = data.get("SITE_NUM")
            elif name == "PTR":
                measurement = _extract_stdf_measurement(data, test_metadata)
                if measurement is not None:
                    current_measurements.append(measurement)
            elif name == "PRR":
                part_flag = data.get("PART_FLG")
                status = "FAIL" if _is_part_fail(part_flag) else "PASS"
                part_id = data.get("PART_ID") or current_serial
                if part_id is None and current_site is not None:
                    part_id = f"SITE{current_site}_{fallback_index}"
                store_attempt(part_id, status)
                current_serial = None
                current_site = None

    pass_count = 0
    fail_count = 0
    measurements: list[MeasurementRow] = []
    for serial, attempt in attempts.items():
        status = attempt["status"]
        if status == "PASS":
            pass_count += 1
        else:
            fail_count += 1
        for measurement in attempt["measurements"]:
            measurements.append(
                MeasurementRow(
                    lot=source.lot,
                    event=source.event,
                    interval=source.interval,
                    source=source.path.name,
                    status=status,
                    test_number=measurement["test_number"],
                    test_name=measurement["test_name"],
                    test_unit=measurement["test_unit"],
                    low_limit=measurement["low_limit"],
                    high_limit=measurement["high_limit"],
                    measurement=measurement["measurement"],
                    serial_number=serial,
                )
            )

    summary = SummaryRow(
        lot=source.lot,
        event=source.event,
        interval=source.interval,
        file_name=source.path.name,
        source_path=source.path,
        pass_count=pass_count,
        fail_count=fail_count,
    )
    return summary, measurements


def _extract_stdf_measurement(data: Dict[str, Any], cache: Dict[str, _TestMetadata]) -> Optional[dict[str, Any]]:
    test_num = data.get("TEST_NUM")
    test_name_raw = data.get("TEST_TXT")
    if test_num is None and not test_name_raw:
        return None

    key = str(test_num) if test_num is not None else f"name:{test_name_raw or ''}"
    entry = cache.get(key, _TestMetadata())

    test_name = str(test_name_raw).strip() if test_name_raw else entry.test_name
    if test_name:
        entry.test_name = test_name

    unit_raw = data.get("UNITS")
    unit_value = str(unit_raw).strip() if unit_raw else entry.unit
    if unit_value:
        entry.unit = unit_value

    res_scale = data.get("RES_SCAL")
    measurement_value = _apply_scale(data.get("RESULT"), res_scale)

    llm_scale = data.get("LLM_SCAL")
    low_limit = _apply_scale(data.get("LO_LIMIT"), llm_scale)
    if low_limit is not None:
        entry.low_limit = low_limit
    else:
        low_limit = entry.low_limit

    hlm_scale = data.get("HLM_SCAL")
    high_limit = _apply_scale(data.get("HI_LIMIT"), hlm_scale)
    if high_limit is not None:
        entry.high_limit = high_limit
    else:
        high_limit = entry.high_limit

    for scale_candidate in (res_scale, llm_scale, hlm_scale):
        if scale_candidate is not None:
            entry.scale = scale_candidate
            break

    cache[key] = entry

    return {
        "test_number": str(test_num) if test_num is not None else "",
        "test_name": entry.test_name,
        "test_unit": _compose_unit(entry.unit, entry.scale),
        "low_limit": low_limit,
        "high_limit": high_limit,
        "measurement": measurement_value,
    }


def _apply_scale(value: Optional[float], scale: Optional[int]) -> Optional[float]:
    if value is None:
        return None
    if scale in (None, 0):
        return value
    try:
        return value * (10 ** scale)
    except Exception:
        return value



_PREFIX_BY_EXPONENT = {
    -24: "y",
    -21: "z",
    -18: "a",
    -15: "f",
    -12: "p",
    -9: "n",
    -6: "u",
    -3: "m",
    0: "",
    3: "k",
    6: "M",
    9: "G",
    12: "T",
    15: "P",
    18: "E",
    21: "Z",
    24: "Y",
}


def _compose_unit(base_unit: str, scale: Optional[int]) -> str:
    unit = (base_unit or "").strip()
    if not unit:
        if scale in (2, -2):
            return "%"
        return ""
    if scale in (None, 0):
        return unit
    prefix = _PREFIX_BY_EXPONENT.get(-scale)
    if prefix is None:
        return unit
    return f"{prefix}{unit}"

def _is_part_fail(flag: Any) -> bool:
    if flag is None:
        return False
    if isinstance(flag, bytes):
        if not flag:
            return False
        numeric = flag[0]
    else:
        try:
            numeric = int(flag)
        except Exception:
            return False
    return (numeric & _PART_FAIL_MASK) != 0


# --- Excel helpers from legacy merge pipeline ---

EXCEL_SHEET_NAME_LIMIT = 31
INVALID_SHEET_CHARS = set('[]:*?/\\')


def analyze_unit_results(sheet: Worksheet) -> Tuple[int, int]:
    header_row_index: Optional[int] = None
    header_values: List[str] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        for cell_value in row:
            if isinstance(cell_value, str) and cell_value.strip().upper() == 'UNIT SN':
                header_row_index = idx
                header_values = [
                    value.strip() if isinstance(value, str) else ''
                    for value in row
                ]
                break
        if header_row_index is not None:
            break
    if header_row_index is None:
        return 0, 0
    try:
        sn_col = next(
            index for index, value in enumerate(header_values)
            if isinstance(value, str) and value.strip().upper() == 'UNIT SN'
        )
        status_col = next(
            index for index, value in enumerate(header_values)
            if isinstance(value, str) and value.strip().upper() == 'PASS/FAIL'
        )
    except StopIteration:
        return 0, 0
    sn_values: Dict[str, str] = {}
    for row in sheet.iter_rows(
        min_row=header_row_index + 1,
        values_only=True,
    ):
        if not row:
            continue
        serial = row[sn_col]
        status = row[status_col]
        if serial is None:
            continue
        serial_text = str(serial).strip()
        if not serial_text:
            continue
        outcome = _normalize_outcome(status, 'FAIL')
        sn_values[serial_text] = outcome
    pass_count = sum(1 for outcome in sn_values.values() if outcome == 'PASS')
    fail_count = sum(1 for outcome in sn_values.values() if outcome == 'FAIL')
    return pass_count, fail_count


def _get_cell(row: Optional[Sequence[Any]], index: int) -> Any:
    if row is None:
        return None
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return None
        except TypeError:
            pass
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        result = float(text)
    except ValueError:
        return None
    try:
        if math.isnan(result):
            return None
    except TypeError:
        pass
    return result


def _normalize_measurement(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        sentinel_tokens = {'N/A', 'NA', 'NULL', 'NONE', 'F'}
        upper = stripped.upper()
        if upper in sentinel_tokens:
            return None
        if stripped.upper().startswith('F-'):
            candidate = stripped[1:]
        else:
            candidate = stripped
        coerced = _coerce_float(candidate)
        if coerced is not None:
            return coerced
        return stripped
    return value


def _clean_test_name(raw: Any) -> str:
    if raw is None:
        return ''
    text = str(raw).strip()
    if not text:
        return ''
    if text.startswith('"') and text.endswith('"') and len(text) > 1:
        text = text[1:-1]
    if '",' in text:
        text = text.split('",', 1)[0]
    return text


def _normalize_outcome(value: Any, default: str) -> str:
    default_upper = default.strip().upper() if isinstance(default, str) else 'FAIL'
    if not default_upper:
        default_upper = 'FAIL'
    if value is None:
        return default_upper
    if isinstance(value, str):
        upper = value.strip().upper()
    else:
        upper = str(value).strip().upper()
    if upper in {'P', 'PASS'}:
        return 'PASS'
    if upper in {'F', 'FAIL'}:
        return 'FAIL'
    return default_upper


def _is_blank_row(row: Sequence[Any]) -> bool:
    if not row:
        return True
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str):
            if cell.strip():
                return False
        else:
            return False
    return True


def _build_test_columns(header_row: Sequence[Any], metadata: Dict[str, Sequence[Any]], status_col: int) -> List[Dict[str, Any]]:
    number_row = metadata.get('number')
    name_row = metadata.get('name')
    unit_row = metadata.get('unit')
    low_row = metadata.get('low')
    high_row = metadata.get('high')
    if not number_row or not name_row or not unit_row:
        return []

    start_index: Optional[int] = None
    for idx, value in enumerate(header_row):
        if isinstance(value, str) and value.strip().upper() == 'RETEST ANALYSIS':
            start_index = idx + 1
            break
    if start_index is None:
        start_index = status_col + 1

    max_len = max(len(number_row), len(name_row), len(unit_row), len(low_row or []), len(high_row or []))
    columns: List[Dict[str, Any]] = []
    for idx in range(start_index, max_len):
        test_number = _get_cell(number_row, idx)
        test_name = _get_cell(name_row, idx)
        test_unit = _get_cell(unit_row, idx)
        if test_number is None or test_name is None or test_unit is None:
            continue
        number_text = str(test_number).strip()
        if not number_text:
            continue
        name_text = _clean_test_name(test_name)
        if not name_text:
            continue
        unit_text = str(test_unit).strip()
        if not unit_text:
            continue
        columns.append(
            {
                'index': idx,
                'test_number': number_text,
                'test_name': name_text,
                'test_unit': unit_text,
                'low_limit': _coerce_float(_get_cell(low_row, idx)),
                'high_limit': _coerce_float(_get_cell(high_row, idx)),
            }
        )
    return columns


def extract_measurements(
    sheet: Worksheet,
    lot: str,
    event: str,
    interval: str,
    source_sheet: str,
) -> List[_ExcelMeasurement]:
    metadata_rows: Dict[str, Sequence[Any]] = {}
    header_row: Optional[Sequence[Any]] = None
    header_row_index: Optional[int] = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        first_cell = row[0]
        if not isinstance(first_cell, str):
            continue
        label = first_cell.strip().upper()
        if label == 'TEST NUMBER':
            metadata_rows['number'] = list(row)
        elif label == 'TEST NAME':
            metadata_rows['name'] = list(row)
        elif label == 'TEST UNIT':
            metadata_rows['unit'] = list(row)
        elif label == 'LOW LIMIT':
            metadata_rows['low'] = list(row)
        elif label == 'HIGH LIMIT':
            metadata_rows['high'] = list(row)
        elif label == 'UNIT SN':
            header_row = list(row)
            header_row_index = idx
            break
    if header_row is None or header_row_index is None:
        return []

    try:
        sn_col = next(
            index for index, value in enumerate(header_row)
            if isinstance(value, str) and value.strip().upper() == 'UNIT SN'
        )
        status_col = next(
            index for index, value in enumerate(header_row)
            if isinstance(value, str) and value.strip().upper() == 'PASS/FAIL'
        )
    except StopIteration:
        return []

    columns = _build_test_columns(header_row, metadata_rows, status_col)
    if not columns:
        return []

    records: List[_ExcelMeasurement] = []
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=header_row_index + 1, values_only=True),
        start=header_row_index + 1,
    ):
        if not row or _is_blank_row(row):
            continue
        serial = row[sn_col]
        if serial is None:
            continue
        serial_text = str(serial).strip()
        if not serial_text:
            continue
        outcome = _normalize_outcome(_get_cell(row, status_col), 'FAIL')
        for column in columns:
            measurement_value = _normalize_measurement(_get_cell(row, column['index']))
            if measurement_value is None:
                continue
            records.append(
                _ExcelMeasurement(
                    lot=lot,
                    event=event,
                    interval=interval,
                    source_sheet=source_sheet,
                    sn=serial_text,
                    status=outcome,
                    test_number=column['test_number'],
                    test_name=column['test_name'],
                    test_unit=column['test_unit'],
                    low_limit=column['low_limit'],
                    high_limit=column['high_limit'],
                    measurement=measurement_value,
                    row_index=row_index,
                )
            )
    return records


def collapse_measurement_records(records: Sequence[_ExcelMeasurement]) -> List[_ExcelMeasurement]:
    if not records:
        return []
    collapsed: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}
    order: List[Tuple[str, str, str, str]] = []
    for record in records:
        key = (record.lot, record.event, record.interval, record.sn)
        bucket = collapsed.get(key)
        if bucket is None:
            bucket = {
                'status': record.status,
                'source_sheet': record.source_sheet,
                'row_index': record.row_index,
                'records': [],
            }
            collapsed[key] = bucket
            order.append(key)
        else:
            if record.status == 'PASS':
                if (
                    bucket['status'] != 'PASS'
                    or bucket['row_index'] != record.row_index
                    or bucket['source_sheet'] != record.source_sheet
                ):
                    bucket['status'] = 'PASS'
                    bucket['source_sheet'] = record.source_sheet
                    bucket['row_index'] = record.row_index
                    bucket['records'] = []
            else:
                if bucket['status'] == 'PASS':
                    if (
                        bucket['source_sheet'] == record.source_sheet
                        and bucket['row_index'] == record.row_index
                    ):
                        pass
                    else:
                        continue
                if (
                    bucket['status'] != 'FAIL'
                    or bucket['row_index'] != record.row_index
                    or bucket['source_sheet'] != record.source_sheet
                ):
                    bucket['status'] = 'FAIL'
                    bucket['source_sheet'] = record.source_sheet
                    bucket['row_index'] = record.row_index
                    bucket['records'] = []
        bucket['records'].append(record)
    result: List[_ExcelMeasurement] = []
    for key in order:
        bucket = collapsed[key]
        status = bucket['status']
        for record in bucket['records']:
            record.status = status
        result.extend(bucket['records'])
    return result


def _create_summary_sheet(workbook: Workbook, entries: Sequence[SummaryRow]) -> None:
    if not entries:
        return
    summary_ws = workbook.create_sheet(title='Summary', index=0)
    headers = ['Source', 'Lot', 'Event', 'Interval', 'Passing Units', 'Failing Units']
    summary_ws.append(headers)
    for entry in entries:
        summary_ws.append([
            entry.file_name,
            entry.lot,
            entry.event,
            entry.interval,
            entry.pass_count,
            entry.fail_count,
        ])
    last_row = len(entries) + 1
    table = Table(displayName='SummaryTable', ref=f'A1:F{last_row}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showRowStripes=True,
        showColumnStripes=False,
    )
    summary_ws.add_table(table)
    summary_ws.freeze_panes = 'A2'
    for index, width in enumerate((36, 14, 18, 16, 18, 18), start=1):
        summary_ws.column_dimensions[get_column_letter(index)].width = width


def _create_measurements_sheet(workbook: Workbook, records: Sequence[MeasurementRow]) -> None:
    if not records:
        return
    max_rows = 1048576
    data_capacity = max_rows - 1
    base_index = 1 if 'Summary' in workbook.sheetnames else 0
    headers = [
        'Lot',
        'Event',
        'Interval',
        'Source',
        'Serial Number',
        'PASS/FAIL',
        'Test Number',
        'Test Name',
        'Test Unit',
        'Low Limit',
        'High Limit',
        'Measurement',
    ]
    widths = [16, 18, 12, 32, 18, 12, 18, 42, 12, 14, 14, 16]
    total = len(records)
    chunks = (total + data_capacity - 1) // data_capacity
    for chunk_idx in range(chunks):
        start_row = chunk_idx * data_capacity
        end_row = min(start_row + data_capacity, total)
        chunk = records[start_row:end_row]
        if not chunk:
            continue
        suffix = '' if chunk_idx == 0 else f'_{chunk_idx + 1}'
        sheet_title = 'Measurements' if chunk_idx == 0 else f'Measurements{suffix}'
        worksheet_index = base_index + chunk_idx
        detail_ws = workbook.create_sheet(title=sheet_title, index=worksheet_index)
        detail_ws.append(headers)
        for record in chunk:
            detail_ws.append([
                record.lot,
                record.event,
                record.interval,
                record.source,
                record.serial_number,
                record.status,
                record.test_number,
                record.test_name,
                record.test_unit,
                record.low_limit,
                record.high_limit,
                record.measurement,
            ])
        last_row = len(chunk) + 1
        table_name = 'MeasurementTable' if chunk_idx == 0 else f'MeasurementTable{chunk_idx + 1}'
        last_col_letter = get_column_letter(len(headers))
        table = Table(displayName=table_name, ref=f'A1:{last_col_letter}{last_row}')
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium9',
            showRowStripes=True,
            showColumnStripes=False,
        )
        detail_ws.add_table(table)
        detail_ws.freeze_panes = 'A2'
        for idx, width in enumerate(widths, start=1):
            detail_ws.column_dimensions[get_column_letter(idx)].width = width
        for row_idx in range(2, last_row + 1):
            detail_ws.cell(row=row_idx, column=10).number_format = numbers.FORMAT_GENERAL
            detail_ws.cell(row=row_idx, column=11).number_format = numbers.FORMAT_GENERAL
            detail_ws.cell(row=row_idx, column=12).number_format = numbers.FORMAT_GENERAL
